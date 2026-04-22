#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sistema de Vendas - Moto Pecas & Mecanica v3.0"""

import sys, os, json, uuid
from datetime import datetime
from collections import defaultdict

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QComboBox, QSpinBox,
    QDoubleSpinBox, QTextEdit, QMessageBox, QFileDialog, QHeaderView,
    QFrame, QDateEdit, QGroupBox, QRadioButton, QSplitter,
    QAbstractItemView, QStatusBar, QStackedWidget, QCheckBox, QColorDialog,
    QScrollArea, QSizePolicy, QInputDialog, QTabWidget, QShortcut, QGraphicsDropShadowEffect
)
import urllib.parse, webbrowser
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPixmap, QBrush, QPalette, QIcon, QKeySequence

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.lib.pagesizes import A6
    from reportlab.pdfgen import canvas as rl_canvas
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import qrcode
    import qrcode.image.pil
    QR_OK = True
except ImportError:
    QR_OK = False

import xml.etree.ElementTree as ET
try:
    import requests as _requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False
import hashlib
import random
import re as _re
import unicodedata as _ud

# ─── PATHS ────────────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE   = os.path.join(BASE_DIR, "autopecas_data.json")
FALTAS_FILE = os.path.join(BASE_DIR, "faltas.json")
TROCAS_FILE = os.path.join(BASE_DIR, "trocas.json")
CAIXA_FILE  = os.path.join(BASE_DIR, "caixa.json")
NPS_FILE    = os.path.join(BASE_DIR, "nps.json")

# ─── PORTUGUESE WEEKDAYS ──────────────────────────────────────────────────────
DIAS_PT = {
    "Monday": "Segunda-feira", "Tuesday": "Terca-feira",
    "Wednesday": "Quarta-feira", "Thursday": "Quinta-feira",
    "Friday": "Sexta-feira", "Saturday": "Sabado", "Sunday": "Domingo",
}

PRESET_COLORS = [
    ("#FF6B35", "Laranja (Padrao)"),
    ("#E53935", "Vermelho"),
    ("#1E88E5", "Azul"),
    ("#43A047", "Verde"),
    ("#8E24AA", "Roxo"),
    ("#FFB300", "Ambar"),
    ("#00ACC1", "Ciano"),
    ("#F06292", "Rosa"),
    ("#5D4037", "Marrom"),
    ("#546E7A", "Ardosia"),
]

CATEGORIES = ["Filtros","Freios","Oleos","Motor","Eletrica","Suspensao",
               "Arrefecimento","Transmissao","Carroceria","Acessorios","Outros"]
UNITS = ["UN","JG","LT","KG","MT","CX","PC","FR","KIT","PAR"]
PAYMENTS = ["Dinheiro","Cartao de Credito","Cartao de Debito",
            "PIX","Boleto","Transferencia","Fiado / A Prazo"]



# ═════════════════════════════════════════════════════════════════════════════
#  HELPERS FISCAIS
# ═════════════════════════════════════════════════════════════════════════════

def _clean(s):
    """Remove accents and special chars, keep alphanumeric spaces."""
    n = _ud.normalize("NFKD", str(s)); return "".join(c for c in n if not _ud.combining(c))

def _digits(s):
    """Extract only digits from string."""
    return _re.sub(r"\D", "", str(s))

def _calc_dv_nf(chave43):
    """Calculate NF-e check digit (mod11)."""
    seq = [2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4]
    soma = sum(int(chave43[i]) * seq[i] for i in range(43))
    resto = soma % 11
    return 0 if resto in (0, 1) else 11 - resto

def _gerar_chave_nfce(cuf, aamm, cnpj, mod, serie, nnf, c_nf):
    """Generate the 44-digit NFC-e access key."""
    cnpj_d = _digits(cnpj).zfill(14)
    serie_d = str(serie).zfill(3)
    nnf_d = str(nnf).zfill(9)
    c_nf_d = str(c_nf).zfill(8)
    chave43 = f"{cuf}{aamm}{cnpj_d}{mod}{serie_d}{nnf_d}1{c_nf_d}"
    dv = _calc_dv_nf(chave43)
    return chave43 + str(dv)

def _formatar_chave(chave44):
    """Format 44-digit key into groups of 4."""
    return " ".join(chave44[i:i+4] for i in range(0, 44, 4))

def _gerar_xml_nfce(empresa, nfce_cfg, sale, nf_num, chave44):
    """Generate NFC-e XML in the correct Brazilian format."""
    cnpj = _digits(empresa.get("cnpj",""))
    ie   = _digits(empresa.get("ie",""))
    cuf  = nfce_cfg.get("cuf", "35")
    serie = str(nfce_cfg.get("serie","001")).zfill(3)
    amb  = nfce_cfg.get("ambiente", "2")   # 1=producao 2=homologacao
    now  = datetime.now()
    dh_emi = now.strftime("%Y-%m-%dT%H:%M:%S") + "-03:00"
    aamm = now.strftime("%y%m")

    root = ET.Element("nfeProc", xmlns="http://www.portalfiscal.inf.br/nfe", versao="4.00")
    nfe  = ET.SubElement(root, "NFe", xmlns="http://www.portalfiscal.inf.br/nfe")
    inf  = ET.SubElement(nfe, "infNFe", Id=f"NFe{chave44}", versao="4.00")

    # --- ide ---
    ide = ET.SubElement(inf, "ide")
    ET.SubElement(ide, "cUF").text = cuf
    ET.SubElement(ide, "cNF").text = chave44[35:43]
    ET.SubElement(ide, "natOp").text = "VENDA AO CONSUMIDOR"
    ET.SubElement(ide, "mod").text = "65"
    ET.SubElement(ide, "serie").text = serie.lstrip("0") or "1"
    ET.SubElement(ide, "nNF").text = str(nf_num)
    ET.SubElement(ide, "dhEmi").text = dh_emi
    ET.SubElement(ide, "tpNF").text = "1"
    ET.SubElement(ide, "idDest").text = "1"
    ET.SubElement(ide, "cMunFG").text = nfce_cfg.get("cmun","3550308")
    ET.SubElement(ide, "tpImp").text = "4"   # 4=DANFE NFC-e
    ET.SubElement(ide, "tpEmis").text = "1"
    ET.SubElement(ide, "cDV").text = chave44[-1]
    ET.SubElement(ide, "tpAmb").text = amb
    ET.SubElement(ide, "finNFe").text = "1"
    ET.SubElement(ide, "indFinal").text = "1"
    ET.SubElement(ide, "indPres").text = "1"
    ET.SubElement(ide, "indIntermed").text = "0"
    ET.SubElement(ide, "procEmi").text = "0"
    ET.SubElement(ide, "verProc").text = "MotoPecas-3.0"

    # --- emit ---
    emit = ET.SubElement(inf, "emit")
    ET.SubElement(emit, "CNPJ").text = cnpj
    ET.SubElement(emit, "xNome").text = _clean(empresa.get("razao_social",""))[:60]
    ET.SubElement(emit, "xFant").text = _clean(empresa.get("nome_fantasia",""))[:60]
    end_e = ET.SubElement(emit, "enderEmit")
    ET.SubElement(end_e, "xLgr").text  = _clean(empresa.get("endereco",""))[:60]
    ET.SubElement(end_e, "xBairro").text = _clean(empresa.get("bairro",""))[:60]
    ET.SubElement(end_e, "cMun").text  = nfce_cfg.get("cmun","3550308")
    ET.SubElement(end_e, "xMun").text  = _clean(empresa.get("municipio",""))[:60]
    ET.SubElement(end_e, "UF").text    = empresa.get("uf","SP")
    ET.SubElement(end_e, "CEP").text   = _digits(empresa.get("cep",""))
    ET.SubElement(end_e, "cPais").text = "1058"
    ET.SubElement(end_e, "xPais").text = "Brasil"
    if empresa.get("telefone"):
        ET.SubElement(end_e, "fone").text = _digits(empresa.get("telefone",""))
    ET.SubElement(emit, "CRT").text = empresa.get("regime","1")

    # --- dest (consumidor final sem identificacao) ---
    dest = ET.SubElement(inf, "dest")
    ET.SubElement(dest, "indIEDest").text = "9"  # nao contribuinte

    # --- det (items) ---
    total_bc_icms = 0.0; total_icms = 0.0
    total_pis = 0.0; total_cofins = 0.0
    total_prod = 0.0; total_nf = 0.0

    for idx, it in enumerate(sale.get("items",[]), 1):
        det = ET.SubElement(inf, "det", nItem=str(idx))
        prod_el = ET.SubElement(det, "prod")
        ET.SubElement(prod_el, "cProd").text = it.get("code","")
        ET.SubElement(prod_el, "cEAN").text = "SEM GTIN"
        ET.SubElement(prod_el, "xProd").text = _clean(it.get("name",""))[:120]
        ET.SubElement(prod_el, "NCM").text = _digits(it.get("ncm","00000000")).zfill(8)
        ET.SubElement(prod_el, "CFOP").text = it.get("cfop","5102")
        ET.SubElement(prod_el, "uCom").text = it.get("unit","UN")
        qcom = float(it.get("quantity",1))
        vuncom = float(it.get("unit_price",0))
        disc_f = 1 - float(it.get("discount",0))/100
        vtotal = round(qcom * vuncom * disc_f, 2)
        ET.SubElement(prod_el, "qCom").text = f"{qcom:.4f}"
        ET.SubElement(prod_el, "vUnCom").text = f"{vuncom:.10f}"
        ET.SubElement(prod_el, "vProd").text = f"{vtotal:.2f}"
        ET.SubElement(prod_el, "cEANTrib").text = "SEM GTIN"
        ET.SubElement(prod_el, "uTrib").text = it.get("unit","UN")
        ET.SubElement(prod_el, "qTrib").text = f"{qcom:.4f}"
        ET.SubElement(prod_el, "vUnTrib").text = f"{vuncom:.10f}"
        ET.SubElement(prod_el, "indTot").text = "1"
        total_prod += vtotal

        # ICMS (Simples Nacional — CSOSN 400 sem destaque)
        imp = ET.SubElement(det, "imposto")
        icms = ET.SubElement(imp, "ICMS")
        icms_sn = ET.SubElement(icms, "ICMSSN400")
        ET.SubElement(icms_sn, "orig").text = it.get("orig","0")
        ET.SubElement(icms_sn, "CSOSN").text = it.get("csosn","400")
        # PIS
        pis_el = ET.SubElement(imp, "PIS")
        pis_nt = ET.SubElement(pis_el, "PISNT")
        ET.SubElement(pis_nt, "CST").text = "07"
        # COFINS
        cof_el = ET.SubElement(imp, "COFINS")
        cof_nt = ET.SubElement(cof_el, "COFINSNT")
        ET.SubElement(cof_nt, "CST").text = "07"

    disc_val = float(sale.get("discount_value",0))
    total_nf = round(float(sale.get("total",0)), 2)

    # --- total ---
    total_el = ET.SubElement(inf, "total")
    icmstot = ET.SubElement(total_el, "ICMSTot")
    for tag, val in [("vBC","0.00"),("vICMS","0.00"),("vICMSDeson","0.00"),
                     ("vFCPDescomp","0.00"),("vBCST","0.00"),("vST","0.00"),
                     ("vFCPST","0.00"),("vFCPSTRet","0.00"),
                     ("vProd",f"{total_prod:.2f}"),("vFrete","0.00"),("vSeg","0.00"),
                     ("vDesc",f"{disc_val:.2f}"),("vII","0.00"),("vIPI","0.00"),
                     ("vIPIDevol","0.00"),("vPIS","0.00"),("vCOFINS","0.00"),
                     ("vOutro","0.00"),("vNF",f"{total_nf:.2f}"),("vTotTrib","0.00")]:
        ET.SubElement(icmstot, tag).text = val

    # --- transp ---
    transp = ET.SubElement(inf, "transp")
    ET.SubElement(transp, "modFrete").text = "9"

    # --- pag ---
    pag_el = ET.SubElement(inf, "pag")
    det_pag = ET.SubElement(pag_el, "detPag")
    pay_map = {
        "Dinheiro":"01","Cheque":"02","Cartao de Credito":"03","Cartao de Debito":"04",
        "Credito Loja":"05","Vale Alimentacao":"10","Vale Refeicao":"11",
        "Vale Presente":"12","Vale Combustivel":"13","PIX":"17","Boleto":"15",
        "Transferencia":"19","Fiado / A Prazo":"90","Outros":"99"
    }
    pm = sale.get("payment_method","Dinheiro")
    tpag = next((v for k,v in pay_map.items() if k.lower() in pm.lower()), "01")
    ET.SubElement(det_pag, "indPag").text = "0"
    ET.SubElement(det_pag, "tPag").text = tpag
    ET.SubElement(det_pag, "vPag").text = f"{total_nf:.2f}"

    # --- infAdic ---
    info_adic = ET.SubElement(inf, "infAdic")
    obs_text = _clean(sale.get("observations",""))
    info_fisco = f"NFC-e emitida pelo Sistema MotoPecas v3.0. Amb:{amb}"
    if obs_text: info_fisco += f" | {obs_text[:100]}"
    ET.SubElement(info_adic, "infAdFisco").text = info_fisco

    # --- infRespTec ---
    resp = ET.SubElement(inf, "infRespTec")
    ET.SubElement(resp, "CNPJ").text = cnpj
    ET.SubElement(resp, "xContato").text = _clean(empresa.get("razao_social",""))[:60]
    ET.SubElement(resp, "email").text = empresa.get("email","sistema@motopecas.com")
    ET.SubElement(resp, "fone").text = _digits(empresa.get("telefone","00000000000"))

    xml_str = ET.tostring(root, encoding="unicode", xml_declaration=False)
    return '<?xml version="1.0" encoding="UTF-8"?>' + xml_str


def _gerar_danfe_nfce(path, empresa, nfce_cfg, sale, nf_num, chave44, xml_bytes=None):
    """Generate DANFE NFC-e PDF (80mm thermal paper style)."""
    if not PDF_OK: raise ImportError("reportlab nao instalado")

    from reportlab.pdfgen import canvas as CV
    from reportlab.lib.units import mm
    from io import BytesIO

    PAGE_W = 80 * mm
    MARGIN = 4 * mm
    CONTENT_W = PAGE_W - 2 * MARGIN

    # --- Calculate page height dynamically ---
    n_items = len(sale.get("items",[]))
    PAGE_H = (85 + n_items * 28 + 120) * mm
    if PAGE_H < 180 * mm: PAGE_H = 180 * mm

    c = CV.Canvas(path, pagesize=(PAGE_W, PAGE_H))
    y = PAGE_H - MARGIN

    def line(w=0.3): c.setLineWidth(w); c.line(MARGIN, y, PAGE_W-MARGIN, y)
    def dashed():
        c.setDash(2, 2); c.setLineWidth(0.3); c.line(MARGIN, y, PAGE_W-MARGIN, y); c.setDash()

    def text(txt, size=7, bold=False, align="L", color=(0,0,0)):
        c.setFillColorRGB(*color)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        txt = str(txt)
        if align == "C":
            c.drawCentredString(PAGE_W/2, y, txt)
        elif align == "R":
            c.drawRightString(PAGE_W-MARGIN, y, txt)
        else:
            c.drawString(MARGIN, y, txt)

    def nl(pts=9): return pts  # pts to subtract from y

    amb = nfce_cfg.get("ambiente","2")
    emp = empresa

    # Header
    y -= 6*mm
    text(_clean(emp.get("nome_fantasia", emp.get("razao_social",""))), size=10, bold=True, align="C")
    y -= nl(12)
    text(_clean(emp.get("razao_social","")), size=7, align="C")
    y -= nl(9)
    cnpj_fmt = emp.get("cnpj",""); ie_fmt = emp.get("ie","")
    text(f"CNPJ: {cnpj_fmt}", size=7, align="C")
    y -= nl(9)
    end_str = f"{_clean(emp.get('endereco',''))} - {_clean(emp.get('bairro',''))}"
    text(end_str[:50], size=6, align="C")
    y -= nl(8)
    text(f"{_clean(emp.get('municipio',''))}-{emp.get('uf','')} CEP:{emp.get('cep','')}", size=6, align="C")
    y -= nl(8)
    text(f"IE: {ie_fmt}  Tel: {emp.get('telefone','')}", size=6, align="C")
    y -= nl(12)

    if amb == "2":
        c.setFillColorRGB(0.9, 0.1, 0.1)
        c.rect(MARGIN, y-5*mm, CONTENT_W, 5*mm, fill=1, stroke=0)
        c.setFillColorRGB(1,1,1)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(PAGE_W/2, y-3.5*mm, "AMBIENTE DE HOMOLOGACAO - SEM VALOR FISCAL")
        c.setFillColorRGB(0,0,0)
        y -= 6*mm

    line(); y -= nl(10)
    text("DANFE NFC-e", size=9, bold=True, align="C")
    y -= nl(10)
    text("Documento Auxiliar da Nota Fiscal de", size=6, align="C")
    y -= nl(8)
    text("Consumidor Eletronica", size=6, align="C")
    y -= nl(8)
    text(f"N. {str(nf_num).zfill(9)}   Serie: {nfce_cfg.get('serie','001')}", size=7, align="C")
    y -= nl(9)
    text(f"Emissao: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", size=7, align="C")
    y -= nl(11)
    dashed(); y -= nl(10)

    # Items header
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(MARGIN, y, "PRODUTO")
    c.drawString(MARGIN + CONTENT_W*0.52, y, "QTD")
    c.drawString(MARGIN + CONTENT_W*0.65, y, "V.UNIT")
    c.drawRightString(PAGE_W-MARGIN, y, "TOTAL")
    y -= nl(9)
    line(0.2); y -= nl(9)

    for it in sale.get("items",[]):
        disc_f = 1 - float(it.get("discount",0))/100
        tot = float(it["quantity"]) * float(it["unit_price"]) * disc_f
        nome = _clean(it.get("name",""))[:28]
        c.setFont("Helvetica", 6.5)
        c.drawString(MARGIN, y, nome)
        y -= nl(8)
        c.setFont("Helvetica", 6.5)
        c.drawString(MARGIN, y, f"  Cod:{it.get('code','')}")
        c.drawString(MARGIN + CONTENT_W*0.52, y, f"{float(it['quantity']):.2f}")
        c.drawString(MARGIN + CONTENT_W*0.65, y, f"R${float(it['unit_price']):.2f}")
        c.drawRightString(PAGE_W-MARGIN, y, f"R${tot:.2f}")
        if it.get("discount",0) > 0:
            y -= nl(7)
            c.setFont("Helvetica", 6)
            c.drawString(MARGIN, y, f"  Desconto: {it['discount']:.1f}%")
        y -= nl(10)

    dashed(); y -= nl(10)

    # Totals
    def tot_row(label, value, bold=False):
        nonlocal y
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 7 if bold else 6.5)
        c.drawString(MARGIN, y, label)
        c.drawRightString(PAGE_W-MARGIN, y, value)
        y -= nl(10 if bold else 9)

    sub = float(sale.get("subtotal",0))
    disc = float(sale.get("discount_value",0))
    tot_nf = float(sale.get("total",0))
    tot_row("Subtotal:", f"R$ {sub:.2f}")
    if disc > 0: tot_row(f"Desconto ({sale.get('discount_pct',0):.1f}%):", f"-R$ {disc:.2f}")
    tot_row("TOTAL:", f"R$ {tot_nf:.2f}", bold=True)
    y -= nl(3)
    tot_row(f"Pagamento: {sale.get('payment_method','')}", f"R$ {tot_nf:.2f}")

    cust = sale.get("customer_name","Consumidor Final")
    y -= nl(4)
    line(0.2); y -= nl(9)
    c.setFont("Helvetica", 6.5)
    c.drawString(MARGIN, y, f"Consumidor: {_clean(cust)}")
    y -= nl(8)
    obs = sale.get("observations","")
    if obs:
        c.setFont("Helvetica", 6)
        c.drawString(MARGIN, y, f"Obs: {_clean(obs)[:55]}")
        y -= nl(8)

    dashed(); y -= nl(10)

    # QR Code
    chave_fmt = _formatar_chave(chave44)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(PAGE_W/2, y, "Chave de Acesso")
    y -= nl(9)
    c.setFont("Helvetica", 5.5)
    # Draw key in segments
    for seg_start in range(0, 44, 11):
        seg = chave44[seg_start:seg_start+11]
        c.drawCentredString(PAGE_W/2, y, seg)
        y -= nl(7)
    y -= nl(4)

    # QR Code generation
    qr_url = (f"https://www.nfce.fazenda.sp.gov.br/consulta?"
              f"p={chave44}|2|{nfce_cfg.get('ambiente','2')}|1|"
              f"{nfce_cfg.get('csc_token','')[:8]}")

    qr_size = 30 * mm
    qr_x = (PAGE_W - qr_size) / 2
    qr_y = y - qr_size - 2*mm

    if QR_OK:
        try:
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M,
                                box_size=3, border=1)
            qr.add_data(qr_url); qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            from io import BytesIO
            buf = BytesIO(); qr_img.save(buf, format="PNG"); buf.seek(0)
            c.drawImage(buf, qr_x, qr_y, width=qr_size, height=qr_size, preserveAspectRatio=True)
        except Exception:
            c.rect(qr_x, qr_y, qr_size, qr_size, fill=0)
            c.setFont("Helvetica",6); c.drawCentredString(PAGE_W/2, qr_y+qr_size/2,"QR Code")
    else:
        c.rect(qr_x, qr_y, qr_size, qr_size, fill=0)
        c.setFont("Helvetica",6); c.drawCentredString(PAGE_W/2, qr_y+qr_size/2,"QR Code")

    y = qr_y - nl(8)
    c.setFont("Helvetica",6); c.drawCentredString(PAGE_W/2, y,"Consulte pela Chave de Acesso em:")
    y -= nl(8); c.setFont("Helvetica",5.5)
    c.drawCentredString(PAGE_W/2, y, "www.nfce.fazenda.sp.gov.br")
    y -= nl(10)
    line(0.3); y -= nl(8)
    c.setFont("Helvetica",5.5)
    c.drawCentredString(PAGE_W/2, y,"MotoPecas Sistema v3.0 — Emitido em "+datetime.now().strftime("%d/%m/%Y %H:%M"))

    c.save()

# ═════════════════════════════════════════════════════════════════════════════
#  NOTA FISCAL PDF
# ═════════════════════════════════════════════════════════════════════════════
def _gerar_nota_fiscal(path, cart, customer_name, customer_obj, subtotal,
                       disc_pct, disc_val, total, payment, observations, dm):
    """Generate a professional invoice PDF using reportlab."""
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm
    )
    styles = getSampleStyleSheet()
    accent = colors.HexColor("#FF6B35")
    dark   = colors.HexColor("#1A1A1A")
    gray   = colors.HexColor("#666666")
    light  = colors.HexColor("#F5F5F5")
    white  = colors.white

    title_style  = ParagraphStyle("Title2",  parent=styles["Normal"], fontSize=22, textColor=accent,  fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)
    sub_style    = ParagraphStyle("Sub2",    parent=styles["Normal"], fontSize=10, textColor=gray,    fontName="Helvetica",      alignment=TA_CENTER, spaceAfter=2)
    h1_style     = ParagraphStyle("H1",      parent=styles["Normal"], fontSize=11, textColor=dark,    fontName="Helvetica-Bold", spaceAfter=2)
    normal_style = ParagraphStyle("Normal2", parent=styles["Normal"], fontSize=9,  textColor=dark,    fontName="Helvetica")
    small_style  = ParagraphStyle("Small",   parent=styles["Normal"], fontSize=8,  textColor=gray,    fontName="Helvetica")
    right_style  = ParagraphStyle("Right",   parent=styles["Normal"], fontSize=10, textColor=dark,    fontName="Helvetica-Bold", alignment=TA_RIGHT)
    total_style  = ParagraphStyle("Total",   parent=styles["Normal"], fontSize=14, textColor=accent,  fontName="Helvetica-Bold", alignment=TA_RIGHT)

    story = []
    W = A4[0] - 30*mm  # usable width

    # ── NOTA FISCAL CONFIG (dados configurados em Configuracoes) ──────────────
    nf_cfg      = dm.get_nota_fiscal_cfg() if dm and hasattr(dm, "get_nota_fiscal_cfg") else {}
    _nome_loja  = nf_cfg.get("nome_loja",  "MOTO PECAS & MECANICA")
    _slogan     = nf_cfg.get("slogan",     "Sistema de Gestao e Vendas")
    _rodape_txt = nf_cfg.get("rodape",     "Obrigado pela preferencia!")

    # ── HEADER ────────────────────────────────────────────────────────────────
    header_data = [[
        Paragraph(f"<b><font size=24 color='#FF6B35'>{_nome_loja}</font></b><br/>"
                  f"<font size=10 color='#666666'>{_slogan}</font>", styles["Normal"]),
        Paragraph(
            f"<b><font size=18 color='#1A1A1A'>NOTA FISCAL</font></b><br/>"
            f"<font size=9 color='#666666'>Emitida em: {datetime.now().strftime('%d/%m/%Y  %H:%M')}</font>",
            ParagraphStyle("RH", parent=styles["Normal"], alignment=TA_RIGHT)
        )
    ]]
    header_table = Table(header_data, colWidths=[W*0.55, W*0.45])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#FFF8F5")),
        ("ROUNDEDCORNERS", [8,8,8,8]),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING", (0,0), (0,-1), 12),
        ("RIGHTPADDING", (-1,0), (-1,-1), 12),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=2, color=accent))
    story.append(Spacer(1, 4*mm))

    # ── CLIENT + INFO ─────────────────────────────────────────────────────────
    veiculo  = customer_obj.get("veiculo","") if customer_obj else ""
    placa    = customer_obj.get("placa","")   if customer_obj else ""
    phone    = customer_obj.get("phone","--") if customer_obj else "--"
    address  = customer_obj.get("address","--") if customer_obj else "--"

    info_data = [[
        [Paragraph("<b>DADOS DO CLIENTE</b>", h1_style),
         Paragraph(f"Nome: {customer_name}", normal_style),
         Paragraph(f"Veiculo: {veiculo}" if veiculo else "Veiculo: --", normal_style),
         Paragraph(f"Placa: {placa}" if placa else "Placa: --", normal_style),
         Paragraph(f"Telefone: {phone}", normal_style),
         Paragraph(f"Endereco: {address}", normal_style)],
        [Paragraph("<b>INFORMACOES DE PAGAMENTO</b>", h1_style),
         Paragraph(f"Forma de Pagamento: <b>{payment}</b>", normal_style),
         Paragraph(f"Data de Emissao: <b>{datetime.now().strftime('%d/%m/%Y')}</b>", normal_style),
         Paragraph(f"Hora: <b>{datetime.now().strftime('%H:%M:%S')}</b>", normal_style),
         Paragraph(f"Status: <b>EMITIDA</b>", normal_style)],
    ]]
    info_table = Table(info_data, colWidths=[W*0.55, W*0.45])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BACKGROUND", (0,0), (-1,-1), light),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6*mm))

    # ── ITEMS TABLE ──────────────────────────────────────────────────────────
    story.append(Paragraph("<b>ITENS DA NOTA FISCAL</b>", h1_style))
    story.append(Spacer(1, 2*mm))

    item_header = ["Cod.", "Produto / Descricao", "Qtd", "Preco Unit.", "Desc.%", "Total"]
    col_ws = [W*0.08, W*0.38, W*0.08, W*0.16, W*0.10, W*0.18]
    item_rows = [item_header]

    for it in cart:
        disc_f = 1 - it["discount"]/100
        row_total = it["quantity"] * it["unit_price"] * disc_f
        item_rows.append([
            it.get("code",""),
            it.get("name",""),
            str(it["quantity"]),
            fmtR(it["unit_price"]),
            f"{it['discount']:.1f}%",
            fmtR(row_total),
        ])

    item_table = Table(item_rows, colWidths=col_ws, repeatRows=1)
    item_table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0,0), (-1,0), accent),
        ("TEXTCOLOR",     (0,0), (-1,0), white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("TOPPADDING",    (0,0), (-1,0), 7),
        ("BOTTOMPADDING", (0,0), (-1,0), 7),
        # Data rows
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ALIGN",         (0,1), (0,-1), "CENTER"),   # code
        ("ALIGN",         (2,1), (2,-1), "CENTER"),   # qty
        ("ALIGN",         (3,1), (-1,-1), "RIGHT"),   # prices
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("LEFTPADDING",   (1,0), (1,-1), 6),
        # Alternating rows
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [white, light]),
        # Grid
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0,0), (-1,0), 1.5, accent),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 5*mm))

    # ── TOTALS ────────────────────────────────────────────────────────────────
    totals_data = [
        ["", "Subtotal:", fmtR(subtotal)],
        ["", f"Desconto ({disc_pct:.1f}%):", f"- {fmtR(disc_val)}"],
        ["", "TOTAL:", fmtR(total)],
    ]
    totals_table = Table(totals_data, colWidths=[W*0.55, W*0.25, W*0.20])
    totals_table.setStyle(TableStyle([
        ("FONTNAME",      (0,0), (-1,-2), "Helvetica"),
        ("FONTSIZE",      (0,0), (-1,-2), 9),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (1,-1), (2,-1), 13),
        ("TEXTCOLOR",     (2,-1), (2,-1), accent),
        ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LINEABOVE",     (1,-1), (-1,-1), 1.5, accent),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    story.append(Spacer(1, 3*mm))

    # ── OBSERVATIONS ─────────────────────────────────────────────────────────
    if observations.strip():
        story.append(Paragraph("<b>Observacoes:</b>", h1_style))
        story.append(Paragraph(observations.strip(), normal_style))
        story.append(Spacer(1, 3*mm))

    # ── FOOTER ───────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=accent))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"<font color='#FF6B35'>{_nome_loja}</font>  |  "
        f"Nota emitida em {datetime.now().strftime('%d/%m/%Y as %H:%M')}  |  "
        f"{_rodape_txt}",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=gray, alignment=TA_CENTER)
    ))

    doc.build(story)


# ═════════════════════════════════════════════════════════════════════════════
#  DATA MANAGER
# ═════════════════════════════════════════════════════════════════════════════
class DataManager:
    def __init__(self):
        self.data = {}
        self.load()

    def load(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                for k, v in self._defaults().items():
                    if k not in self.data:
                        self.data[k] = v
                # ensure built-in users exist
                self._ensure_builtin_users()
            except Exception:
                self.data = self._defaults()
                self._add_samples()
        else:
            self.data = self._defaults()
            self._add_samples()
            self.save()

    def _defaults(self):
        return {
            "products": [], "customers": [], "sales": [],
            "users": [],
            "settings": {"theme": "dark", "background_image": "", "accent_color": "#FF6B35", "app_icon": "",
                         "nota_fiscal": {"nome_loja": "MOTO PECAS & MECANICA",
                                          "slogan":    "Sistema de Gestao e Vendas",
                                          "endereco":  "", "telefone": "",
                                          "cnpj":      "",
                                          "rodape":    "Obrigado pela preferencia!"},
                         "meu_whatsapp": ""},
            "empresa": {
                "razao_social": "MOTO PECAS E MECANICA LTDA",
                "nome_fantasia": "MOTO PECAS",
                "cnpj": "00.000.000/0001-00",
                "ie": "000.000.000.000",
                "im": "",
                "endereco": "Rua das Pecas, 100",
                "bairro": "Centro",
                "municipio": "Sao Paulo",
                "uf": "SP",
                "cep": "01000-000",
                "telefone": "(11) 0000-0000",
                "email": "",
                "regime": "1",
                "cnae": "4530-7/03"
            },
            "nfce": {
                "serie": "001",
                "proxima_nf": 1,
                "ambiente": "2",
                "csc_id": "000001",
                "csc_token": "AAAA0000000000000000000000000000",
                "cuf": "35",
                "cmun": "3550308"
            }
        }

    def _ensure_builtin_users(self):
        """Always guarantee admin and funcionario accounts exist."""
        existing = {u["username"] for u in self.data.get("users", [])}
        if "admin" not in existing:
            self.data["users"].insert(0, {
                "id": "admin-fixed", "username": "admin", "password": "admin",
                "role": "admin", "name": "Administrador", "active": True
            })
        if "funcionario" not in existing:
            self.data["users"].append({
                "id": "func-fixed", "username": "funcionario", "password": "123moto",
                "role": "operator", "name": "Funcionario", "active": True
            })

    def _add_samples(self):
        self._ensure_builtin_users()
        self.data["products"] = self._sample_products()
        self.data["customers"] = self._sample_customers()

    def save(self):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _sample_products(self):
        rows = [
            ("FLT001","Filtro de Oleo","Filtros","Bosch","UN",15,35,50,10),
            ("FLT002","Filtro de Ar","Filtros","Mann","UN",20,45,30,8),
            ("FLT003","Filtro de Combustivel","Filtros","WIX","UN",18,40,25,5),
            ("FRE001","Pastilha de Freio Dianteira","Freios","Bosch","JG",45,95,20,5),
            ("FRE002","Disco de Freio Dianteiro","Freios","Brembo","UN",80,165,15,4),
            ("FRE003","Fluido de Freio DOT4","Freios","ATE","FR",12,28,40,10),
            ("OLE001","Oleo 5W30 Sintetico","Oleos","Castrol","LT",28,55,100,20),
            ("OLE002","Oleo 10W40 Semissintetico","Oleos","Mobil","LT",22,45,80,15),
            ("OLE003","Oleo de Cambio","Oleos","Valvoline","LT",30,60,30,8),
            ("VEL001","Vela de Ignicao","Motor","NGK","UN",12,25,5,10),
            ("BAT001","Bateria 60Ah","Eletrica","Moura","UN",280,450,10,3),
            ("ALT001","Correia Alternador","Motor","Dayco","UN",35,72,18,5),
            ("SUS001","Amortecedor Dianteiro","Suspensao","Monroe","UN",120,220,8,2),
            ("SUS002","Barra Estabilizadora","Suspensao","TRW","UN",65,130,6,2),
            ("RAD001","Radiador","Arrefecimento","Valeo","UN",350,620,4,1),
        ]
        return [{"id": str(uuid.uuid4()), "code": r[0], "name": r[1],
                 "category": r[2], "brand": r[3], "unit": r[4],
                 "cost_price": float(r[5]), "sale_price": float(r[6]),
                 "stock": r[7], "min_stock": r[8], "description": ""} for r in rows]

    def _sample_customers(self):
        return [
            {"id": str(uuid.uuid4()), "name": "Joao Carlos Silva",
             "cpf_cnpj": "123.456.789-00", "phone": "(21) 98765-4321",
             "email": "joao@email.com", "address": "Rua das Flores, 100"},
            {"id": str(uuid.uuid4()), "name": "Maria Santos Oliveira",
             "cpf_cnpj": "987.654.321-00", "phone": "(21) 91234-5678",
             "email": "maria@email.com", "address": "Av. Brasil, 200"},
            {"id": str(uuid.uuid4()), "name": "Auto Center Mega Ltda",
             "cpf_cnpj": "12.345.678/0001-90", "phone": "(21) 3333-4444",
             "email": "contato@mega.com", "address": "Rod. Dutra, 300"},
            {"id": str(uuid.uuid4()), "name": "Pedro Melo Mecanica",
             "cpf_cnpj": "45.678.901/0001-23", "phone": "(21) 97777-8888",
             "email": "pedro@mec.com", "address": "Rua Industrial, 55"},
        ]

    # PRODUCTS
    def get_products(self): return self.data["products"]
    def add_product(self, p): p["id"] = str(uuid.uuid4()); self.data["products"].append(p); self.save()
    def update_product(self, pid, u):
        for i, p in enumerate(self.data["products"]):
            if p["id"] == pid: self.data["products"][i] = u; break
        self.save()
    def delete_product(self, pid):
        self.data["products"] = [p for p in self.data["products"] if p["id"] != pid]; self.save()
    def get_product_by_id(self, pid):
        return next((p for p in self.data["products"] if p["id"] == pid), None)

    def get_product_by_barcode(self, barcode):
        """Find product by EAN/barcode. Also tries matching last digits against code."""
        bc = barcode.strip()
        if not bc: return None
        # Exact barcode match
        exact = next((p for p in self.data["products"] if p.get("barcode","") == bc), None)
        if exact: return exact
        # Try matching barcode against product code
        by_code = next((p for p in self.data["products"] if p.get("code","").upper() == bc.upper()), None)
        if by_code: return by_code
        return None

    # CUSTOMERS
    def get_customers(self): return self.data["customers"]
    def add_customer(self, c): c["id"] = str(uuid.uuid4()); self.data["customers"].append(c); self.save()
    def update_customer(self, cid, u):
        for i, c in enumerate(self.data["customers"]):
            if c["id"] == cid: self.data["customers"][i] = u; break
        self.save()
    def delete_customer(self, cid):
        self.data["customers"] = [c for c in self.data["customers"] if c["id"] != cid]; self.save()
    def get_customer_by_id(self, cid):
        return next((c for c in self.data["customers"] if c["id"] == cid), None)

    # SALES
    def get_sales(self): return self.data["sales"]
    def add_sale(self, sale):
        sale["id"] = str(uuid.uuid4()); sale["date"] = datetime.now().isoformat()
        for item in sale.get("items", []):
            for p in self.data["products"]:
                if p["id"] == item["product_id"]:
                    p["stock"] = max(0, p["stock"] - item["quantity"]); break
        self.data["sales"].append(sale); self.save(); return sale["id"]

    # USERS
    def get_users(self): return self.data.get("users", [])
    def get_user_by_username(self, username):
        return next((u for u in self.get_users() if u.get("username","").lower() == username.lower()), None)
    def authenticate(self, username, password):
        u = self.get_user_by_username(username)
        return u if (u and u.get("password") == password and u.get("active", True)) else None
    def add_user(self, user):
        if self.get_user_by_username(user["username"]): return False, "Login ja existe."
        user["id"] = str(uuid.uuid4()); user["active"] = True
        self.data["users"].append(user); self.save(); return True, "Usuario criado!"
    def update_user(self, uid, upd):
        for i, u in enumerate(self.data["users"]):
            if u["id"] == uid: upd["id"] = uid; self.data["users"][i] = upd; break
        self.save()
    def delete_user(self, uid):
        self.data["users"] = [u for u in self.data["users"] if u["id"] != uid]; self.save()
    def toggle_user_active(self, uid):
        for u in self.data["users"]:
            if u["id"] == uid: u["active"] = not u.get("active", True); break
        self.save()

    # SALES CANCEL/DELETE
    def cancel_sale(self, sale_id):
        """Mark sale as cancelled and restore stock."""
        for s in self.data["sales"]:
            if s["id"] == sale_id:
                if s.get("status") == "cancelada":
                    return False, "Esta venda ja esta cancelada."
                for item in s.get("items", []):
                    for p in self.data["products"]:
                        if p["id"] == item["product_id"]:
                            p["stock"] += item["quantity"]; break
                s["status"] = "cancelada"
                s["cancelled_at"] = datetime.now().isoformat()
                self.save()
                return True, "Venda cancelada com sucesso!"
        return False, "Venda nao encontrada."

    def delete_sale(self, sale_id):
        self.data["sales"] = [s for s in self.data["sales"] if s["id"] != sale_id]; self.save()

    # ── PARTIAL RETURN (devolucao parcial de itens) ──────────────────────────
    def partial_return(self, sale_id, returns):
        """Devolver itens parcialmente. returns = [{product_id, qty, motivo}]"""
        for s in self.data["sales"]:
            if s["id"] == sale_id:
                if s.get("status") == "cancelada":
                    return False, "Venda cancelada."
                total_refund = 0.0
                for ret in returns:
                    pid = ret["product_id"]; rqty = ret["qty"]; motivo = ret["motivo"]
                    for it in s.get("items", []):
                        if it["product_id"] == pid:
                            already = it.get("returned_qty", 0)
                            avail = it["quantity"] - already
                            if rqty > avail:
                                return False, f"Qtd maxima para devolver: {avail}"
                            it["returned_qty"] = already + rqty
                            disc_f = 1 - it.get("discount", 0) / 100
                            refund = rqty * it["unit_price"] * disc_f
                            total_refund += refund
                            # Restaurar estoque se arrependimento
                            if motivo in ("Arrependimento", "Outro"):
                                for p in self.data["products"]:
                                    if p["id"] == pid: p["stock"] += rqty; break
                            # Avaria/Defeito -> vai para Trocas
                            if motivo in ("Avaria", "Defeito"):
                                troca = {
                                    "id": str(uuid.uuid4()),
                                    "product_id": pid,
                                    "product_name": it.get("name", ""),
                                    "product_code": it.get("code", ""),
                                    "quantity": rqty,
                                    "motivo": motivo,
                                    "sale_id": sale_id,
                                    "customer_name": s.get("customer_name", "Consumidor Final"),
                                    "date": datetime.now().isoformat(),
                                    "status": "pendente",
                                    "unit_price": it["unit_price"],
                                }
                                self._save_troca(troca)
                            break
                # Atualizar totais da venda
                s["refunded"] = s.get("refunded", 0) + total_refund
                s.setdefault("returns_log", []).append({
                    "date": datetime.now().isoformat(),
                    "items": returns,
                    "refund": total_refund
                })
                # Se todos itens devolvidos, marcar como devolvida
                all_ret = all(it.get("returned_qty", 0) >= it["quantity"] for it in s.get("items", []))
                if all_ret: s["status"] = "devolvida"
                self.save()
                return True, f"Devolucao registrada! Estorno: {total_refund:.2f}"
        return False, "Venda nao encontrada."

    # ── TROCAS (avaria/defeito) ──────────────────────────────────────────────
    def _load_trocas(self):
        if os.path.exists(TROCAS_FILE):
            try:
                with open(TROCAS_FILE, "r", encoding="utf-8") as f: return json.load(f)
            except: pass
        return []

    def _save_trocas(self, data):
        with open(TROCAS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _save_troca(self, troca):
        data = self._load_trocas(); data.append(troca); self._save_trocas(data)

    def get_trocas(self): return self._load_trocas()

    def update_troca(self, tid, status):
        data = self._load_trocas()
        for t in data:
            if t["id"] == tid:
                old_status = t.get("status")
                t["status"] = status
                t["updated_at"] = datetime.now().isoformat()
                
                # Handling stock changes based on status transition
                if status == "devolvido_estoque" and old_status != "devolvido_estoque":
                    for p in self.data["products"]:
                        if p["id"] == t["product_id"]:
                            p["stock"] += t["quantity"]; break
                    self.save()
                elif old_status == "devolvido_estoque" and status != "devolvido_estoque":
                    for p in self.data["products"]:
                        if p["id"] == t["product_id"]:
                            p["stock"] -= t["quantity"]; break
                    self.save()
                break
        self._save_trocas(data)

    def delete_troca(self, tid):
        data = [t for t in self._load_trocas() if t["id"] != tid]
        self._save_trocas(data)

    # ── CAIXA (controle de caixa PDV) ────────────────────────────────────────
    def _load_caixa(self):
        if os.path.exists(CAIXA_FILE):
            try:
                with open(CAIXA_FILE, "r", encoding="utf-8") as f: return json.load(f)
            except: pass
        return {"aberto": False, "historico": []}

    def _save_caixa(self, data):
        with open(CAIXA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def get_caixa(self): return self._load_caixa()

    def abrir_caixa(self, valor_inicial, operador):
        cx = self._load_caixa()
        if cx.get("aberto"): return False, "Caixa ja esta aberto."
        cx["aberto"] = True
        cx["abertura"] = {
            "valor_inicial": valor_inicial,
            "operador": operador,
            "data": datetime.now().isoformat(),
            "movimentacoes": []
        }
        self._save_caixa(cx); return True, "Caixa aberto!"

    def fechar_caixa(self, valor_contado, operador):
        cx = self._load_caixa()
        if not cx.get("aberto"): return False, "Caixa nao esta aberto.", {}
        ab = cx.get("abertura", {})
        vi = ab.get("valor_inicial", 0)
        # Calcular vendas do periodo
        dt_abertura = ab.get("data", "")
        vendas_periodo = [s for s in self.data["sales"]
                          if s.get("date", "") >= dt_abertura and s.get("status") != "cancelada"]
        total_vendas = sum(s.get("total", 0) for s in vendas_periodo)
        # Movimentacoes
        movs = ab.get("movimentacoes", [])
        total_sangria = sum(m["valor"] for m in movs if m["tipo"] == "sangria")
        total_suprimento = sum(m["valor"] for m in movs if m["tipo"] == "suprimento")
        esperado = vi + total_vendas + total_suprimento - total_sangria
        diferenca = valor_contado - esperado
        # Por forma de pagamento
        por_pagamento = defaultdict(float)
        for s in vendas_periodo:
            por_pagamento[s.get("payment_method", "Outros")] += s.get("total", 0)
        resumo = {
            "valor_inicial": vi, "total_vendas": total_vendas,
            "total_sangria": total_sangria, "total_suprimento": total_suprimento,
            "esperado": esperado, "contado": valor_contado, "diferenca": diferenca,
            "n_vendas": len(vendas_periodo), "por_pagamento": dict(por_pagamento),
            "operador_abertura": ab.get("operador", ""),
            "operador_fechamento": operador,
            "data_abertura": dt_abertura,
            "data_fechamento": datetime.now().isoformat(),
            "movimentacoes": movs
        }
        cx["historico"].append(resumo)
        cx["aberto"] = False; cx["abertura"] = {}
        self._save_caixa(cx); return True, "Caixa fechado!", resumo

    def add_movimentacao_caixa(self, tipo, valor, motivo, operador):
        cx = self._load_caixa()
        if not cx.get("aberto"): return False, "Caixa nao esta aberto."
        mov = {"tipo": tipo, "valor": valor, "motivo": motivo,
               "operador": operador, "data": datetime.now().isoformat()}
        cx.setdefault("abertura", {}).setdefault("movimentacoes", []).append(mov)
        self._save_caixa(cx); return True, f"{'Sangria' if tipo=='sangria' else 'Suprimento'} registrado!"

    # ── NPS ──────────────────────────────────────────────────────────────────
    def _load_nps(self):
        if os.path.exists(NPS_FILE):
            try:
                with open(NPS_FILE, "r", encoding="utf-8") as f: return json.load(f)
            except: pass
        return []

    def _save_nps_data(self, data):
        with open(NPS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def add_nps(self, record):
        data = self._load_nps(); record["id"] = str(uuid.uuid4())
        record["date"] = datetime.now().isoformat()
        data.append(record); self._save_nps_data(data)

    def get_nps(self): return self._load_nps()

    # ── IMPORT XML NF-e ──────────────────────────────────────────────────────
    def import_xml_nfe(self, xml_path):
        """Parse XML NF-e de entrada e retorna lista de itens."""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
            # Try with and without namespace
            dets = root.findall('.//nfe:det', ns)
            if not dets:
                dets = root.findall('.//det')
            items = []
            for det in dets:
                prod = det.find('nfe:prod', ns) or det.find('prod')
                if prod is None: continue
                def _gt(tag):
                    el = prod.find(f'nfe:{tag}', ns) or prod.find(tag)
                    return el.text.strip() if el is not None and el.text else ""
                item = {
                    "code": _gt("cProd"),
                    "ean": _gt("cEAN"),
                    "name": _gt("xProd"),
                    "ncm": _gt("NCM"),
                    "cfop": _gt("CFOP"),
                    "unit": _gt("uCom"),
                    "quantity": float(_gt("qCom") or 0),
                    "unit_price": float(_gt("vUnCom") or 0),
                    "total": float(_gt("vProd") or 0),
                }
                items.append(item)
            return items, ""
        except Exception as e:
            return [], str(e)

    # EMPRESA & NFCE CONFIG
    def get_empresa(self):
        e = self.data.get("empresa", {})
        if not e: self.data["empresa"] = self._defaults()["empresa"]
        return self.data.get("empresa", self._defaults()["empresa"])

    def save_empresa(self, e): self.data["empresa"] = e; self.save()

    def get_nfce(self):
        n = self.data.get("nfce", {})
        if not n: self.data["nfce"] = self._defaults()["nfce"]
        return self.data.get("nfce", self._defaults()["nfce"])

    def save_nfce(self, n): self.data["nfce"] = n; self.save()

    def next_nf_number(self):
        nfce = self.get_nfce(); num = int(nfce.get("proxima_nf", 1))
        nfce["proxima_nf"] = num + 1; self.save_nfce(nfce); return num

    # SETTINGS
    def get_settings(self):
        s = self.data.get("settings", {})
        s.setdefault("accent_color", "#FF6B35"); s.setdefault("theme", "dark")
        s.setdefault("background_image", ""); s.setdefault("app_icon", "")
        s.setdefault("meu_whatsapp", "")
        return s
    def save_settings(self, s): self.data["settings"] = s; self.save()

    def get_nota_fiscal_cfg(self):
        nf = self.get_settings().get("nota_fiscal", {})
        nf.setdefault("nome_loja", "MOTO PECAS & MECANICA")
        nf.setdefault("slogan",    "Sistema de Gestao e Vendas")
        nf.setdefault("endereco",  ""); nf.setdefault("telefone",  "")
        nf.setdefault("cnpj",      ""); nf.setdefault("rodape",    "Obrigado pela preferencia!")
        return nf

    def save_nota_fiscal_cfg(self, cfg):
        s = self.get_settings(); s["nota_fiscal"] = cfg; self.save_settings(s)


# ═════════════════════════════════════════════════════════════════════════════
#  STYLESHEET
# ═════════════════════════════════════════════════════════════════════════════
def _lighter(h):
    c = QColor(h); hue, s, v, a = c.getHsvF()
    c.setHsvF(hue, max(0, s * 0.8), min(1.0, v * 1.25), a); return c.name()

def _darker(h):
    c = QColor(h); hue, s, v, a = c.getHsvF()
    c.setHsvF(hue, s, v * 0.72, a); return c.name()

def build_stylesheet(theme="dark", accent="#FF6B35"):
    ah = _lighter(accent) if hasattr(sys.modules[__name__], '_lighter') else "#F97316"
    ad = _darker(accent) if hasattr(sys.modules[__name__], '_darker') else "#EA580C"
    
    bg="#F8FAFC"; surf="#FFFFFF"; surf2="#F1F5F9"; brd="#E2E8F0"
    txt="#0F172A"; txt2="#64748B"; talt="#FFFFFF"; inbg="#F1F5F9"
    side="#FFFFFF"; card="#FFFFFF"; hdr="#FFFFFF"
    ok="#10B981"; warn="#F59E0B"; err="#EF4444"

    return f"""
QMainWindow,QDialog{{background:{bg};color:{txt};}}
QWidget{{background:{bg};color:{txt};font-family:'Segoe UI',Arial,sans-serif;font-size:13px;}}
QFrame#card{{background:{card};border:1px solid {brd};border-radius:12px;}}
QFrame#sidebar{{background:{side};border-right:1px solid {brd};}}
QLabel#title{{font-size:22px;font-weight:bold;color:{txt};}}
QLabel#subtitle{{font-size:12px;color:{txt2};}}
QLabel#metric_value{{font-size:26px;font-weight:bold;color:{accent};}}
QLabel#metric_label{{font-size:11px;color:{txt2};}}
QLabel#section_title{{font-size:16px;font-weight:800;color:{txt};}}
QPushButton{{background:{accent};color:#FFF;border:none;border-radius:8px;padding:8px 18px;font-size:13px;font-weight:bold;}}
QPushButton:hover{{background:{ah};}}
QPushButton:pressed{{background:{ad};}}
QPushButton:disabled{{background:{brd};color:{txt2};}}
QPushButton#btn_secondary{{background:{surf};color:{txt};border:1px solid {brd};}}
QPushButton#btn_secondary:hover{{background:{surf2};}}
QPushButton#btn_danger{{background:{err};color:#FFF;}}
QPushButton#btn_danger:hover{{background:#DC2626;}}
QPushButton#btn_success{{background:{ok};color:#FFF;}}
QPushButton#btn_success:hover{{background:#059669;}}
QPushButton#btn_warning{{background:{warn};color:#111;}}
QPushButton#nav_btn{{background:transparent;color:{txt2};border:none;border-radius:8px;padding:12px 16px;text-align:left;font-size:14px;}}
QPushButton#nav_btn:hover{{background:{surf2};color:{txt};}}
QPushButton#nav_btn_active{{background:{accent};color:#FFF;border:none;border-radius:8px;padding:12px 16px;text-align:left;font-size:14px;font-weight:bold;}}
QLineEdit,QTextEdit,QComboBox,QSpinBox,QDoubleSpinBox,QDateEdit{{background:{inbg};color:{txt};border:1px solid {brd};border-radius:8px;padding:8px 12px;font-size:13px;}}
QLineEdit:focus,QTextEdit:focus,QComboBox:focus,QSpinBox:focus,QDoubleSpinBox:focus,QDateEdit:focus{{border:2px solid {accent};background:{surf};}}
QSpinBox::up-button, QSpinBox::down-button, QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{ width: 0px; }}
QComboBox::drop-down{{border:none;width:28px;}}
QComboBox::down-arrow{{image:none;border-left:5px solid transparent;border-right:5px solid transparent;border-top:6px solid {txt2};margin-right:8px;}}
QComboBox QAbstractItemView{{background:{surf};color:{txt};border:1px solid {brd};selection-background-color:{surf2};selection-color:{txt};border-radius:8px;}}
QTableWidget{{background:{surf};color:{txt};border:1px solid {brd};border-radius:10px;gridline-color:{brd};alternate-background-color:{talt};}}
QTableWidget::item{{padding:6px 10px;}}
QTableWidget::item:selected{{background:{surf2};color:{txt};font-weight:bold;}}
QHeaderView::section{{background:{surf};color:{txt2};padding:10px;border:none;border-bottom:1px solid {brd};font-weight:bold;font-size:12px;}}
QGroupBox{{border:1px solid {brd};border-radius:12px;margin-top:12px;padding:16px 12px 12px 12px;font-weight:bold;color:{txt};}}
QGroupBox::title{{subcontrol-origin:margin;subcontrol-position:top left;padding:0 8px;color:{txt2};font-size:13px;}}
QScrollBar:vertical{{background:{surf};width:8px;border-radius:4px;}}
QScrollBar::handle:vertical{{background:{brd};border-radius:4px;min-height:24px;}}
QScrollBar::handle:vertical:hover{{background:{txt2};}}
QScrollBar:horizontal{{background:{surf};height:8px;border-radius:4px;}}
QScrollBar::handle:horizontal{{background:{brd};border-radius:4px;min-width:24px;}}
QScrollBar::add-line,QScrollBar::sub-line{{width:0;height:0;}}
QMessageBox{{background:{surf};color:{txt};}}
QStatusBar{{background:{hdr};color:{txt2};border-top:1px solid {brd};font-size:12px;padding:4px 8px;}}
QRadioButton{{color:{txt};spacing:8px;}}
QRadioButton::indicator{{width:18px;height:18px;border-radius:9px;border:2px solid {brd};background:{inbg};}}
QRadioButton::indicator:checked{{background:{accent};border-color:{accent};}}
QCheckBox{{color:{txt};spacing:8px;}}
QCheckBox::indicator{{width:18px;height:18px;border-radius:4px;border:2px solid {brd};background:{inbg};}}
QCheckBox::indicator:checked{{background:{accent};border-color:{accent};}}
QSplitter::handle{{background:{brd};}}
"""


# ═════════════════════════════════════════════════════════════════════════════
#  HELPERS / REUSABLE WIDGETS
# ═════════════════════════════════════════════════════════════════════════════

class DraggableDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._dp = None
    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton: self._dp = e.globalPos() - self.pos()
    def mouseMoveEvent(self, e):
        if e.buttons() == Qt.LeftButton and self._dp: self.move(e.globalPos() - self._dp)

class ModernFormDialog(DraggableDialog):
    """Modern frameless base form dialog with shadow."""
    def __init__(self, parent=None, min_width=500):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumWidth(min_width)
        self._build_base()

    def _build_base(self):
        self._outer = QVBoxLayout(self)
        self._outer.setContentsMargins(16, 16, 16, 16)
        self.container = QFrame()
        self.container.setStyleSheet(f"QFrame{{background:#FFFFFF;border-radius:12px;border:1px solid #E2E8F0;}}")
        # Soft shadow
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(30); shadow.setXOffset(0); shadow.setYOffset(6)
        shadow.setColor(QColor(0, 0, 0, 40))
        self.container.setGraphicsEffect(shadow)

        self.main_lay = QVBoxLayout(self.container)
        self.main_lay.setContentsMargins(24, 24, 24, 24)
        self.main_lay.setSpacing(16)
        self._outer.addWidget(self.container)

    def add_title_bar(self, title):
        top_row = QHBoxLayout()
        tlbl = QLabel(title)
        tlbl.setStyleSheet(f"font-size:18px;font-weight:800;color:#0F172A;background:transparent;border:none;")
        top_row.addWidget(tlbl)
        top_row.addStretch()
        btn_close = QPushButton("✕")
        btn_close.setStyleSheet(f"background:transparent; border:none; font-size:16px; color:#64748B; font-weight:bold;")
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.clicked.connect(self.reject)
        top_row.addWidget(btn_close)
        self.main_lay.addLayout(top_row)
        self.main_lay.addSpacing(8)

class ConfirmDialog(DraggableDialog):
    def __init__(self, parent, title, message, icon="⚠️", danger=False, confirm_text=None):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setFixedWidth(420)
        self._build(title, message, icon, danger, confirm_text)

    def _build(self, title, message, icon, danger, confirm_text):
        outer = QVBoxLayout(self); outer.setContentsMargins(16, 16, 16, 16)
        container = QFrame()
        container.setStyleSheet(f"QFrame{{background:#FFFFFF;border-radius:16px;border:1px solid #E2E8F0;padding:0px;}}")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40); shadow.setXOffset(0); shadow.setYOffset(8)
        shadow.setColor(QColor(0, 0, 0, 60))
        container.setGraphicsEffect(shadow)

        lay = QVBoxLayout(container); lay.setContentsMargins(32, 36, 32, 28); lay.setSpacing(0)
        ic_bg = QWidget(); ic_bg.setFixedSize(72, 72)
        base_color = "#FEF2F2" if danger else "#FFFBEB"
        ic_bg.setStyleSheet(f"background:{base_color};border-radius:36px;border:none;")
        ic_l = QVBoxLayout(ic_bg); ic_l.setContentsMargins(0, 0, 0, 0)
        ic_lbl = QLabel(icon); ic_lbl.setAlignment(Qt.AlignCenter)
        ic_lbl.setStyleSheet("font-size:32px;background:transparent;border:none;")
        ic_l.addWidget(ic_lbl)
        ic_row = QHBoxLayout(); ic_row.addStretch(); ic_row.addWidget(ic_bg); ic_row.addStretch()
        lay.addLayout(ic_row); lay.addSpacing(18)

        tlbl = QLabel(title); tlbl.setAlignment(Qt.AlignCenter)
        tlbl.setStyleSheet(f"font-size:19px;font-weight:800;color:#0F172A;background:transparent;border:none;")
        lay.addWidget(tlbl); lay.addSpacing(8)

        mlbl = QLabel(message); mlbl.setWordWrap(True); mlbl.setAlignment(Qt.AlignCenter)
        mlbl.setStyleSheet(f"font-size:13px;color:#64748B;background:transparent;border:none;line-height:1.6;")
        lay.addWidget(mlbl); lay.addSpacing(26)

        btn_row = QHBoxLayout(); btn_row.setSpacing(12)
        btn_cancel = QPushButton("Cancelar"); btn_cancel.setObjectName("btn_secondary")
        btn_cancel.setFixedHeight(46); btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_cancel.clicked.connect(self.reject)
        
        ok_txt = confirm_text or ("Excluir" if danger else "Confirmar")
        ok_obj = "btn_danger" if danger else "btn_success"
        btn_ok = QPushButton(ok_txt); btn_ok.setObjectName(ok_obj)
        btn_ok.setFixedHeight(46); btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.clicked.connect(self.accept)

        btn_row.addWidget(btn_cancel, 1); btn_row.addWidget(btn_ok, 1)
        lay.addLayout(btn_row); outer.addWidget(container)

    @staticmethod
    def ask(parent, title, message, icon="⚠️", danger=False, confirm_text=None):
        dlg = ConfirmDialog(parent, title, message, icon, danger, confirm_text)
        return dlg.exec_() == QDialog.Accepted

class AlertDialog(DraggableDialog):
    def __init__(self, parent, title, message, icon="ℹ️"):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setFixedWidth(380)
        self._build(title, message, icon)

    def _build(self, title, message, icon):
        outer = QVBoxLayout(self); outer.setContentsMargins(16, 16, 16, 16)
        container = QFrame()
        container.setStyleSheet(f"QFrame{{background:#FFFFFF;border-radius:16px;border:1px solid #E2E8F0;padding:0px;}}")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40); shadow.setXOffset(0); shadow.setYOffset(8)
        shadow.setColor(QColor(0, 0, 0, 60))
        container.setGraphicsEffect(shadow)

        lay = QVBoxLayout(container); lay.setContentsMargins(28, 32, 28, 24); lay.setSpacing(10)
        ic_lbl = QLabel(icon); ic_lbl.setAlignment(Qt.AlignCenter)
        ic_lbl.setStyleSheet("font-size:40px;background:transparent;border:none;")
        lay.addWidget(ic_lbl)

        tlbl = QLabel(title); tlbl.setAlignment(Qt.AlignCenter)
        tlbl.setStyleSheet(f"font-size:18px;font-weight:800;color:#0F172A;background:transparent;border:none;")
        lay.addWidget(tlbl)

        mlbl = QLabel(message); mlbl.setWordWrap(True); mlbl.setAlignment(Qt.AlignCenter)
        mlbl.setStyleSheet(f"font-size:13px;color:#64748B;background:transparent;border:none;")
        lay.addWidget(mlbl); lay.addSpacing(14)

        btn_ok = QPushButton("OK"); btn_ok.setObjectName("btn_success")
        btn_ok.setFixedHeight(44); btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.clicked.connect(self.accept)
        brow = QHBoxLayout(); brow.addStretch(1); brow.addWidget(btn_ok, 2); brow.addStretch(1)
        lay.addLayout(brow); outer.addWidget(container)

    @staticmethod
    def show_info(parent, title, message, icon="ℹ️"):
        AlertDialog(parent, title, message, icon).exec_()

def fmtR(v):
    return "R$ {:,.2f}".format(v).replace(",","X").replace(".",",").replace("X",".")

def fmt_date(iso_str, show_time=True):
    """Convert ISO datetime string to Brazilian format DD/MM/YYYY HH:MM."""
    if not iso_str:
        return "--"
    try:
        iso_str = iso_str.replace("T", " ").strip()
        if len(iso_str) >= 16:
            dt = datetime.strptime(iso_str[:16], "%Y-%m-%d %H:%M")
            return dt.strftime("%d/%m/%Y %H:%M") if show_time else dt.strftime("%d/%m/%Y")
        elif len(iso_str) >= 10:
            dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        return iso_str
    except Exception:
        return iso_str

class MetricCard(QFrame):
    def __init__(self, title, value, icon="", color=None, parent=None):
        super().__init__(parent); self.setObjectName("card")
        self.setMinimumHeight(110); self.setMinimumWidth(175)
        lay = QVBoxLayout(self); lay.setContentsMargins(16,14,16,14); lay.setSpacing(6)
        top = QHBoxLayout()
        ico = QLabel(icon); ico.setFont(QFont("Segoe UI Emoji", 22))
        self.val = QLabel(str(value)); self.val.setObjectName("metric_value")
        if color: self.val.setStyleSheet(f"color:{color};font-size:26px;font-weight:bold;")
        top.addWidget(ico); top.addStretch(); top.addWidget(self.val); lay.addLayout(top)
        t = QLabel(title); t.setObjectName("metric_label"); lay.addWidget(t)
    def set_value(self, v): self.val.setText(str(v))

class SearchBar(QWidget):
    search_changed = pyqtSignal(str)
    def __init__(self, ph="Buscar...", parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        self.edit = QLineEdit(); self.edit.setPlaceholderText(f"  {ph}")
        self.edit.textChanged.connect(self.search_changed); lay.addWidget(self.edit)

class SectionTitle(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent); self.setObjectName("section_title")
        f = QFont(); f.setPointSize(13); f.setBold(True); self.setFont(f)

class SwatchBtn(QPushButton):
    def __init__(self, color, label, parent=None):
        super().__init__(label, parent); self.color = color; self._sel = False; self._paint()
    def set_selected(self, v): self._sel = v; self._paint()
    def _paint(self):
        brd = "3px solid #FFF" if self._sel else f"2px solid {self.color}"
        fw = "bold" if self._sel else "normal"
        self.setStyleSheet(f"QPushButton{{background:{self.color};color:#FFF;border:{brd};"
                           f"border-radius:8px;padding:6px 10px;font-size:12px;font-weight:{fw};"
                           f"text-shadow:0 1px 2px rgba(0,0,0,.7);}}"
                           f"QPushButton:hover{{border:3px solid #FFF;font-weight:bold;}}")


# ═════════════════════════════════════════════════════════════════════════════
#  GTIN LOOKUP WORKER (async thread)
# ═════════════════════════════════════════════════════════════════════════════

from PyQt5.QtCore import QThread

class GtinWorker(QThread):
    """Background thread to query GTIN/barcode APIs without freezing UI."""
    result_ready = pyqtSignal(dict)   # emitted with product data dict
    error        = pyqtSignal(str)    # emitted with error message

    # api_choice: "cosmos" | "openfoodfacts" | "upcitemdb" | "auto"
    def __init__(self, barcode, api_choice="auto", parent=None):
        super().__init__(parent)
        self.barcode    = barcode.strip()
        self.api_choice = api_choice  # which API(s) to use

    def run(self):
        if not REQUESTS_OK:
            self.error.emit("Biblioteca 'requests' nao instalada. Execute: pip install requests")
            return
        if len(self.barcode) < 8:
            self.error.emit("Codigo muito curto. Minimo 8 digitos.")
            return

        choice = self.api_choice
        data   = {}

        # ── API 1: Open Food Facts (free, sem auth) ──────────────
        if choice in ("auto", "openfoodfacts"):
          try:
            r = _requests.get(
                f"https://world.openfoodfacts.org/api/v0/product/{self.barcode}.json",
                headers={"User-Agent": "MotoPecasSistema/3.0"},
                timeout=6
            )
            if r.status_code == 200:
                j = r.json()
                if j.get("status") == 1:
                    p = j.get("product", {})
                    data["name"]     = p.get("product_name_pt") or p.get("product_name") or ""
                    data["brand"]    = p.get("brands","").split(",")[0].strip()
                    data["ncm"]      = ""
                    data["source"]   = "Open Food Facts"
                    if data["name"]:
                        self.result_ready.emit(data)
                        return
          except Exception:
            pass

        # ── API 2: UPC Item DB (free trial — funciona para produtos em geral) ─
        if choice in ("auto", "upcitemdb"):
          try:
            r = _requests.get(
                f"https://api.upcitemdb.com/prod/trial/lookup?upc={self.barcode}",
                headers={"User-Agent": "MotoPecasSistema/3.0"},
                timeout=6
            )
            if r.status_code == 200:
                j = r.json()
                items = j.get("items", [])
                if items:
                    it = items[0]
                    data["name"]   = it.get("title","")
                    data["brand"]  = it.get("brand","")
                    data["ncm"]    = ""
                    data["source"] = "UPC Item DB"
                    if data["name"]:
                        self.result_ready.emit(data)
                        return
          except Exception:
            pass

        # ── API 3: Bluesoft Cosmos (base brasileira — principal para peças BR) ──
        if choice in ("auto", "cosmos"):
          COSMOS_TOKEN = "2kte6zOrEl-HZlQ23RTpWQ"
          try:
            r = _requests.get(
                f"https://api.cosmos.bluesoft.com.br/gtins/{self.barcode}.json",
                headers={
                    "X-Cosmos-Token": COSMOS_TOKEN,
                    "Content-Type":   "application/json",
                    "User-Agent":     "MotoPecasSistema/3.0"
                },
                timeout=8
            )
            if r.status_code == 200:
                j = r.json()
                # Campo principal de descricao
                name = (j.get("description") or j.get("commercial_name") or "").strip()
                # Marca: pode ser objeto {"name":...} ou string
                brand_raw = j.get("brand") or j.get("manufacturer") or ""
                if isinstance(brand_raw, dict):
                    brand = brand_raw.get("name","").strip()
                else:
                    brand = str(brand_raw).strip()
                # NCM: pode ser objeto {"code":...} ou string
                ncm_raw = j.get("ncm") or {}
                if isinstance(ncm_raw, dict):
                    ncm = ncm_raw.get("code","").strip()
                else:
                    ncm = str(ncm_raw).strip()
                # Origem / classificacao adicional
                category = ""
                if j.get("gtins"):
                    gtin_info = j["gtins"][0] if isinstance(j["gtins"], list) else {}
                    category = gtin_info.get("commercial_name","")

                data["name"]     = name or category
                data["brand"]    = brand
                data["ncm"]      = ncm
                data["source"]   = "Bluesoft Cosmos"
                if data["name"]:
                    self.result_ready.emit(data)
                    return
            elif r.status_code == 401:
                self.error.emit("Token Bluesoft Cosmos invalido ou expirado. Verifique em cosmos.bluesoft.com.br")
                return
            elif r.status_code == 404:
                pass
          except Exception:
            pass

        # Nao encontrado em nenhuma API
        api_names = {
            "cosmos":        "Bluesoft Cosmos",
            "openfoodfacts": "Open Food Facts",
            "upcitemdb":     "UPC Item DB",
            "auto":          "Bluesoft Cosmos + Open Food Facts + UPC Item DB",
        }
        self.error.emit(
            f"Produto nao encontrado: {self.barcode}\n"
            f"API consultada: {api_names.get(choice, choice)}"
        )


# ═════════════════════════════════════════════════════════════════════════════
#  GOOGLE IMAGES WORKER + MODAL DE IMAGEM DO PRODUTO
# ═════════════════════════════════════════════════════════════════════════════

class ImageSearchWorker(QThread):
    """Busca imagem do produto com cadeia de fallback: Bing -> DuckDuckGo."""
    image_ready = pyqtSignal(bytes)
    error       = pyqtSignal(str)

    def __init__(self, query, parent=None):
        super().__init__(parent)
        self.query = query

    # ── helpers ────────────────────────────────────────────────────────────
    @staticmethod
    def _is_valid_image(data):
        hd = data[:12]
        return (hd[:2]  == b'\xff\xd8' or   # JPEG
                hd[:4]  == b'\x89PNG' or    # PNG
                hd[:4]  == b'RIFF' or       # WEBP
                hd[:6]  in (b'GIF87a', b'GIF89a'))  # GIF

    def _download_first(self, urls, headers):
        """Tenta baixar a primeira URL valida da lista."""
        for url in urls:
            try:
                r = _requests.get(url, headers=headers, timeout=8, allow_redirects=True)
                if r.status_code == 200 and len(r.content) > 800 and self._is_valid_image(r.content):
                    return r.content
            except Exception:
                continue
        return None

    # ── fonte 1: Bing Images ────────────────────────────────────────────────
    def _try_bing(self, query_encoded, headers):
        try:
            import re
            url  = f"https://www.bing.com/images/search?q={query_encoded}&form=HDRSC2&first=1"
            resp = _requests.get(url, headers=headers, timeout=12)
            # Bing armazena a URL original em 'murl'
            urls = re.findall(r'"murl":"(https?://[^"]+)"', resp.text)
            if not urls:
                # Formato alternativo (query string)
                import urllib.parse
                raw = re.findall(r'mediaurl=(https[^&"\s]+)', resp.text)
                urls = [urllib.parse.unquote(u) for u in raw]
            return list(dict.fromkeys(urls))
        except Exception:
            return []

    # ── fonte 2: DuckDuckGo Images ─────────────────────────────────────────
    def _try_duckduckgo(self, query_encoded, headers):
        try:
            import re
            url  = f"https://duckduckgo.com/?q={query_encoded}&iax=images&ia=images"
            resp = _requests.get(url, headers=headers, timeout=12)
            urls = re.findall(r'"u":"(https?://[^"]+)"', resp.text)
            if not urls:
                urls = re.findall(r'"(https?://[^"]+\.(?:jpg|jpeg|png|webp))"', resp.text)
            return list(dict.fromkeys(u for u in urls if len(u) > 40))
        except Exception:
            return []

    # ── run ────────────────────────────────────────────────────────────────
    def run(self):
        if not REQUESTS_OK:
            self.error.emit("Instale: pip install requests")
            return
        try:
            import urllib.parse
            q = urllib.parse.quote(self.query + " peca automotiva moto")
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            }

            # 1) Bing Images
            urls = self._try_bing(q, headers)
            if urls:
                data = self._download_first(urls[:15], headers)
                if data:
                    self.image_ready.emit(data); return

            # 2) DuckDuckGo Images
            urls = self._try_duckduckgo(q, headers)
            if urls:
                data = self._download_first(urls[:15], headers)
                if data:
                    self.image_ready.emit(data); return

            self.error.emit(
                "Nenhuma imagem encontrada.\n"
                "Verifique a conexao com a internet e tente novamente."
            )
        except Exception as e:
            self.error.emit(f"Erro ao buscar imagem: {str(e)[:120]}")


class ProductImageDialog(ModernFormDialog):
    """Modal que exibe imagem do produto buscada via Bing / DuckDuckGo."""

    def __init__(self, product_name, parent=None):
        super().__init__(parent, min_width=560)
        self.add_title_bar(f"Imagem \u2014 {product_name}")
        self.container.setFixedSize(560, 540)
        self._worker = None

        lay = self.main_lay

        # Titulo
        tit = QLabel(f"<b>{product_name}</b>")
        tit.setStyleSheet("font-size:14px;")
        tit.setWordWrap(True)
        lay.addWidget(tit)

        src = QLabel("  Fonte: Google Imagens")
        src.setObjectName("subtitle")
        lay.addWidget(src)

        # Area da imagem
        self.img_lbl = QLabel()
        self.img_lbl.setAlignment(Qt.AlignCenter)
        self.img_lbl.setFixedSize(500, 370)
        self.img_lbl.setStyleSheet(
            "border:2px solid #3A3A3A; border-radius:10px; background:#1A1A1A;"
        )
        self.img_lbl.setText("  Buscando no Google Imagens...")
        self.img_lbl.setFont(QFont("Segoe UI", 11))
        lay.addWidget(self.img_lbl)

        # Status
        self.status = QLabel("Conectando ao Google...")
        self.status.setAlignment(Qt.AlignCenter)
        self.status.setObjectName("subtitle")
        lay.addWidget(self.status)

        # Botao fechar
        btn = QPushButton("Fechar")
        btn.setObjectName("btn_secondary")
        btn.clicked.connect(self.accept)
        lay.addWidget(btn, alignment=Qt.AlignRight)

        # Inicia busca
        self._worker = ImageSearchWorker(product_name, self)
        self._worker.image_ready.connect(self._on_image)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_image(self, data):
        pix = QPixmap()
        if pix.loadFromData(data):
            scaled = pix.scaled(500, 370, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.img_lbl.setText("")
            self.img_lbl.setPixmap(scaled)
            self.status.setText("  Imagem carregada via Google Imagens")
            self.status.setStyleSheet("color:#00C853;")
        else:
            self._on_error("Formato de imagem invalido")

    def _on_error(self, msg):
        self.img_lbl.setText(f"  {msg}")
        self.img_lbl.setStyleSheet(
            "border:2px solid #FF1744; border-radius:10px; background:#1A1A1A;"
            "color:#FF6B35; padding:10px;"
        )
        self.status.setText("Falha ao carregar imagem")
        self.status.setStyleSheet("color:#FF1744;")

    def closeEvent(self, e):
        if self._worker and self._worker.isRunning():
            self._worker.quit()
            self._worker.wait(2000)
        super().closeEvent(e)

# ═════════════════════════════════════════════════════════════════════════════
#  PRICE RESEARCH WORKER (pesquisa precos regionais via Mercado Livre)
# ═════════════════════════════════════════════════════════════════════════════
class PriceResearchWorker(QThread):
    result_ready = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, query, parent=None):
        super().__init__(parent)
        self.query = query

    def run(self):
        if not REQUESTS_OK:
            self.error.emit("Biblioteca 'requests' nao instalada.")
            return
        try:
            import urllib.parse, re
            q = urllib.parse.quote(self.query.replace(" ", "-"))
            url = f"https://lista.mercadolivre.com.br/{q}"
            r = _requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
            if r.status_code == 200:
                prices_str = re.findall(r'"price":(\d+(?:\.\d+)?)', r.text)
                prices = [float(p) for p in prices_str if float(p) > 0]
                self.result_ready.emit({"prices": prices, "count": len(prices)})
            else:
                self.error.emit(f"Erro na API Mercado Livre: {r.status_code}")
        except Exception as e:
            self.error.emit(f"Erro: {str(e)[:100]}")


# ═════════════════════════════════════════════════════════════════════════════
#  DIALOGS — Product / Customer / User / Settings
# ═════════════════════════════════════════════════════════════════════════════
class ProductDialog(ModernFormDialog):
    def __init__(self, parent=None, product=None):
        super().__init__(parent, min_width=580)
        self.product = product
        self.add_title_bar("Novo Produto" if not product else "Editar Produto")
        screen = QApplication.primaryScreen().availableGeometry()
        max_h = min(820, int(screen.height() * 0.90))
        self.container.setMaximumHeight(max_h)

        main_lay = self.main_lay

        # ── Scroll area envolve todos os campos ───────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        scroll_content = QWidget()
        scroll_lay = QVBoxLayout(scroll_content)
        scroll_lay.setContentsMargins(0, 8, 8, 8)
        scroll_lay.setSpacing(10)
        scroll.setWidget(scroll_content)
        main_lay.addWidget(scroll, 1)

        # ── Grid de campos (dentro do scroll) ────────────────────
        g = QGridLayout()
        g.setSpacing(9)
        g.setColumnStretch(1, 1)
        scroll_lay.addLayout(g)

        # ─ Codigo ─
        self.code = QLineEdit()
        self.code.setPlaceholderText("Codigo interno ou EAN — leia o codigo de barras aqui")
        self.code.setMinimumHeight(36)
        self.code.returnPressed.connect(self._lookup_gtin)
        g.addWidget(QLabel("Codigo *"), 0, 0)
        g.addWidget(self.code, 0, 1)

        # ─ API selector (linha abaixo do campo Codigo) ─
        api_widget = QWidget()
        api_inner = QHBoxLayout(api_widget)
        api_inner.setContentsMargins(0, 0, 0, 0)
        api_inner.setSpacing(6)

        self.api_cb = QComboBox()
        self.api_cb.setMinimumHeight(30)
        self.api_cb.addItem("Sem busca por API", "none")
        self.api_cb.addItem("Buscar por Todas", "auto")
        self.api_cb.addItem("Bluesoft Cosmos (BR)", "cosmos")
        self.api_cb.addItem("Open Food Facts", "openfoodfacts")
        self.api_cb.addItem("UPC Item DB", "upcitemdb")
        self.api_cb.setCurrentIndex(0)
        self.api_cb.currentIndexChanged.connect(self._on_api_changed)
        api_inner.addWidget(self.api_cb, 1)

        self.btn_lookup = QPushButton("Consultar GTIN")
        self.btn_lookup.setMinimumHeight(30)
        self.btn_lookup.setAutoDefault(False)
        self.btn_lookup.setDefault(False)
        self.btn_lookup.setEnabled(False)  # disabled by default (none selected)
        self.btn_lookup.clicked.connect(self._lookup_gtin)
        api_inner.addWidget(self.btn_lookup)

        g.addWidget(QLabel("API GTIN"), 1, 0)
        g.addWidget(api_widget, 1, 1)

        # ─ Status da consulta GTIN ─
        self.gtin_status = QLabel("")
        self.gtin_status.setMinimumHeight(22)
        self.gtin_status.setWordWrap(True)
        self.gtin_status.setStyleSheet("font-size:12px; padding:2px 6px; border-radius:4px;")
        g.addWidget(QLabel(""), 2, 0)
        g.addWidget(self.gtin_status, 2, 1)

        # ─ Demais campos ─
        self.name = QLineEdit()
        self.name.returnPressed.connect(lambda: self.cat.setFocus())
        self.cat = QComboBox(); self.cat.addItems(CATEGORIES)
        self.brand = QLineEdit()
        self.brand.returnPressed.connect(lambda: self.unit.setFocus())
        self.unit = QComboBox(); self.unit.addItems(UNITS)
        self.cost = QDoubleSpinBox()
        self.cost.setRange(0, 999999); self.cost.setPrefix("R$ "); self.cost.setDecimals(2)
        self.price = QDoubleSpinBox()
        self.price.setRange(0, 999999); self.price.setPrefix("R$ "); self.price.setDecimals(2)
        self.stock = QSpinBox(); self.stock.setRange(0, 99999)
        self.minst = QSpinBox(); self.minst.setRange(0, 99999)
        self.desc = QTextEdit(); self.desc.setMaximumHeight(58)
        self.ncm = QLineEdit(); self.ncm.setPlaceholderText("Ex: 87089990 (8 digitos)")
        self.cfop = QLineEdit("5102"); self.cfop.setPlaceholderText("Ex: 5102")
        self.csosn = QComboBox()
        for v, t in [("400","400 - Nao tributada (Simples)"), ("102","102 - Permissao ICMS SN"),
                     ("103","103 - Isencao SN"), ("300","300 - Imune"), ("500","500 - ICMS cobrado")]:
            self.csosn.addItem(t, v)
        self.orig = QComboBox()
        for v, t in [("0","0 - Nacional"), ("1","1 - Estrangeira Importada"),
                     ("2","2 - Estrangeira Adq. Interior"), ("3","3 - Nacional 40-70% conteudo")]:
            self.orig.addItem(t, v)

        other_rows = [
            ("Nome *",          self.name),
            ("Categoria",       self.cat),
            ("Marca",           self.brand),
            ("Unidade",         self.unit),
            ("Custo",           self.cost),
            ("Preco Venda *",   self.price),
            ("Estoque",         self.stock),
            ("Minimo",          self.minst),
            ("Descricao",       self.desc),
            ("NCM (Fiscal)",    self.ncm),
            ("CFOP (Fiscal)",   self.cfop),
            ("CSOSN (Fiscal)",  self.csosn),
            ("Origem (Fiscal)", self.orig),
        ]
        for offset, (lbl, w) in enumerate(other_rows, start=3):
            g.addWidget(QLabel(lbl), offset, 0)
            g.addWidget(w, offset, 1)

        # ─ Markup 30% checkbox ─
        markup_row = 3 + len(other_rows)
        self.chk_markup = QCheckBox("  Preco Venda = Custo + 30%")
        self.chk_markup.setToolTip("Se marcado, preco de venda sera automaticamente custo × 1.30")
        self.chk_markup.stateChanged.connect(self._on_markup_changed)
        self.cost.valueChanged.connect(self._on_cost_changed)
        g.addWidget(QLabel("Markup"), markup_row, 0)
        g.addWidget(self.chk_markup, markup_row, 1)

        # ─ Pesquisa de precos regionais ─
        self.btn_price_research = QPushButton("  Pesquisar Precos na Regiao")
        self.btn_price_research.setObjectName("btn_secondary")
        self.btn_price_research.setAutoDefault(False)
        self.btn_price_research.setDefault(False)
        self.btn_price_research.clicked.connect(self._research_price)
        g.addWidget(QLabel("Referencia"), markup_row + 1, 0)
        g.addWidget(self.btn_price_research, markup_row + 1, 1)

        self.price_ref_lbl = QLabel("")
        self.price_ref_lbl.setWordWrap(True)
        self.price_ref_lbl.setStyleSheet("font-size:11px;padding:4px 8px;")
        g.addWidget(QLabel(""), markup_row + 2, 0)
        g.addWidget(self.price_ref_lbl, markup_row + 2, 1)

        scroll_lay.addStretch()

        # ─ Populate when editing ─
        if product:
            self.code.setText(product.get("code", "") or product.get("barcode", ""))
            self.name.setText(product.get("name", ""))
            i = self.cat.findText(product.get("category", ""))
            if i >= 0: self.cat.setCurrentIndex(i)
            self.brand.setText(product.get("brand", ""))
            i = self.unit.findText(product.get("unit", "UN"))
            if i >= 0: self.unit.setCurrentIndex(i)
            self.cost.setValue(float(product.get("cost_price", 0)))
            self.price.setValue(float(product.get("sale_price", 0)))
            self.stock.setValue(int(product.get("stock", 0)))
            self.minst.setValue(int(product.get("min_stock", 0)))
            self.desc.setPlainText(product.get("description", ""))
            self.ncm.setText(product.get("ncm", ""))
            self.cfop.setText(product.get("cfop", "5102"))
            ix = self.csosn.findData(product.get("csosn", "400"))
            if ix >= 0: self.csosn.setCurrentIndex(ix)
            ix = self.orig.findData(product.get("orig", "0"))
            if ix >= 0: self.orig.setCurrentIndex(ix)

        # ─ Buttons fixos na parte inferior (fora do scroll) ─
        footer = QWidget()
        footer.setStyleSheet("border-top: 1px solid #E2E8F0;")
        footer_lay = QHBoxLayout(footer)
        footer_lay.setContentsMargins(0, 10, 0, 0)
        footer_lay.addStretch()
        cn = QPushButton("Cancelar")
        cn.setObjectName("btn_secondary")
        cn.setAutoDefault(False); cn.setDefault(False)
        cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Produto")
        sv.setMinimumWidth(140)
        sv.setAutoDefault(False); sv.setDefault(False)
        sv.setObjectName("btn_success")
        sv.clicked.connect(self._save)
        footer_lay.addWidget(cn); footer_lay.addWidget(sv)
        main_lay.addWidget(footer)

        # alias kept for compatibility
        self.barcode_e = self.code

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            focused = self.focusWidget()
            if isinstance(focused, QLineEdit):
                event.ignore()
                return
        super().keyPressEvent(event)

    def _on_api_changed(self):
        choice = self.api_cb.currentData()
        self.btn_lookup.setEnabled(choice != "none")

    def _set_gtin_status(self, msg, ok=True, loading=False):
        if loading:
            s = "font-size:12px;padding:3px 8px;border-radius:4px;background:#1A2A3A;color:#5599FF;border:1px solid #5599FF;"
        elif ok:
            s = "font-size:12px;padding:3px 8px;border-radius:4px;background:#003300;color:#00C853;border:1px solid #00C853;"
        else:
            s = "font-size:12px;padding:3px 8px;border-radius:4px;background:#330008;color:#FF1744;border:1px solid #FF1744;"
        self.gtin_status.setText(msg)
        self.gtin_status.setStyleSheet(s)

    def _lookup_gtin(self):
        choice = self.api_cb.currentData()
        if choice == "none":
            self.name.setFocus()
            return

        bc = self.code.text().strip()
        if not bc:
            self._set_gtin_status("Digite ou leia um codigo de barras no campo Codigo.", False)
            return
        if len(bc) < 8:
            self._set_gtin_status("Codigo muito curto — minimo 8 digitos para consulta GTIN.", False)
            self.name.setFocus()
            return

        api_label = self.api_cb.currentText()
        self._set_gtin_status(f"Consultando {api_label} para: {bc}...", loading=True)
        self.btn_lookup.setEnabled(False)
        self.btn_lookup.setText("Consultando...")

        if self._gtin_worker and self._gtin_worker.isRunning():
            self._gtin_worker.quit()
            self._gtin_worker.wait()

        self._gtin_worker = GtinWorker(bc, choice, self)
        self._gtin_worker.result_ready.connect(self._on_gtin_result)
        self._gtin_worker.error.connect(self._on_gtin_error)
        self._gtin_worker.finished.connect(self._on_gtin_done)
        self._gtin_worker.start()

    def _on_gtin_result(self, data):
        name  = data.get("name", "").strip()
        brand = data.get("brand", "").strip()
        ncm   = data.get("ncm", "").strip()
        src   = data.get("source", "")
        if name:
            self.name.blockSignals(True)
            self.name.setText(name)
            self.name.blockSignals(False)
        if brand:
            self.brand.setText(brand)
        if ncm:
            self.ncm.setText(ncm)
        preview = name[:38] + ("..." if len(name) > 38 else "")
        self._set_gtin_status(f"Encontrado via {src}: {preview}", ok=True)

    def _on_gtin_error(self, msg):
        self._set_gtin_status(msg.splitlines()[0], ok=False)

    def _on_gtin_done(self):
        choice = self.api_cb.currentData()
        self.btn_lookup.setEnabled(choice != "none")
        self.btn_lookup.setText("Consultar GTIN")

    def _generate_code_from_name(self, text):
        if not text.strip(): return ""
        import unicodedata, re as _re
        nfkd  = unicodedata.normalize("NFKD", text.upper())
        clean = "".join(c for c in nfkd if not unicodedata.combining(c))
        clean = _re.sub(r"[^A-Z0-9 ]", "", clean).strip()
        words = clean.split()
        if not words: return ""
        if len(words) == 1:   prefix = words[0][:4]
        elif len(words) == 2: prefix = words[0][:3] + words[1][:1]
        else:                 prefix = words[0][:2] + words[1][:1] + words[2][:1]
        prefix = prefix[:4].ljust(3, "X")
        dm = getattr(self.parent(), "dm", None)
        existing = {p.get("code","") for p in dm.get_products()} if dm else set()
        num = 1
        while True:
            cand = f"{prefix}{num:03d}"
            if cand not in existing: break
            num += 1
        return cand

    def _auto_code(self, text): pass  # kept for compatibility
    def _update_api_info(self): pass  # kept for compatibility

    def _on_markup_changed(self, state):
        if state == Qt.Checked and self.cost.value() > 0:
            self.price.setValue(round(self.cost.value() * 1.30, 2))

    def _on_cost_changed(self, val):
        if self.chk_markup.isChecked() and val > 0:
            self.price.setValue(round(val * 1.30, 2))

    def _research_price(self):
        name = self.name.text().strip()
        brand = self.brand.text().strip()
        if not name:
            QMessageBox.warning(self, "Atencao", "Informe o nome do produto primeiro.")
            return
        self.btn_price_research.setEnabled(False)
        self.btn_price_research.setText("Pesquisando...")
        self.price_ref_lbl.setText("Buscando precos no Mercado Livre...")
        self.price_ref_lbl.setStyleSheet("font-size:11px;padding:4px 8px;color:#5599FF;background:#1A2A3A;border-radius:4px;")
        self._price_worker = PriceResearchWorker(f"{name} {brand}".strip(), self)
        self._price_worker.result_ready.connect(self._on_price_result)
        self._price_worker.error.connect(self._on_price_error)
        self._price_worker.start()

    def _on_price_result(self, data):
        self.btn_price_research.setEnabled(True)
        self.btn_price_research.setText("  Pesquisar Precos na Regiao")
        prices = data.get("prices", [])
        if prices:
            avg = sum(prices) / len(prices)
            mn = min(prices)
            mx = max(prices)
            txt = f"Mercado Livre ({len(prices)} resultados):\n"
            txt += f"  Min: R$ {mn:.2f}  |  Med: R$ {avg:.2f}  |  Max: R$ {mx:.2f}"
            self.price_ref_lbl.setText(txt)
            self.price_ref_lbl.setStyleSheet("font-size:11px;padding:6px 8px;color:#00C853;background:#003300;border:1px solid #00C853;border-radius:4px;")
        else:
            self.price_ref_lbl.setText("Nenhum resultado encontrado.")
            self.price_ref_lbl.setStyleSheet("font-size:11px;padding:4px 8px;color:#FF6B35;")

    def _on_price_error(self, msg):
        self.btn_price_research.setEnabled(True)
        self.btn_price_research.setText("  Pesquisar Precos na Regiao")
        self.price_ref_lbl.setText(msg)
        self.price_ref_lbl.setStyleSheet("font-size:11px;padding:4px 8px;color:#FF1744;background:#330008;border-radius:4px;")

    def _save(self):
        if not self.code.text().strip():
            gen = self._generate_code_from_name(self.name.text().strip())
            if gen:
                self.code.setText(gen)
            else:
                QMessageBox.warning(self, "Atencao", "Informe o codigo do produto.")
                return
        if not self.name.text().strip():
            QMessageBox.warning(self, "Atencao", "Informe o nome do produto.")
            return
        if self.price.value() <= 0:
            QMessageBox.warning(self, "Atencao", "Informe o preco de venda.")
            return
        self.result_data = {
            "id":           self.product["id"] if self.product else "",
            "code":         self.code.text().strip().upper(),
            "barcode":      self.code.text().strip(),
            "name":         self.name.text().strip(),
            "category":     self.cat.currentText(),
            "brand":        self.brand.text().strip(),
            "unit":         self.unit.currentText(),
            "cost_price":   self.cost.value(),
            "sale_price":   self.price.value(),
            "stock":        self.stock.value(),
            "min_stock":    self.minst.value(),
            "description":  self.desc.toPlainText().strip(),
            "ncm":          self.ncm.text().strip(),
            "cfop":         self.cfop.text().strip(),
            "csosn":        self.csosn.currentData(),
            "orig":         self.orig.currentData(),
        }
        self.accept()


class CustomerDialog(ModernFormDialog):
    def __init__(self, parent=None, customer=None):
        super().__init__(parent, min_width=460); self.customer = customer
        self.add_title_bar("Novo Cliente" if not customer else "Editar Cliente")
        lay = self.main_lay
        g = QGridLayout(); g.setSpacing(9)
        self.nm  = QLineEdit(); self.nm.setPlaceholderText("Nome completo ou razao social")
        self.ph  = QLineEdit(); self.ph.setPlaceholderText("(DD) XXXXX-XXXX")
        self.ad  = QLineEdit(); self.ad.setPlaceholderText("Endereco completo")
        self.vei = QLineEdit(); self.vei.setPlaceholderText("Ex: Honda CG 160, Yamaha Factor 150...")
        self.pla = QLineEdit(); self.pla.setPlaceholderText("Ex: ABC-1234 ou ABC1D23")
        self.pla.setMaxLength(8)
        rows = [("Nome / Razao Social *", self.nm),
                ("Telefone / WhatsApp",   self.ph),
                ("Endereco",              self.ad),
                ("Veiculo",               self.vei),
                ("Placa",                 self.pla)]
        for r,(l,w) in enumerate(rows): g.addWidget(QLabel(l),r,0); g.addWidget(w,r,1)
        g.setColumnStretch(1,1); lay.addLayout(g); lay.addSpacing(8)
        if customer:
            self.nm.setText(customer.get("name",""))
            self.ph.setText(customer.get("phone",""))
            self.ad.setText(customer.get("address",""))
            self.vei.setText(customer.get("veiculo",""))
            self.pla.setText(customer.get("placa",""))
        b = QHBoxLayout(); b.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Cliente"); sv.setObjectName("btn_success"); sv.clicked.connect(self._save)
        b.addWidget(cn); b.addWidget(sv); lay.addLayout(b)
    def _save(self):
        if not self.nm.text().strip(): QMessageBox.warning(self,"Atencao","Informe o nome."); return
        self.result_data = {"id": self.customer["id"] if self.customer else "",
            "name":    self.nm.text().strip(),
            "phone":   self.ph.text().strip(),
            "address": self.ad.text().strip(),
            "veiculo": self.vei.text().strip(),
            "placa":   self.pla.text().strip().upper()}
        # Auto-salvar contato no WhatsApp
        phone = self.ph.text().strip()
        name = self.nm.text().strip()
        if phone and not self.customer:  # Apenas para novos clientes
            digits = "".join(filter(str.isdigit, phone))
            if len(digits) in [10, 11]: digits = "55" + digits
            try:
                msg = f"Contato salvo: {name}"
                encoded = urllib.parse.quote(msg)
                webbrowser.open(f"https://wa.me/{digits}?text={encoded}")
            except: pass
        self.accept()

class UserDialog(ModernFormDialog):
    def __init__(self, parent=None, user=None):
        super().__init__(parent, min_width=450); self.user = user
        self.add_title_bar("Novo Usuario" if not user else "Editar Usuario")
        lay = self.main_lay
        info = QLabel("  Operador: Produtos, Nova Venda, Historico e Clientes.  "
                      "   Administrador: acesso completo a todos os modulos.")
        info.setWordWrap(True); info.setObjectName("subtitle")
        info.setStyleSheet("font-size:12px;padding:8px 4px;background:transparent;"); lay.addWidget(info)
        g = QGridLayout(); g.setSpacing(10)
        self.nm = QLineEdit(); self.nm.setPlaceholderText("Nome completo")
        self.us = QLineEdit(); self.us.setPlaceholderText("Login de acesso (sem espacos)")
        self.pw = QLineEdit(); self.pw.setEchoMode(QLineEdit.Password); self.pw.setPlaceholderText("Senha")
        self.p2 = QLineEdit(); self.p2.setEchoMode(QLineEdit.Password); self.p2.setPlaceholderText("Confirme a senha")
        self.rl = QComboBox(); self.rl.addItems(["  Operador (acesso limitado)","  Administrador (acesso total)"])
        self.can_edit_prod = QCheckBox("  Permitir que este Operador edite/exclua produtos")
        self.can_edit_prod.setToolTip("Se desmarcado, o operador so pode visualizar produtos")
        rows = [("Nome Completo *",self.nm),("Login / Usuario *",self.us),
                ("Senha *",self.pw),("Confirmar Senha",self.p2),("Perfil de Acesso",self.rl)]
        for r,(l,w) in enumerate(rows): g.addWidget(QLabel(l),r,0); g.addWidget(w,r,1)
        g.addWidget(QLabel(""), len(rows), 0); g.addWidget(self.can_edit_prod, len(rows), 1)
        g.setColumnStretch(1,1); lay.addLayout(g); lay.addSpacing(8)
        if user:
            self.nm.setText(user.get("name","")); self.us.setText(user.get("username",""))
            self.us.setEnabled(False)
            self.pw.setPlaceholderText("Deixe em branco para manter a senha")
            self.p2.setPlaceholderText("Confirme apenas se alterar")
            self.rl.setCurrentIndex(1 if user.get("role")=="admin" else 0)
            self.can_edit_prod.setChecked(user.get("can_edit_products", False))
        b = QHBoxLayout(); b.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Usuario"); sv.setObjectName("btn_success"); sv.clicked.connect(self._save)
        b.addWidget(cn); b.addWidget(sv); lay.addLayout(b)
    def _save(self):
        nm = self.nm.text().strip(); us = self.us.text().strip().lower().replace(" ","")
        pw = self.pw.text(); p2 = self.p2.text()
        if not nm: QMessageBox.warning(self,"Atencao","Informe o nome."); return
        if not us: QMessageBox.warning(self,"Atencao","Informe o login."); return
        if not self.user and not pw: QMessageBox.warning(self,"Atencao","Informe a senha."); return
        if pw and pw != p2: QMessageBox.warning(self,"Atencao","As senhas nao coincidem."); return
        self.result_data = {"id": self.user["id"] if self.user else "", "name": nm, "username": us,
            "password": pw if pw else (self.user.get("password","") if self.user else ""),
            "role": "admin" if self.rl.currentIndex()==1 else "operator",
            "can_edit_products": self.can_edit_prod.isChecked(), "active": True}
        self.accept()

class EmpresaDialog(ModernFormDialog):
    """Company fiscal data configuration dialog."""
    REGIMES = [("1","Simples Nacional"),("2","Simples Nacional - Excesso"),("3","Regime Normal")]
    UFS = ["AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT","PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO"]

    def __init__(self, parent, dm):
        super().__init__(parent, min_width=590); self.dm=dm
        self.add_title_bar("Configuracao da Empresa / Emitente")
        lay = self.main_lay
        emp = dm.get_empresa()

        tabs = QTabWidget()

        # Tab 1: Empresa
        t1 = QWidget(); f1 = QFormLayout(t1); f1.setSpacing(10)
        self.razao = QLineEdit(emp.get("razao_social","")); self.razao.setMaxLength(60)
        self.fantasia = QLineEdit(emp.get("nome_fantasia","")); self.fantasia.setMaxLength(60)
        self.cnpj = QLineEdit(emp.get("cnpj","")); self.cnpj.setPlaceholderText("00.000.000/0001-00")
        self.ie = QLineEdit(emp.get("ie","")); self.ie.setPlaceholderText("Inscricao Estadual")
        self.im = QLineEdit(emp.get("im","")); self.im.setPlaceholderText("Inscricao Municipal (opcional)")
        self.cnae = QLineEdit(emp.get("cnae","")); self.cnae.setPlaceholderText("Ex: 4530-7/03")
        self.regime_cb = QComboBox()
        for code,desc in self.REGIMES: self.regime_cb.addItem(f"{code} - {desc}", code)
        cur_reg = emp.get("regime","1")
        for i in range(self.regime_cb.count()):
            if self.regime_cb.itemData(i) == cur_reg: self.regime_cb.setCurrentIndex(i); break
        for lbl,w in [("Razao Social *",self.razao),("Nome Fantasia",self.fantasia),
                      ("CNPJ *",self.cnpj),("Insc. Estadual",self.ie),
                      ("Insc. Municipal",self.im),("CNAE",self.cnae),("Regime Tributario",self.regime_cb)]:
            f1.addRow(lbl, w)
        tabs.addTab(t1, "  Empresa")

        # Tab 2: Endereco
        t2 = QWidget(); f2 = QFormLayout(t2); f2.setSpacing(10)
        self.end = QLineEdit(emp.get("endereco",""))
        self.bairro = QLineEdit(emp.get("bairro",""))
        self.mun = QLineEdit(emp.get("municipio",""))
        self.uf_cb = QComboBox(); self.uf_cb.addItems(self.UFS)
        uf_idx = self.uf_cb.findText(emp.get("uf","SP"))
        if uf_idx>=0: self.uf_cb.setCurrentIndex(uf_idx)
        self.cep = QLineEdit(emp.get("cep","")); self.cep.setPlaceholderText("00000-000")
        self.tel = QLineEdit(emp.get("telefone",""))
        self.email = QLineEdit(emp.get("email",""))
        for lbl,w in [("Logradouro",self.end),("Bairro",self.bairro),("Municipio",self.mun),
                      ("UF",self.uf_cb),("CEP",self.cep),("Telefone",self.tel),("E-mail",self.email)]:
            f2.addRow(lbl, w)
        tabs.addTab(t2, "  Endereco")

        # Tab 3: NFC-e
        t3 = QWidget(); f3 = QFormLayout(t3); f3.setSpacing(10)
        nfce = dm.get_nfce()
        self.serie = QLineEdit(nfce.get("serie","001")); self.serie.setPlaceholderText("001")
        self.prox_nf = QSpinBox(); self.prox_nf.setRange(1,999999999); self.prox_nf.setValue(int(nfce.get("proxima_nf",1)))
        self.amb_cb = QComboBox()
        self.amb_cb.addItem("2 - Homologacao (Testes)", "2")
        self.amb_cb.addItem("1 - Producao (Real)", "1")
        amb_idx = 1 if nfce.get("ambiente","2")=="1" else 0
        self.amb_cb.setCurrentIndex(amb_idx)
        self.cuf = QLineEdit(nfce.get("cuf","35")); self.cuf.setPlaceholderText("Ex: 35=SP, 33=RJ")
        self.cmun = QLineEdit(nfce.get("cmun","3550308")); self.cmun.setPlaceholderText("Codigo IBGE municipio")
        self.csc_id = QLineEdit(nfce.get("csc_id","000001"))
        self.csc_token = QLineEdit(nfce.get("csc_token","")); self.csc_token.setPlaceholderText("Token CSC da SEFAZ")
        info_csc = QLabel("O CSC (Codigo de Seguranca do Contribuinte) e obtido junto a SEFAZ do seu estado.")
        info_csc.setObjectName("subtitle"); info_csc.setWordWrap(True)
        for lbl,w in [("Serie NFC-e",self.serie),("Proxima NF",self.prox_nf),
                      ("Ambiente",self.amb_cb),("Codigo UF (cUF)",self.cuf),
                      ("Cod. Municipio",self.cmun),("CSC ID",self.csc_id),("CSC Token",self.csc_token)]:
            f3.addRow(lbl, w)
        f3.addRow("", info_csc)
        tabs.addTab(t3, "  NFC-e")

        lay.addWidget(tabs)
        btns = QHBoxLayout(); btns.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Configuracoes"); sv.setObjectName("btn_success"); sv.clicked.connect(self._save)
        btns.addWidget(cn); btns.addWidget(sv); lay.addLayout(btns)

    def _save(self):
        if not self.cnpj.text().strip():
            QMessageBox.warning(self,"Atencao","Informe o CNPJ da empresa."); return
        emp = {
            "razao_social": self.razao.text().strip(),
            "nome_fantasia": self.fantasia.text().strip(),
            "cnpj": self.cnpj.text().strip(),
            "ie": self.ie.text().strip(),
            "im": self.im.text().strip(),
            "cnae": self.cnae.text().strip(),
            "regime": self.regime_cb.currentData(),
            "endereco": self.end.text().strip(),
            "bairro": self.bairro.text().strip(),
            "municipio": self.mun.text().strip(),
            "uf": self.uf_cb.currentText(),
            "cep": self.cep.text().strip(),
            "telefone": self.tel.text().strip(),
            "email": self.email.text().strip(),
        }
        nfce = {
            "serie": self.serie.text().strip().zfill(3),
            "proxima_nf": self.prox_nf.value(),
            "ambiente": self.amb_cb.currentData(),
            "cuf": self.cuf.text().strip(),
            "cmun": self.cmun.text().strip(),
            "csc_id": self.csc_id.text().strip(),
            "csc_token": self.csc_token.text().strip(),
        }
        self.dm.save_empresa(emp)
        self.dm.save_nfce(nfce)
        self.accept()


class SettingsDialog(ModernFormDialog):
    def __init__(self, parent, dm):
        super().__init__(parent, min_width=660); self.dm = dm; self.settings = dm.get_settings().copy()
        self._sel = self.settings.get("accent_color","#FF6B35")
        self.add_title_bar("Configuracoes")
        
        screen = QApplication.primaryScreen().availableGeometry()
        max_h = min(820, int(screen.height() * 0.90))
        self.container.setMaximumHeight(max_h)

        outer = self.main_lay

        # ScrollArea para o conteudo responsivo
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")

        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setContentsMargins(24, 24, 24, 12)
        lay.setSpacing(14)
        scroll.setWidget(content)
        outer.addWidget(scroll)

        # Linha de botoes fixos na base
        btn_bar = QFrame()
        btn_bar.setFixedHeight(60)
        btn_bar.setStyleSheet("border-top:1px solid #E2E8F0;")
        btn_l = QHBoxLayout(btn_bar)
        btn_l.setContentsMargins(0, 10, 0, 10)
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        ap = QPushButton("  Aplicar Configuracoes"); ap.setObjectName("btn_success"); ap.clicked.connect(self._apply)
        btn_l.addStretch(); btn_l.addWidget(cn); btn_l.addWidget(ap)
        outer.addWidget(btn_bar)

        lay.addWidget(SectionTitle("  Aparencia Oculta"))

        # THEME
        tg = QGroupBox("  Tema da Interface"); tl = QHBoxLayout(tg); tl.setSpacing(20)
        self.rdark = QRadioButton("  Modo Escuro (Dark)")
        self.rlight = QRadioButton("  Modo Claro (Light)")
        self.rdark.setChecked(self.settings.get("theme","dark") == "dark")
        self.rlight.setChecked(self.settings.get("theme","dark") == "light")
        tl.addWidget(self.rdark); tl.addWidget(self.rlight); tl.addStretch(); lay.addWidget(tg)

        # ACCENT COLOR
        cg = QGroupBox("  Cor Principal do Sistema"); cl = QVBoxLayout(cg); cl.setSpacing(10)
        cl.addWidget(QLabel("Escolha uma cor predefinida ou use o seletor personalizado:"))
        gw = QWidget(); gg = QGridLayout(gw); gg.setSpacing(8); gg.setContentsMargins(0,0,0,0)
        self._swatches = []
        for i,(color,label) in enumerate(PRESET_COLORS):
            btn = SwatchBtn(color, label)
            btn.clicked.connect(lambda _,c=color,b=btn: self._pick_preset(c,b))
            btn.set_selected(color.lower() == self._sel.lower())
            gg.addWidget(btn, i//2, i%2); self._swatches.append((btn,color))
        cl.addWidget(gw)
        cr = QHBoxLayout()
        self.cprev = QPushButton(); self.cprev.setFixedSize(40,40)
        self.cprev.setToolTip("Clique para cor personalizada")
        self.cprev.clicked.connect(self._pick_custom); self._paint_prev(); cr.addWidget(self.cprev)
        cr.addWidget(QLabel("  Cor personalizada (clique no quadrado)", objectName="subtitle"))
        cr.addStretch()
        self.hex_lbl = QLabel(self._sel.upper())
        self.hex_lbl.setStyleSheet("font-family:monospace;font-size:13px;"); cr.addWidget(self.hex_lbl)
        cl.addLayout(cr); lay.addWidget(cg)

        # ICONE DO SISTEMA
        icon_g = QGroupBox("  Icone do Sistema (janela e barra de tarefas)"); il = QVBoxLayout(icon_g); il.setSpacing(8)
        icon_row = QHBoxLayout()
        self.icon_prev = QLabel(); self.icon_prev.setFixedSize(64,64); self.icon_prev.setAlignment(Qt.AlignCenter)
        self.icon_prev.setStyleSheet("border:1px dashed #666;border-radius:8px;")
        cur_icon = self.settings.get("app_icon","")
        if cur_icon and os.path.exists(cur_icon):
            self.icon_prev.setPixmap(QPixmap(cur_icon).scaled(56,56,Qt.KeepAspectRatio,Qt.SmoothTransformation))
        else: self.icon_prev.setText("Sem\nicone")
        icon_row.addWidget(self.icon_prev)
        icon_btns = QVBoxLayout(); icon_btns.setSpacing(6)
        bic = QPushButton("  Escolher Icone (ICO, PNG)"); bic.clicked.connect(self._choose_icon)
        bicr = QPushButton("  Remover Icone"); bicr.setObjectName("btn_secondary"); bicr.clicked.connect(self._clear_icon)
        icon_btns.addWidget(bic); icon_btns.addWidget(bicr); icon_btns.addStretch()
        icon_row.addLayout(icon_btns); icon_row.addStretch()
        il.addLayout(icon_row)
        il.addWidget(QLabel("Recomendado: arquivo .ICO 256x256 ou PNG 256x256", objectName="subtitle"))
        lay.addWidget(icon_g)

        # BACKGROUND
        bg_g = QGroupBox("  Papel de Parede"); bl = QVBoxLayout(bg_g); bl.setSpacing(8)
        self.bgprev = QLabel(); self.bgprev.setFixedHeight(100); self.bgprev.setAlignment(Qt.AlignCenter)
        self.bgprev.setStyleSheet("border:1px dashed #666;border-radius:6px;")
        cur = self.settings.get("background_image","")
        if cur and os.path.exists(cur):
            self.bgprev.setPixmap(QPixmap(cur).scaled(440,100,Qt.KeepAspectRatio,Qt.SmoothTransformation))
        else: self.bgprev.setText("Nenhuma imagem selecionada")
        bl.addWidget(self.bgprev)
        br = QHBoxLayout()
        bc = QPushButton("  Escolher Imagem"); bc.clicked.connect(self._choose_bg)
        bclr = QPushButton("  Remover"); bclr.setObjectName("btn_secondary"); bclr.clicked.connect(self._clear_bg)
        br.addWidget(bc); br.addWidget(bclr); br.addStretch(); bl.addLayout(br)
        bl.addWidget(QLabel("Formatos: JPG, PNG, BMP, WEBP", objectName="subtitle")); lay.addWidget(bg_g)

        # COMPORTAMENTO
        comp_g = QGroupBox("  Comportamento do Sistema"); comp_l = QVBoxLayout(comp_g)
        self.chk_fs = QCheckBox("  Iniciar sistema em Tela Inteira (Fullscreen)")
        self.chk_fs.setChecked(self.settings.get("fullscreen", False))
        comp_l.addWidget(self.chk_fs)
        lay.addWidget(comp_g)

        # NOTA FISCAL SIMPLES CONFIG
        nf_g = QGroupBox("  Configuracao — Nota Fiscal Simples (PDF)")
        nf_g.setCheckable(False)
        nf_l = QFormLayout(nf_g)
        nf_l.setSpacing(10)
        nf_l.setLabelAlignment(Qt.AlignLeft)
        nf_l.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        cur_nf = dm.get_nota_fiscal_cfg()
        self.nf_nome   = QLineEdit(cur_nf.get("nome_loja", "")); self.nf_nome.setPlaceholderText("Nome da loja exibido no cabecalho da nota")
        self.nf_slogan = QLineEdit(cur_nf.get("slogan", ""));    self.nf_slogan.setPlaceholderText("Ex: Sistema de Gestao e Vendas")
        self.nf_end    = QLineEdit(cur_nf.get("endereco", ""));   self.nf_end.setPlaceholderText("Endereco completo")
        self.nf_tel    = QLineEdit(cur_nf.get("telefone", ""));   self.nf_tel.setPlaceholderText("Telefone de contato")
        self.nf_cnpj   = QLineEdit(cur_nf.get("cnpj", ""));      self.nf_cnpj.setPlaceholderText("CNPJ (opcional)")
        self.nf_rodape = QLineEdit(cur_nf.get("rodape", ""));    self.nf_rodape.setPlaceholderText("Ex: Obrigado pela preferencia!")
        for lbl, w in [("Nome da Loja *",  self.nf_nome),
                       ("Slogan",           self.nf_slogan),
                       ("Endereco",         self.nf_end),
                       ("Telefone",         self.nf_tel),
                       ("CNPJ",             self.nf_cnpj),
                       ("Rodape",           self.nf_rodape)]:
            nf_l.addRow(lbl, w)
        lay.addWidget(nf_g)

        # WHATSAPP CONFIG
        wpp_g = QGroupBox("  Meu WhatsApp (para Faltas e NPS)"); wpp_l = QFormLayout(wpp_g)
        wpp_l.setSpacing(10)
        self.wpp_num = QLineEdit(self.settings.get("meu_whatsapp", ""))
        self.wpp_num.setPlaceholderText("Ex: 21999999999 (com DDD, sem +55)")
        wpp_l.addRow("Numero:", self.wpp_num)
        lay.addWidget(wpp_g)

        lay.addStretch()

    def _pick_preset(self, color, clicked):
        self._sel = color
        for btn,c in self._swatches: btn.set_selected(c.lower()==color.lower())
        self._paint_prev(); self.hex_lbl.setText(color.upper())
    def _pick_custom(self):
        c = QColorDialog.getColor(QColor(self._sel), self, "Cor Personalizada")
        if c.isValid():
            self._sel = c.name()
            for btn,_ in self._swatches: btn.set_selected(False)
            self._paint_prev(); self.hex_lbl.setText(self._sel.upper())
    def _paint_prev(self):
        self.cprev.setStyleSheet(f"QPushButton{{background:{self._sel};border:2px solid rgba(255,255,255,.4);border-radius:8px;}}"
                                 f"QPushButton:hover{{border:2px solid #FFF;}}")
    def _choose_icon(self):
        p,_ = QFileDialog.getOpenFileName(self,"Escolher Icone","","Icones (*.ico *.png *.jpg)")
        if p:
            self.settings["app_icon"] = p
            self.icon_prev.setPixmap(QPixmap(p).scaled(56,56,Qt.KeepAspectRatio,Qt.SmoothTransformation))
    def _clear_icon(self):
        self.settings["app_icon"] = ""; self.icon_prev.clear(); self.icon_prev.setText("Sem\nicone")
    def _choose_bg(self):
        p,_ = QFileDialog.getOpenFileName(self,"Imagem de Fundo","","Imagens (*.jpg *.jpeg *.png *.bmp *.webp)")
        if p:
            self.settings["background_image"] = p
            self.bgprev.setPixmap(QPixmap(p).scaled(440,100,Qt.KeepAspectRatio,Qt.SmoothTransformation))
    def _clear_bg(self):
        self.settings["background_image"] = ""; self.bgprev.clear(); self.bgprev.setText("Nenhuma imagem selecionada")
    def _apply(self):
        self.settings["theme"] = "dark" if self.rdark.isChecked() else "light"
        self.settings["accent_color"] = self._sel
        self.settings["fullscreen"] = self.chk_fs.isChecked()
        self.settings["meu_whatsapp"] = self.wpp_num.text().strip()
        self.dm.save_settings(self.settings)
        # Salva configuracao da Nota Fiscal Simples
        nf_cfg = {
            "nome_loja": self.nf_nome.text().strip()   or "MOTO PECAS & MECANICA",
            "slogan":    self.nf_slogan.text().strip()  or "Sistema de Gestao e Vendas",
            "endereco":  self.nf_end.text().strip(),
            "telefone":  self.nf_tel.text().strip(),
            "cnpj":      self.nf_cnpj.text().strip(),
            "rodape":    self.nf_rodape.text().strip()  or "Obrigado pela preferencia!",
        }
        self.dm.save_nota_fiscal_cfg(nf_cfg)
        self.accept()


# ═════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
class DashboardTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm = dm; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(18)
        hdr = QHBoxLayout(); t = QLabel("  Dashboard"); t.setObjectName("title")
        self.tl = QLabel(); self.tl.setObjectName("subtitle")
        hdr.addWidget(t); hdr.addStretch(); hdr.addWidget(self.tl); lay.addLayout(hdr)
        tmr = QTimer(self); tmr.timeout.connect(self._tick); tmr.start(1000); self._tick()
        cr = QHBoxLayout(); cr.setSpacing(14)
        self.c1 = MetricCard("Vendas Hoje","R$ 0,00","","#FF6B35")
        self.c2 = MetricCard("Vendas do Mes","R$ 0,00","","#00C853")
        self.c3 = MetricCard("Produtos","0","","#FFD600")
        self.c4 = MetricCard("Estoque Critico","0","","#FF1744")
        for c in [self.c1,self.c2,self.c3,self.c4]: cr.addWidget(c)
        lay.addLayout(cr)
        bot = QHBoxLayout(); bot.setSpacing(16)
        sf = QFrame(); sf.setObjectName("card"); sl = QVBoxLayout(sf); sl.setContentsMargins(14,14,14,14); sl.setSpacing(8)
        sl.addWidget(SectionTitle("  Ultimas Vendas"))
        self.st = QTableWidget(0,5); self.st.setHorizontalHeaderLabels(["Data","Cliente","Itens","Total","Pagamento"])
        self.st.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.st.setEditTriggers(QAbstractItemView.NoEditTriggers); self.st.setAlternatingRowColors(True)
        self.st.setSelectionBehavior(QAbstractItemView.SelectRows); self.st.verticalHeader().setVisible(False)
        self.st.setMaximumHeight(260); sl.addWidget(self.st); bot.addWidget(sf,3)
        ef = QFrame(); ef.setObjectName("card"); el = QVBoxLayout(ef); el.setContentsMargins(14,14,14,14); el.setSpacing(8)
        el.addWidget(SectionTitle("  Estoque Critico"))
        self.et = QTableWidget(0,4); self.et.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Minimo"])
        self.et.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.et.setEditTriggers(QAbstractItemView.NoEditTriggers); self.et.setAlternatingRowColors(True)
        self.et.verticalHeader().setVisible(False); self.et.setMaximumHeight(260)
        el.addWidget(self.et); bot.addWidget(ef,2); lay.addLayout(bot); lay.addStretch()
    def _tick(self): self.tl.setText(datetime.now().strftime("  %d/%m/%Y     %H:%M:%S"))
    def refresh(self):
        prods = self.dm.get_products(); sales = self.dm.get_sales(); now = datetime.now()
        td = sum(s.get("total",0) for s in sales if s.get("date","")[:10]==now.strftime("%Y-%m-%d"))
        mo = sum(s.get("total",0) for s in sales if s.get("date","")[:7]==now.strftime("%Y-%m"))
        alerts = [p for p in prods if p.get("stock",0)<=p.get("min_stock",0)]
        self.c1.set_value(fmtR(td)); self.c2.set_value(fmtR(mo))
        self.c3.set_value(str(len(prods))); self.c4.set_value(str(len(alerts)))
        recent = sorted(sales,key=lambda x:x.get("date",""),reverse=True)[:10]
        self.st.setRowCount(len(recent))
        for r,s in enumerate(recent):
            self.st.setItem(r,0,QTableWidgetItem(fmt_date(s.get("date",""))))
            self.st.setItem(r,1,QTableWidgetItem(s.get("customer_name","—")))
            self.st.setItem(r,2,QTableWidgetItem(str(len(s.get("items",[])))))
            ti = QTableWidgetItem(fmtR(s.get("total",0))); ti.setForeground(QColor("#00C853")); self.st.setItem(r,3,ti)
            self.st.setItem(r,4,QTableWidgetItem(s.get("payment_method","—")))
        self.et.setRowCount(len(alerts))
        for r,p in enumerate(alerts):
            self.et.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.et.setItem(r,1,QTableWidgetItem(p.get("name","")))
            si = QTableWidgetItem(str(p.get("stock",0))); si.setForeground(QColor("#FF1744")); self.et.setItem(r,2,si)
            self.et.setItem(r,3,QTableWidgetItem(str(p.get("min_stock",0))))


# ═════════════════════════════════════════════════════════════════════════════
#  PRODUCTS
# ═════════════════════════════════════════════════════════════════════════════
class ProductsTab(QWidget):
    def __init__(self, dm, user=None, parent=None):
        super().__init__(parent)
        self.dm = dm
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"
        self.can_edit = self.is_admin or self.user.get("can_edit_products", False)
        self._all = []; self._ids = []; self._build(); self.refresh()

    def set_user(self, user):
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"
        self.can_edit = self.is_admin or self.user.get("can_edit_products", False)
        self._apply_permissions()

    def _apply_permissions(self):
        """Show/hide buttons and cost column based on user role."""
        self.btn_add.setVisible(self.can_edit)
        self.btn_edit.setVisible(self.can_edit)
        self.btn_del.setVisible(self.can_edit)
        # Column 5 = Custo — hide for operators
        self.tbl.setColumnHidden(5, not self.is_admin)

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gestao de Produtos"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Codigo, nome ou marca..."); self.srch.search_changed.connect(self._filter); tb.addWidget(self.srch,3)
        self.catf = QComboBox(); self.catf.addItem("Todas"); self.catf.addItems(CATEGORIES)
        self.catf.currentTextChanged.connect(self._filter); tb.addWidget(self.catf,1)
        self.lowcb = QCheckBox("  Critico"); self.lowcb.stateChanged.connect(self._filter); tb.addWidget(self.lowcb)
        self.recentcb = QCheckBox("  Adicionados Recentemente"); self.recentcb.stateChanged.connect(self._filter); tb.addWidget(self.recentcb)
        tb.addStretch()
        self.btn_add = QPushButton("  Novo"); self.btn_add.clicked.connect(self._add); tb.addWidget(self.btn_add)
        self.btn_edit = QPushButton("  Editar"); self.btn_edit.setObjectName("btn_secondary"); self.btn_edit.clicked.connect(self._edit); tb.addWidget(self.btn_edit)
        self.btn_del = QPushButton("  Excluir"); self.btn_del.setObjectName("btn_danger"); self.btn_del.clicked.connect(self._delete); tb.addWidget(self.btn_del)
        self.btn_xml = QPushButton("  Importar XML"); self.btn_xml.setObjectName("btn_success"); self.btn_xml.clicked.connect(self._import_xml); tb.addWidget(self.btn_xml)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0,11)
        self.tbl.setHorizontalHeaderLabels(["Codigo","Nome","Categoria","Marca","Und","Custo","Venda","Estoque","Min","Status","Foto"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.doubleClicked.connect(self._edit)
        for c,w in enumerate([80,-1,100,100,55,90,90,70,60,85,52]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt = QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
        self._apply_permissions()

    def refresh(self):
        self._all = self.dm.get_products(); self._filter()

    def _filter(self):
        txt = self.srch.edit.text().lower(); cat = self.catf.currentText()
        low = self.lowcb.isChecked(); recent = self.recentcb.isChecked()
        prods = self._all
        if txt: prods = [p for p in prods if txt in p.get("code","").lower() or txt in p.get("name","").lower() or txt in p.get("brand","").lower()]
        if cat != "Todas": prods = [p for p in prods if p.get("category","") == cat]
        if low: prods = [p for p in prods if p.get("stock",0) <= p.get("min_stock",0)]
        if recent:
            # sort by added_at descending, show last 20
            prods = sorted(prods, key=lambda x: x.get("added_at",""), reverse=True)[:20]
        self.tbl.setRowCount(len(prods)); self._ids = []
        for r,p in enumerate(prods):
            self._ids.append(p["id"]); stk=p.get("stock",0); mn=p.get("min_stock",0)
            status = "Esgotado" if stk==0 else ("Baixo" if stk<=mn else "Ativo")
            self.tbl.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.tbl.setItem(r,1,QTableWidgetItem(p.get("name","")))
            self.tbl.setItem(r,2,QTableWidgetItem(p.get("category",""))); self.tbl.setItem(r,3,QTableWidgetItem(p.get("brand","")))
            self.tbl.setItem(r,4,QTableWidgetItem(p.get("unit","")))
            self.tbl.setItem(r,5,QTableWidgetItem(fmtR(p.get("cost_price",0)))); self.tbl.setItem(r,6,QTableWidgetItem(fmtR(p.get("sale_price",0))))
            si = QTableWidgetItem(str(stk)); si.setTextAlignment(Qt.AlignCenter)
            si.setForeground(QColor("#FF1744" if stk==0 else "#FFD600" if stk<=mn else "#00C853")); self.tbl.setItem(r,7,si)
            mi = QTableWidgetItem(str(mn)); mi.setTextAlignment(Qt.AlignCenter); self.tbl.setItem(r,8,mi)
            sti = QTableWidgetItem(status); sti.setTextAlignment(Qt.AlignCenter)
            sti.setForeground(QColor({"Ativo":"#00C853","Baixo":"#FFD600","Esgotado":"#FF1744"}.get(status,"#AAA"))); self.tbl.setItem(r,9,sti)
            # highlight recent row
            if recent:
                for c in range(11):
                    cell = self.tbl.item(r, c)
                    if cell: cell.setBackground(QColor("#1A2A1A") if self.is_admin or True else QColor("#1A2A1A"))
            # Botao de imagem do produto
            _pname = p.get("name", "")
            _pbrand = p.get("brand", "")
            _search_term = f"{_pname} {_pbrand}".strip()
            btn_img = QPushButton("\U0001f5bc")
            btn_img.setFixedSize(42, 28)
            btn_img.setToolTip(f"Ver imagem: {_search_term}")
            btn_img.setStyleSheet(
                "QPushButton{background:#2A2A2A;border:1px solid #3A3A3A;border-radius:5px;"
                "font-size:14px;padding:0;}QPushButton:hover{background:#FF6B35;}"
            )
            btn_img.clicked.connect(lambda _, n=_search_term: self._view_product_image(n))
            self.tbl.setCellWidget(r, 10, btn_img)
            self.tbl.setRowHeight(r,36)
        self.cnt.setText(f"Exibindo {len(prods)} de {len(self._all)} produtos")

    def _view_product_image(self, name):
        """Abre modal com imagem do produto buscada no Google Imagens."""
        dlg = ProductImageDialog(name, self)
        dlg.exec_()

    def _sel(self):
        r = self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None

    def _check_edit_permission(self):
        if self.can_edit: return True
        QMessageBox.warning(self,"Sem Permissao",
            "Voce nao tem permissao para editar produtos.\n\nSolicite ao administrador que habilite\nesta permissao no seu cadastro de usuario.")
        return False

    def _add(self):
        if not self._check_edit_permission(): return
        dlg = ProductDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            dlg.result_data["added_at"] = datetime.now().isoformat()
            self.dm.add_product(dlg.result_data); self.refresh()

    def _edit(self):
        if not self._check_edit_permission(): return
        pid = self._sel()
        if not pid: QMessageBox.information(self,"Atencao","Selecione um produto."); return
        p = self.dm.get_product_by_id(pid)
        if not p: return
        dlg = ProductDialog(self, p)
        if dlg.exec_() == QDialog.Accepted: dlg.result_data["id"]=pid; self.dm.update_product(pid,dlg.result_data); self.refresh()

    def _delete(self):
        if not self._check_edit_permission(): return
        pid = self._sel()
        if not pid: QMessageBox.information(self,"Atencao","Selecione um produto."); return
        p = self.dm.get_product_by_id(pid)
        if p and QMessageBox.question(self,"Confirmar",f"Excluir '{p['name']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_product(pid); self.refresh()

    def _import_xml(self):
        dlg = ImportXMLDialog(self, self.dm)
        if dlg.exec_() == QDialog.Accepted:
            self.refresh()

# ═════════════════════════════════════════════════════════════════════════════
#  SALES (PDV)
# ═════════════════════════════════════════════════════════════════════════════
class SalesTab(QWidget):
    sale_completed = pyqtSignal()
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self.cart=[]; self._pids=[]; self._build(); self._sp("")
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        top_l = QHBoxLayout()
        top_l.addWidget(SectionTitle("  Nova Venda"))
        top_l.addStretch()
        self.btn_focus = QPushButton(" ⚡ Modo Foco PDV [OFF] ")
        self.btn_focus.setObjectName("btn_secondary")
        self.btn_focus.setCheckable(True)
        self.btn_focus.clicked.connect(self._toggle_focus)
        top_l.addWidget(self.btn_focus)
        lay.addLayout(top_l)
        
        spl = QSplitter(Qt.Horizontal); spl.setHandleWidth(8)
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,0,0); ll.setSpacing(10)
        sg = QGroupBox("Adicionar Produto"); g = QGridLayout(sg); g.setSpacing(8)
        g.addWidget(QLabel("Produto:"),0,0)
        self.psr = QLineEdit(); self.psr.setPlaceholderText("Codigo ou nome...")
        self.psr.textChanged.connect(self._sp); g.addWidget(self.psr,0,1,1,3)
        self.pl = QTableWidget(0,5); self.pl.setHorizontalHeaderLabels(["Codigo","Nome","Marca","Preco","Estoque"])
        self.pl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.pl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.pl.verticalHeader().setVisible(False); self.pl.setMaximumHeight(155)
        self.pl.selectionModel().selectionChanged.connect(self._osel); g.addWidget(self.pl,1,0,1,4)
        g.addWidget(QLabel("Qtd:"),2,0); self.qty = QSpinBox(); self.qty.setRange(1,9999); self.qty.setValue(1); g.addWidget(self.qty,2,1)
        g.addWidget(QLabel("Preco:"),2,2); self.up = QDoubleSpinBox(); self.up.setRange(0,999999); self.up.setPrefix("R$ "); self.up.setDecimals(2); g.addWidget(self.up,2,3)
        g.addWidget(QLabel("Desc. Item:"),3,0); self.idc = QDoubleSpinBox(); self.idc.setRange(0,100); self.idc.setSuffix("%"); g.addWidget(self.idc,3,1)
        bai = QPushButton("  Adicionar ao Carrinho"); bai.clicked.connect(self._add_item); g.addWidget(bai,3,2,1,2)
        ll.addWidget(sg)
        cg = QGroupBox("Carrinho de Compras"); cl = QVBoxLayout(cg)
        self.ct = QTableWidget(0,7); self.ct.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Preco","Desc%","Total",""])
        self.ct.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.ct.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.ct.setEditTriggers(QAbstractItemView.NoEditTriggers); self.ct.setAlternatingRowColors(True)
        self.ct.verticalHeader().setVisible(False)
        for c,w in enumerate([80,-1,50,90,60,90,40]):
            if w>0: self.ct.setColumnWidth(c,w)
        cl.addWidget(self.ct)
        br = QPushButton("  Remover Selecionado"); br.setObjectName("btn_danger"); br.clicked.connect(self._rsel); cl.addWidget(br)
        ll.addWidget(cg); spl.addWidget(left)
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(12)
        cug = QGroupBox("  Cliente"); cul = QVBoxLayout(cug)
        self.ccb = QComboBox(); self.ccb.addItem("-- Consumidor Final --",None)
        for c in self.dm.get_customers():
            lbl = c["name"]
            if c.get("placa"): lbl += f" | {c['placa']}"
            elif c.get("veiculo"): lbl += f" | {c.get('veiculo','')[:20]}"
            self.ccb.addItem(lbl, c["id"])
        cul.addWidget(self.ccb); rl.addWidget(cug)
        pg = QGroupBox("  Pagamento"); pf = QFormLayout(pg)
        self.pm = QComboBox(); self.pm.addItems(PAYMENTS); pf.addRow("Forma:",self.pm)
        self.gd = QDoubleSpinBox(); self.gd.setRange(0,100); self.gd.setSuffix("%"); self.gd.valueChanged.connect(self._utot); pf.addRow("Desconto Geral:",self.gd)
        self.obs = QTextEdit(); self.obs.setMaximumHeight(58); self.obs.setPlaceholderText("Observacoes..."); pf.addRow("Obs:",self.obs)
        rl.addWidget(pg)
        tg = QGroupBox("  Totais"); tgl = QGridLayout(tg)
        def trow(lbl,attr,r,big=False):
            l=QLabel(lbl); v=QLabel("R$ 0,00"); v.setAlignment(Qt.AlignRight|Qt.AlignVCenter)
            if big: l.setStyleSheet("font-size:16px;font-weight:bold;"); v.setStyleSheet("font-size:20px;font-weight:bold;color:#FF6B35;")
            setattr(self,attr,v); tgl.addWidget(l,r,0); tgl.addWidget(v,r,1)
        trow("Subtotal:","ls",0); trow("Desconto:","ld",1)
        sep=QFrame(); sep.setFrameShape(QFrame.HLine); tgl.addWidget(sep,2,0,1,2)
        trow("TOTAL:","lt",3,True); rl.addWidget(tg); rl.addStretch()
        bf = QPushButton("  FINALIZAR VENDA"); bf.setObjectName("btn_success"); bf.setMinimumHeight(50)
        bf.setFont(QFont("Segoe UI",14,QFont.Bold)); bf.clicked.connect(self._fin); rl.addWidget(bf)
        bnfc = QPushButton("  Emitir NFC-e (DANFE + XML)"); bnfc.setObjectName("btn_secondary"); bnfc.setMinimumHeight(40)
        bnfc.clicked.connect(self._emitir_nfce); rl.addWidget(bnfc)
        bnps = QPushButton("  Enviar NPS WhatsApp"); bnps.setObjectName("btn_success"); bnps.setMinimumHeight(40)
        bnps.clicked.connect(self._send_nps); rl.addWidget(bnps)
        bnf = QPushButton("  Nota Fiscal Simples (PDF)"); bnf.setObjectName("btn_secondary"); bnf.setMinimumHeight(36)
        bnf.clicked.connect(self._emitir_nf); rl.addWidget(bnf)
        bc = QPushButton("  Limpar Venda"); bc.setObjectName("btn_secondary"); bc.clicked.connect(self._clr); rl.addWidget(bc)
        spl.addWidget(right); spl.setSizes([650,320]); lay.addWidget(spl,1)
    def _sp(self,txt=""):
        prods = self.dm.get_products()
        if txt: prods = [p for p in prods if txt.lower() in p.get("code","").lower() or txt.lower() in p.get("name","").lower()]
        prods = prods[:60]; self.pl.setRowCount(len(prods)); self._pids=[]
        for r,p in enumerate(prods):
            self._pids.append(p["id"])
            self.pl.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.pl.setItem(r,1,QTableWidgetItem(p.get("name","")))
            self.pl.setItem(r,2,QTableWidgetItem(p.get("brand",""))); self.pl.setItem(r,3,QTableWidgetItem(fmtR(p.get("sale_price",0))))
            si=QTableWidgetItem(str(p.get("stock",0))); si.setForeground(QColor("#FF1744" if p.get("stock",0)==0 else "#00C853")); self.pl.setItem(r,4,si)
            self.pl.setRowHeight(r,30)
    def _osel(self):
        r=self.pl.currentRow()
        if 0<=r<len(self._pids):
            p=self.dm.get_product_by_id(self._pids[r])
            if p: self.up.setValue(p.get("sale_price",0))
    def _add_item(self):
        r=self.pl.currentRow()
        if r<0 or r>=len(self._pids): QMessageBox.information(self,"Atencao","Selecione um produto."); return
        pid=self._pids[r]; p=self.dm.get_product_by_id(pid)
        if not p: return
        qty=self.qty.value(); price=self.up.value(); disc=self.idc.value()
        if p.get("stock",0)<qty: QMessageBox.warning(self,"Estoque",f"Disponivel: {p.get('stock',0)} {p.get('unit','UN')}"); return
        for it in self.cart:
            if it["product_id"]==pid: it["quantity"]+=qty; it["unit_price"]=price; it["discount"]=disc; self._rc(); return
        self.cart.append({"product_id":pid,"code":p.get("code",""),"name":p.get("name",""),"quantity":qty,"unit_price":price,"discount":disc})
        self._rc(); self.qty.setValue(1); self.idc.setValue(0)
    def _rc(self):
        self.ct.setRowCount(len(self.cart))
        for r,it in enumerate(self.cart):
            df=1-it["discount"]/100; tot=it["quantity"]*it["unit_price"]*df
            self.ct.setItem(r,0,QTableWidgetItem(it["code"])); self.ct.setItem(r,1,QTableWidgetItem(it["name"]))
            self.ct.setItem(r,2,QTableWidgetItem(str(it["quantity"]))); self.ct.setItem(r,3,QTableWidgetItem(fmtR(it["unit_price"])))
            self.ct.setItem(r,4,QTableWidgetItem(f"{it['discount']:.1f}%"))
            ti=QTableWidgetItem(fmtR(tot)); ti.setForeground(QColor("#00C853")); self.ct.setItem(r,5,ti)
            db=QPushButton("x"); db.setFixedSize(28,28); db.setObjectName("btn_danger")
            db.clicked.connect(lambda _,i=r: self._ri(i)); self.ct.setCellWidget(r,6,db)
            self.ct.setRowHeight(r,36)
        self._utot()
    def _ri(self,r):
        if 0<=r<len(self.cart): self.cart.pop(r); self._rc()
    def _rsel(self): self._ri(self.ct.currentRow())
    def _utot(self):
        sub=sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp=self.gd.value(); dv=sub*dp/100; tot=sub-dv
        self.ls.setText(fmtR(sub)); self.ld.setText(f"-{fmtR(dv)}"); self.lt.setText(fmtR(tot))
    def _fin(self):
        if not self.cart: QMessageBox.warning(self,"Vazio","Adicione produtos ao carrinho."); return
        sub=sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp=self.gd.value(); dv=sub*dp/100; tot=sub-dv
        
        pm = self.pm.currentText()
        settings = self.dm.get_settings()
        tk = settings.get("mp_access_token", "")
        dv_id = settings.get("mp_device_id", "")
        
        # MOCK INTERCEPTOR MERCADO PAGO
        if tk and dv_id:
            if pm == "PIX":
                dlg = PaymentProcessDialog(self, tot, "pix")
                if dlg.exec_() != QDialog.Accepted:
                    QMessageBox.warning(self, "Cancelado", "Pagamento PIX cancelado pelo app do Mercado Pago."); return
            elif pm in ["Cartão de Crédito", "Cartão de Débito"]:
                dlg = PaymentProcessDialog(self, tot, "credit")
                if dlg.exec_() != QDialog.Accepted:
                    QMessageBox.warning(self, "Cancelado", "Transacao cancelada na Maquininha."); return
        
        cid=self.ccb.currentData(); cn="Consumidor Final" if cid is None else self.ccb.currentText().split("|")[0].strip()
        sale={"customer_id":cid or "","customer_name":cn,"items":self.cart.copy(),"subtotal":sub,
              "discount_pct":dp,"discount_value":dv,"total":tot,"payment_method":self.pm.currentText(),
              "observations":self.obs.toPlainText(),"status":"concluida"}
        sid=self.dm.add_sale(sale); self.sale_completed.emit()
        
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Venda Concluida")
        msg_box.setText(f"Venda registrada!\n\nCliente: {cn}\nTotal: {fmtR(tot)}\nPagamento: {self.pm.currentText()}\nID: {sid[:8]}...")
        btn_whats = msg_box.addButton("Enviar nota para cliente", QMessageBox.ActionRole)
        btn_ok = msg_box.addButton("OK", QMessageBox.AcceptRole)
        msg_box.exec_()
        
        if msg_box.clickedButton() == btn_whats:
            phone = ""
            c_obj = None
            if cid:
                c_obj = self.dm.get_customer_by_id(cid)
                if c_obj: phone = c_obj.get("phone", "")
            
            if not phone:
                phone, ok = QInputDialog.getText(self, "WhatsApp", "Digite o WhatsApp do cliente com DDD (Apenas numeros):")
                if not ok or not phone: return
            
            phone = "".join(filter(str.isdigit, phone))
            if len(phone) in [10, 11]: phone = "55" + phone
            
            folder = os.path.join(BASE_DIR, "notas_emitidas")
            if not os.path.exists(folder): os.makedirs(folder)
            aamm = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_path = os.path.join(folder, f"nota_{sid[:6]}_{aamm}.pdf")
            has_pdf = False
            try:
                if PDF_OK:
                    _gerar_nota_fiscal(pdf_path, self.cart, cn, c_obj, sub, dp, dv, tot, self.pm.currentText(), self.obs.toPlainText(), self.dm)
                    has_pdf = True
            except: pass

            items_lines = "\n".join([f"- {it['quantity']}x {it['name']} ({fmtR(it['unit_price'])})" for it in self.cart])
            msg = f"Ola! Agradecemos a preferencia.\n\n*Resumo da compra - Moto Pecas:*\n{items_lines}\n\n*Total:* {fmtR(tot)}\n*Pagamento:* {self.pm.currentText()}\n\nAgilize arrastando a nota PDF nesta conversa. Volte sempre!"
            encoded = urllib.parse.quote(msg)
            webbrowser.open(f"https://wa.me/{phone}?text={encoded}")
            
            if has_pdf:
                import subprocess, platform
                try:
                    if platform.system() == "Windows": subprocess.Popen(rf'explorer /select,"{pdf_path}"')
                    elif platform.system() == "Darwin": subprocess.Popen(["open", "-R", pdf_path])
                    else: subprocess.Popen(["xdg-open", folder])
                except: pass
            
        self._clr()
    def _toggle_focus(self, checked):
        if checked:
            self.btn_focus.setText(" ⚡ Modo Foco PDV [ON] ")
            self.btn_focus.setStyleSheet("background: #FFD600; color: #111; font-weight: bold; border-radius: 8px; padding: 6px 12px;")
        else:
            self.btn_focus.setText(" ⚡ Modo Foco PDV [OFF] ")
            self.btn_focus.setStyleSheet("")
        if parent := self.window():
            if hasattr(parent, "toggle_focus_mode"):
                parent.toggle_focus_mode(checked)

    def _emitir_nfce(self):
        if not self.cart:
            QMessageBox.warning(self, "Carrinho Vazio", "Adicione produtos ao carrinho antes de emitir a NFC-e."); return
        if not PDF_OK:
            QMessageBox.critical(self, "Erro", "Biblioteca 'reportlab' nao instalada."); return
        empresa = self.dm.get_empresa()
        nfce_cfg = self.dm.get_nfce()
        sub = sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp = self.gd.value(); dv = sub*dp/100; tot = sub-dv
        cid = self.ccb.currentData()
        cn = "Consumidor Final" if cid is None else self.ccb.currentText().split("|")[0].strip()
        cust_obj = self.dm.get_customer_by_id(cid) if cid else None
        # Enrich cart items with fiscal fields from product data
        enriched = []
        for it in self.cart:
            prod = self.dm.get_product_by_id(it["product_id"]) or {}
            enriched.append({**it, "ncm": prod.get("ncm","00000000"),
                             "cfop": prod.get("cfop","5102"),
                             "csosn": prod.get("csosn","400"),
                             "orig": prod.get("orig","0"),
                             "unit": prod.get("unit","UN")})
        sale = {"customer_id": cid or "", "customer_name": cn,
                "items": enriched, "subtotal": sub,
                "discount_pct": dp, "discount_value": dv, "total": tot,
                "payment_method": self.pm.currentText(),
                "observations": self.obs.toPlainText(), "status": "concluida"}
        nf_num = self.dm.next_nf_number()
        aamm = datetime.now().strftime("%y%m")
        c_nf = random.randint(10000000, 99999999)
        chave44 = _gerar_chave_nfce(nfce_cfg.get("cuf","35"), aamm,
                                    empresa.get("cnpj","00000000000000"),
                                    "65", nfce_cfg.get("serie","001"), nf_num, c_nf)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_base = f"NFC-e_{nf_num:09d}_{ts}"
        folder = QFileDialog.getExistingDirectory(self, "Escolher Pasta para Salvar NFC-e")
        if not folder: return
        pdf_path = os.path.join(folder, f"{default_base}_DANFE.pdf")
        xml_path = os.path.join(folder, f"{default_base}_NFCe.xml")
        try:
            _gerar_danfe_nfce(pdf_path, empresa, nfce_cfg, sale, nf_num, chave44)
            xml_str = _gerar_xml_nfce(empresa, nfce_cfg, sale, nf_num, chave44)
            with open(xml_path, "w", encoding="utf-8") as f: f.write(xml_str)
            amb = nfce_cfg.get("ambiente","2")
            msg = (f"NFC-e N. {nf_num:09d} gerada!\n\n"
                   f"DANFE PDF: {os.path.basename(pdf_path)}\n"
                   f"XML: {os.path.basename(xml_path)}\n\n"
                   f"Chave: {chave44[:22]}...\n\n"
                   f"{'HOMOLOGACAO - sem valor fiscal' if amb=='2' else 'PRODUCAO - transmita o XML a SEFAZ'}\n\n"
                   f"Deseja abrir o DANFE agora?")
            reply = QMessageBox.information(self, "NFC-e Gerada", msg, QMessageBox.Open|QMessageBox.Close)
            if reply == QMessageBox.Open:
                import subprocess
                if sys.platform=="win32": os.startfile(pdf_path)
                elif sys.platform=="darwin": subprocess.call(["open", pdf_path])
                else: subprocess.call(["xdg-open", pdf_path])
        except Exception as e:
            QMessageBox.critical(self, "Erro NFC-e", str(e))

    def _emitir_nf(self):
        if not self.cart:
            QMessageBox.warning(self, "Carrinho Vazio", "Adicione produtos ao carrinho antes de emitir a nota."); return
        if not PDF_OK:
            QMessageBox.critical(self, "Erro", "Biblioteca 'reportlab' nao instalada.\nExecute: pip install reportlab"); return
        sub = sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp = self.gd.value(); dv = sub*dp/100; tot = sub-dv
        cid = self.ccb.currentData()
        cn = "Consumidor Final" if cid is None else self.ccb.currentText().split("|")[0].strip()
        cust_obj = self.dm.get_customer_by_id(cid) if cid else None
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"nota_fiscal_{ts}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Nota Fiscal", default_name, "PDF (*.pdf)")
        if not path: return
        try:
            _gerar_nota_fiscal(path, self.cart, cn, cust_obj, sub, dp, dv, tot,
                               self.pm.currentText(), self.obs.toPlainText(), self.dm)
            reply = QMessageBox.information(self, "  Nota Fiscal Gerada",
                f"Nota fiscal salva em:\n{path}\n\nDeseja abrir o arquivo agora?",
                QMessageBox.Open | QMessageBox.Close)
            if reply == QMessageBox.Open:
                import subprocess, sys
                if sys.platform == "win32": os.startfile(path)
                elif sys.platform == "darwin": subprocess.call(["open", path])
                else: subprocess.call(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Gerar PDF", str(e))

    def _clr(self):
        self.cart=[]; self._rc(); self.ccb.setCurrentIndex(0); self.pm.setCurrentIndex(0); self.gd.setValue(0); self.obs.clear()
    def refresh_customers(self):
        self.ccb.clear(); self.ccb.addItem("-- Consumidor Final --",None)
        for c in self.dm.get_customers():
            label = c["name"]
            if c.get("placa"): label += f" | {c['placa']}"
            elif c.get("veiculo"): label += f" | {c['veiculo'][:20]}"
            self.ccb.addItem(label, c["id"])

    def _send_nps(self):
        cid = self.ccb.currentData()
        cust_name = self.ccb.currentText().split("|")[0].strip() if cid else ""
        if not cid or cust_name == "Consumidor Final":
            QMessageBox.warning(self, "Atencao", "Selecione um cliente valido para enviar NPS.")
            return
            
        c = self.dm.get_customer_by_id(cid)
        phone = c.get("phone", "")
        
        if not phone:
            phone, ok = QInputDialog.getText(self, "NPS WhatsApp", "Digite o numero do cliente (com DDD):")
            if not ok or not phone: return
            
        digits = "".join(filter(str.isdigit, phone))
        if len(digits) in [10, 11]: digits = "55" + digits
        loja = self.dm.get_nota_fiscal_cfg().get("nome_loja", "Moto Pecas")
        msg = f"Ola {cust_name}! 😊\n\n"
        msg += f"Voce comprou recentemente na *{loja}* e gostaríamos de saber sua opiniao.\n\n"
        msg += f"De 0 a 10, qual nota voce da para nosso atendimento? ⭐\n\n"
        msg += f"0️⃣ 1️⃣ 2️⃣ 3️⃣ 4️⃣ 5️⃣ 6️⃣ 7️⃣ 8️⃣ 9️⃣ 🔟\n\n"
        msg += f"Responda com o numero! Sua opiniao e muito importante. 🙏"
        import urllib.parse, webbrowser
        encoded = urllib.parse.quote(msg)
        webbrowser.open(f"https://wa.me/{digits}?text={encoded}")
        # Registrar NPS enviado - nao temos sale_id aqui antes de finalizar, 
        # mas pode-se registrar o NPS
        self.dm.add_nps({
            "sale_id": "Venda no PDV",
            "customer_name": cust_name,
            "phone": phone,
            "status": "enviado"
        })


# ═════════════════════════════════════════════════════════════════════════════
#  CUSTOMERS
# ═════════════════════════════════════════════════════════════════════════════
class CustomersTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._all=[]; self._ids=[]; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gestao de Clientes"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Nome, CPF/CNPJ ou telefone..."); self.srch.search_changed.connect(self._filter); tb.addWidget(self.srch,3); tb.addStretch()
        for lbl,obj,cb in [("  Novo","ba",self._add),("  Editar","be",self._edit),("  Excluir","bd",self._del),("  Historico","bh",self._hist)]:
            btn=QPushButton(lbl)
            if lbl=="  Excluir": btn.setObjectName("btn_danger")
            elif lbl!="  Novo": btn.setObjectName("btn_secondary")
            btn.clicked.connect(cb); tb.addWidget(btn)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0,6); self.tbl.setHorizontalHeaderLabels(["Nome","Telefone","Veiculo","Placa","Endereco","Total Compras"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.doubleClicked.connect(self._edit)
        for c,w in enumerate([-1,130,160,100,180,110]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt = QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
    def refresh(self): self._all=self.dm.get_customers(); self._filter()
    def _filter(self):
        txt=self.srch.edit.text().lower()
        custs=[c for c in self._all if not txt or txt in c.get("name","").lower()
                or txt in c.get("phone","").lower() or txt in c.get("veiculo","").lower()
                or txt in c.get("placa","").lower()]
        self.tbl.setRowCount(len(custs)); self._ids=[]
        sales=self.dm.get_sales()
        for r,c in enumerate(custs):
            self._ids.append(c["id"]); tot=sum(s.get("total",0) for s in sales if s.get("customer_id")==c["id"])
            self.tbl.setItem(r,0,QTableWidgetItem(c.get("name","")))
            self.tbl.setItem(r,1,QTableWidgetItem(c.get("phone","")))
            self.tbl.setItem(r,2,QTableWidgetItem(c.get("veiculo","")))
            placa_item = QTableWidgetItem(c.get("placa",""))
            placa_item.setTextAlignment(Qt.AlignCenter)
            if c.get("placa",""):
                placa_item.setForeground(QColor("#1E88E5"))
                placa_item.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.tbl.setItem(r,3,placa_item)
            self.tbl.setItem(r,4,QTableWidgetItem(c.get("address","")))
            ti=QTableWidgetItem(fmtR(tot)); ti.setForeground(QColor("#00C853")); self.tbl.setItem(r,5,ti); self.tbl.setRowHeight(r,36)
        self.cnt.setText(f"Exibindo {len(custs)} de {len(self._all)} clientes")
    def _sid(self):
        r=self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _add(self):
        dlg=CustomerDialog(self)
        if dlg.exec_()==QDialog.Accepted: self.dm.add_customer(dlg.result_data); self.refresh()
    def _edit(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if not c: return
        dlg=CustomerDialog(self,c)
        if dlg.exec_()==QDialog.Accepted: dlg.result_data["id"]=cid; self.dm.update_customer(cid,dlg.result_data); self.refresh()
    def _del(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if c and QMessageBox.question(self,"Confirmar",f"Excluir '{c['name']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_customer(cid); self.refresh()
    def _hist(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if not c: return
        sales=[s for s in self.dm.get_sales() if s.get("customer_id")==cid]
        dlg=QDialog(self); dlg.setWindowTitle(f"Historico — {c['name']}"); dlg.setMinimumSize(640,400)
        l=QVBoxLayout(dlg); l.setContentsMargins(16,16,16,16)
        vei_placa = ""
        if c.get("veiculo"): vei_placa = f"  |  <b>Veiculo:</b> {c['veiculo']}"
        if c.get("placa"):   vei_placa += f"  <b>Placa:</b> {c['placa']}"
        l.addWidget(QLabel(f"<b>Cliente:</b> {c['name']} | <b>Total vendas:</b> {len(sales)}{vei_placa}"))
        t=QTableWidget(len(sales),5); t.setHorizontalHeaderLabels(["Data","Itens","Subtotal","Desconto","Total"])
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch); t.setEditTriggers(QAbstractItemView.NoEditTriggers); t.verticalHeader().setVisible(False)
        for r,s in enumerate(sorted(sales,key=lambda x:x.get("date",""),reverse=True)):
            t.setItem(r,0,QTableWidgetItem(fmt_date(s.get("date",""))))
            t.setItem(r,1,QTableWidgetItem(str(len(s.get("items",[])))))
            t.setItem(r,2,QTableWidgetItem(fmtR(s.get("subtotal",0)))); t.setItem(r,3,QTableWidgetItem(fmtR(s.get("discount_value",0))))
            ti=QTableWidgetItem(fmtR(s.get("total",0))); ti.setForeground(QColor("#00C853")); t.setItem(r,4,ti)
        l.addWidget(t)
        tg=sum(s.get("total",0) for s in sales)
        lb=QLabel(f"<b>Total Gasto: {fmtR(tg)}</b>"); lb.setStyleSheet("color:#FF6B35;font-size:14px;"); l.addWidget(lb)
        btn=QPushButton("Fechar"); btn.clicked.connect(dlg.accept); l.addWidget(btn); dlg.exec_()


# ═════════════════════════════════════════════════════════════════════════════
#  REPORTS — dias em Portugues + exportacao Excel
# ═════════════════════════════════════════════════════════════════════════════
class ReportsTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._hdr=[]; self._data=[]; self._build()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(14)
        lay.addWidget(SectionTitle("  Relatorios"))
        ctrl = QHBoxLayout(); ctrl.setSpacing(10)
        ctrl.addWidget(QLabel("Relatorio:"))
        self.rt = QComboBox(); self.rt.setMinimumWidth(245)
        self.rt.addItems(["Vendas por Periodo","Vendas por Cliente","Vendas por Produto",
            "Estoque Atual","Produtos Criticos (Estoque Baixo)","Resumo Financeiro","Ranking Mais Vendidos"])
        ctrl.addWidget(self.rt)
        ctrl.addWidget(QLabel("De:"))
        self.df = QDateEdit(QDate.currentDate().addDays(-30)); self.df.setCalendarPopup(True); self.df.setDisplayFormat("dd/MM/yyyy"); ctrl.addWidget(self.df)
        ctrl.addWidget(QLabel("Ate:"))
        self.dt2 = QDateEdit(QDate.currentDate()); self.dt2.setCalendarPopup(True); self.dt2.setDisplayFormat("dd/MM/yyyy"); ctrl.addWidget(self.dt2)
        bg = QPushButton("  Gerar Relatorio"); bg.clicked.connect(self._gen); ctrl.addWidget(bg)
        if EXCEL_OK:
            bx = QPushButton("  Exportar Excel"); bx.setObjectName("btn_success"); bx.clicked.connect(self._export); ctrl.addWidget(bx)
        ctrl.addStretch(); lay.addLayout(ctrl)
        sr = QHBoxLayout(); sr.setSpacing(12)
        self.s1=MetricCard("Registros","0","","#FF6B35"); self.s2=MetricCard("Valor Total","R$ 0,00","","#00C853")
        self.s3=MetricCard("Media","R$ 0,00","","#FFD600"); self.s4=MetricCard("Periodo","0 dias","","#1E88E5")
        for c in [self.s1,self.s2,self.s3,self.s4]: c.setMaximumHeight(100); sr.addWidget(c)
        lay.addLayout(sr)
        self.tbl = QTableWidget(0,1); self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True); self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.verticalHeader().setVisible(False); self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl,1)
        self.info = QLabel("Selecione um relatorio e clique em Gerar"); self.info.setObjectName("subtitle"); self.info.setAlignment(Qt.AlignCenter); lay.addWidget(self.info)
        self._gen()
    def _gen(self):
        rt=self.rt.currentText(); d0=self.df.date().toString("yyyy-MM-dd"); d1=self.dt2.date().toString("yyyy-MM-dd")+"T23:59:59"
        sales=[s for s in self.dm.get_sales() if d0<=s.get("date","")<=d1]
        self.s4.set_value(f"{self.df.date().daysTo(self.dt2.date())+1} dias")
        m={"Vendas por Periodo":self._period,"Vendas por Cliente":self._customer,"Vendas por Produto":self._product,
           "Estoque Atual":self._stock,"Produtos Criticos (Estoque Baixo)":self._lowstock,
           "Resumo Financeiro":self._financial,"Ranking Mais Vendidos":self._top}
        fn=m.get(rt)
        if fn: fn() if rt in ("Estoque Atual","Produtos Criticos (Estoque Baixo)") else fn(sales)
    def _st(self, headers, rows, mc=None):
        self._hdr=headers; self._data=rows; mc=mc or []
        self.tbl.setColumnCount(len(headers)); self.tbl.setHorizontalHeaderLabels(headers); self.tbl.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for c,val in enumerate(row):
                it=QTableWidgetItem(str(val))
                if c in mc: it.setForeground(QColor("#00C853"))
                self.tbl.setItem(r,c,it); self.tbl.setRowHeight(r,32)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.info.setText(f"{len(rows)} registro(s).")
    def _upd(self,count,total):
        self.s1.set_value(str(count)); self.s2.set_value(fmtR(total)); self.s3.set_value(fmtR(total/max(1,count)))
    def _period(self, sales):
        bd=defaultdict(lambda:{"count":0,"total":0})
        for s in sales: d=s.get("date","")[:10]; bd[d]["count"]+=1; bd[d]["total"]+=s.get("total",0)
        rows=[]
        for day in sorted(bd.keys(),reverse=True):
            d=datetime.strptime(day,"%Y-%m-%d")
            dia = DIAS_PT.get(d.strftime("%A"), d.strftime("%A"))  # PORTUGUES!
            rows.append([d.strftime("%d/%m/%Y"), dia, bd[day]["count"], fmtR(bd[day]["total"])])
        self._st(["Data","Dia da Semana","N Vendas","Total"],rows,[3]); self._upd(len(sales),sum(s.get("total",0) for s in sales))
    def _customer(self, sales):
        bc=defaultdict(lambda:{"count":0,"total":0,"name":""})
        for s in sales: k=s.get("customer_id","") or "cf"; bc[k]["count"]+=1; bc[k]["total"]+=s.get("total",0); bc[k]["name"]=s.get("customer_name","Consumidor Final")
        rows=sorted(bc.values(),key=lambda x:x["total"],reverse=True)
        self._st(["Cliente","N Vendas","Total"],[[r["name"],r["count"],fmtR(r["total"])] for r in rows],[2]); self._upd(len(sales),sum(s.get("total",0) for s in sales))
    def _product(self, sales):
        bp=defaultdict(lambda:{"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]): pid=it.get("product_id",""); bp[pid]["qty"]+=it.get("quantity",0); bp[pid]["total"]+=it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100); bp[pid]["name"]=it.get("name",""); bp[pid]["code"]=it.get("code","")
        rows=sorted(bp.values(),key=lambda x:x["total"],reverse=True)
        self._st(["Codigo","Produto","Qtd Vendida","Total"],[[r["code"],r["name"],r["qty"],fmtR(r["total"])] for r in rows],[3]); self._upd(len(rows),sum(r["total"] for r in rows))
    def _stock(self):
        prods=self.dm.get_products(); rows=[]
        for p in sorted(prods,key=lambda x:x.get("name","")):
            stk=p.get("stock",0); status="Esgotado" if stk==0 else ("Critico" if stk<=p.get("min_stock",0) else "Normal")
            rows.append([p.get("code",""),p.get("name",""),p.get("category",""),p.get("brand",""),str(stk),str(p.get("min_stock",0)),fmtR(p.get("sale_price",0)),fmtR(stk*p.get("cost_price",0)),status])
        tv=sum(p.get("stock",0)*p.get("cost_price",0) for p in prods)
        self._st(["Codigo","Produto","Categoria","Marca","Estoque","Minimo","Preco","Val.Estoque","Status"],rows,[6,7])
        self.s1.set_value(str(len(prods))); self.s2.set_value(fmtR(tv)); self.s3.set_value(fmtR(tv/max(1,len(prods)))); self.info.setText(f"{len(prods)} produtos.")
    def _lowstock(self):
        prods=[p for p in self.dm.get_products() if p.get("stock",0)<=p.get("min_stock",0)]
        rows=[[p.get("code",""),p.get("name",""),p.get("category",""),str(p.get("stock",0)),str(p.get("min_stock",0)),str(max(0,p.get("min_stock",0)-p.get("stock",0))),fmtR(p.get("sale_price",0))] for p in sorted(prods,key=lambda x:x.get("stock",0))]
        self._st(["Codigo","Produto","Categoria","Estoque","Minimo","Qtd p/Repor","Preco"],rows,[6])
        self.s1.set_value(str(len(prods))); self.s2.set_value("--"); self.s3.set_value("--"); self.info.setText(f"{len(prods)} criticos.")
    def _financial(self, sales):
        sb=sum(s.get("subtotal",0) for s in sales); sd=sum(s.get("discount_value",0) for s in sales); sl=sum(s.get("total",0) for s in sales)
        bp=defaultdict(float)
        for s in sales: bp[s.get("payment_method","--")]+=s.get("total",0)
        rows=[[m,fmtR(v),f"{v/max(1,sl)*100:.1f}%"] for m,v in sorted(bp.items(),key=lambda x:x[1],reverse=True)]
        rows+=[["-"*20,"-"*10,""],[" TOTAL BRUTO",fmtR(sb),"100%"],[" DESCONTOS",f"-{fmtR(sd)}",""],[" TOTAL LIQUIDO",fmtR(sl),""]]
        self._st(["Forma de Pagamento","Valor","% do Total"],rows,[1]); self._upd(len(sales),sl)
    def _top(self, sales):
        bp=defaultdict(lambda:{"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]): pid=it.get("product_id",""); bp[pid]["qty"]+=it.get("quantity",0); bp[pid]["total"]+=it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100); bp[pid]["name"]=it.get("name",""); bp[pid]["code"]=it.get("code","")
        rs=sorted(bp.values(),key=lambda x:x["qty"],reverse=True)[:20]
        self._st(["Pos","Codigo","Produto","Qtd Vendida","Total"],[[f"#{i+1}",r["code"],r["name"],r["qty"],fmtR(r["total"])] for i,r in enumerate(rs)],[4]); self._upd(len(rs),sum(r["total"] for r in rs))
    def _export(self):
        if not EXCEL_OK: QMessageBox.warning(self,"Erro","openpyxl nao esta instalado."); return
        if not self._data: QMessageBox.information(self,"Atencao","Gere um relatorio antes de exportar."); return
        ts=datetime.now().strftime("%Y%m%d_%H%M%S")
        sn=self.rt.currentText().replace(" ","_").replace("/","_").replace("(","").replace(")","")
        path,_=QFileDialog.getSaveFileName(self,"Salvar Excel",f"relatorio_{sn}_{ts}.xlsx","Excel (*.xlsx)")
        if not path: return
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Relatorio"
        hf=PatternFill(start_color="FF6B35",end_color="FF6B35",fill_type="solid")
        hfont=XLFont(bold=True,color="FFFFFF",size=11)
        halign=Alignment(horizontal="center",vertical="center")
        thin=Side(style="thin",color="CCCCCC"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
        altf=PatternFill(start_color="F5F5F5",end_color="F5F5F5",fill_type="solid")
        nc=len(self._hdr); lc=get_column_letter(nc)
        ws.merge_cells(f"A1:{lc}1")
        t=ws["A1"]; t.value=f"SISTEMA MOTO PECAS -- {self.rt.currentText().upper()}"
        t.font=XLFont(bold=True,size=14,color="FF6B35"); t.alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=28
        ws.merge_cells(f"A2:{lc}2")
        p=ws["A2"]; p.value=f"Periodo: {self.df.date().toString('dd/MM/yyyy')} a {self.dt2.date().toString('dd/MM/yyyy')}  |  Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        p.font=XLFont(size=10,color="888888"); p.alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=18
        for c,h in enumerate(self._hdr,1):
            cell=ws.cell(row=4,column=c,value=h); cell.fill=hf; cell.font=hfont; cell.alignment=halign; cell.border=bdr
        ws.row_dimensions[4].height=22
        for r,row in enumerate(self._data,5):
            for c,val in enumerate(row,1):
                cell=ws.cell(row=r,column=c,value=val); cell.alignment=Alignment(horizontal="left",vertical="center"); cell.border=bdr
                if (r-5)%2==1: cell.fill=altf
            ws.row_dimensions[r].height=18
        for col in ws.columns:
            ml=0; cl2=get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value: ml=max(ml,len(str(cell.value)))
                except: pass
            ws.column_dimensions[cl2].width=min(50,max(10,ml+4))
        lr=len(self._data)+6
        ws.cell(row=lr,column=1,value="Total de Registros:").font=XLFont(bold=True); ws.cell(row=lr,column=2,value=self.s1.val.text())
        ws.cell(row=lr+1,column=1,value="Valor Total:").font=XLFont(bold=True); ws.cell(row=lr+1,column=2,value=self.s2.val.text())
        try: wb.save(path); QMessageBox.information(self,"  Exportado!",f"Salvo em:\n{path}")
        except Exception as e: QMessageBox.critical(self,"Erro",str(e))


# ═════════════════════════════════════════════════════════════════════════════
#  SALES HISTORY
# ═════════════════════════════════════════════════════════════════════════════
class HistoryTab(QWidget):
    data_changed = pyqtSignal()
    def __init__(self, dm, user=None, parent=None):
        super().__init__(parent)
        self.dm = dm
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"
        self._sales = []
        self._build()
        self.refresh()

    def set_user(self, user):
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Historico de Vendas"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Cliente ou ID..."); self.srch.search_changed.connect(self.refresh)
        tb.addWidget(self.srch, 2)
        tb.addWidget(QLabel("De:"))
        self.df = QDateEdit(QDate.currentDate().addDays(-30))
        self.df.setCalendarPopup(True); self.df.setDisplayFormat("dd/MM/yyyy"); tb.addWidget(self.df)
        tb.addWidget(QLabel("Ate:"))
        self.dt = QDateEdit(QDate.currentDate())
        self.dt.setCalendarPopup(True); self.dt.setDisplayFormat("dd/MM/yyyy"); tb.addWidget(self.dt)
        bf = QPushButton("  Filtrar"); bf.clicked.connect(self.refresh); tb.addWidget(bf)
        self.show_cancelled = QCheckBox("Mostrar canceladas")
        self.show_cancelled.stateChanged.connect(self.refresh); tb.addWidget(self.show_cancelled)
        tb.addStretch()
        bd = QPushButton("  Detalhar"); bd.setObjectName("btn_secondary"); bd.clicked.connect(self._detail); tb.addWidget(bd)
        self.btn_cancel = QPushButton("  Cancelar Venda"); self.btn_cancel.setObjectName("btn_warning")
        self.btn_cancel.clicked.connect(self._cancel_sale); tb.addWidget(self.btn_cancel)
        self.btn_delete = QPushButton("  Excluir Venda"); self.btn_delete.setObjectName("btn_danger")
        self.btn_delete.clicked.connect(self._delete_sale); tb.addWidget(self.btn_delete)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0, 8)
        self.tbl.setHorizontalHeaderLabels(["Data/Hora","Cliente","Itens","Subtotal","Desconto","Total","Pagamento","Status"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnWidth(0, 140); self.tbl.setColumnWidth(7, 100)
        self.tbl.doubleClicked.connect(self._detail)
        lay.addWidget(self.tbl, 1)
        self.info = QLabel(); self.info.setObjectName("subtitle"); lay.addWidget(self.info)

    def refresh(self):
        d0 = self.df.date().toString("yyyy-MM-dd")
        d1 = self.dt.date().toString("yyyy-MM-dd") + "T23:59:59"
        txt = self.srch.edit.text().lower()
        show_all = self.show_cancelled.isChecked()
        sales = [s for s in self.dm.get_sales() if d0 <= s.get("date","") <= d1]
        if not show_all:
            sales = [s for s in sales if s.get("status","concluida") != "cancelada"]
        if txt:
            sales = [s for s in sales if txt in s.get("customer_name","").lower()
                     or txt in s.get("id","").lower()]
        self._sales = sorted(sales, key=lambda x: x.get("date",""), reverse=True)
        self.tbl.setRowCount(len(self._sales))
        tot = 0
        for r, s in enumerate(self._sales):
            cancelled = s.get("status","") == "cancelada"
            if not cancelled: tot += s.get("total", 0)
            self.tbl.setItem(r, 0, QTableWidgetItem(fmt_date(s.get("date",""))))
            self.tbl.setItem(r, 1, QTableWidgetItem(s.get("customer_name","--")))
            self.tbl.setItem(r, 2, QTableWidgetItem(str(len(s.get("items",[])))))
            self.tbl.setItem(r, 3, QTableWidgetItem(fmtR(s.get("subtotal",0))))
            self.tbl.setItem(r, 4, QTableWidgetItem(fmtR(s.get("discount_value",0))))
            ti = QTableWidgetItem(fmtR(s.get("total",0)))
            ti.setForeground(QColor("#888888" if cancelled else "#00C853"))
            self.tbl.setItem(r, 5, ti)
            self.tbl.setItem(r, 6, QTableWidgetItem(s.get("payment_method","--")))
            st = QTableWidgetItem("Cancelada" if cancelled else "Concluida")
            st.setTextAlignment(Qt.AlignCenter)
            st.setForeground(QColor("#FF1744" if cancelled else "#00C853"))
            self.tbl.setItem(r, 7, st)
            if cancelled:
                for c in range(8):
                    cell = self.tbl.item(r, c)
                    if cell and c != 7: cell.setForeground(QColor("#888888"))
            self.tbl.setRowHeight(r, 36)
        self.info.setText(f"{len(self._sales)} venda(s) | Total do periodo: {fmtR(tot)}")

    def _sel_sale(self):
        r = self.tbl.currentRow()
        if r < 0 or r >= len(self._sales):
            QMessageBox.information(self, "Atencao", "Selecione uma venda na tabela.")
            return None
        return self._sales[r]

    def _request_auth(self, action):
        dlg = QDialog(self); dlg.setWindowTitle("Autorizacao Necessaria")
        dlg.setFixedWidth(390); dlg.setModal(True)
        lay = QVBoxLayout(dlg); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        lay.addWidget(QLabel("  Autorizacao Necessaria",
                             styleSheet="font-size:15px;font-weight:bold;"))
        lay.addWidget(QLabel(
            f"Para {action.lower()} esta venda, informe a senha de autorizacao do administrador.",
            wordWrap=True, styleSheet="font-size:12px;"))
        form = QHBoxLayout(); form.addWidget(QLabel("Senha:"))
        self._pwd_e = QLineEdit(); self._pwd_e.setEchoMode(QLineEdit.Password)
        self._pwd_e.setPlaceholderText("Senha de autorizacao")
        self._pwd_e.setMinimumHeight(38); form.addWidget(self._pwd_e, 1); lay.addLayout(form)
        self._pwd_err = QLabel(""); self._pwd_err.setStyleSheet("color:#FF1744;font-size:12px;")
        lay.addWidget(self._pwd_err)
        btns = QHBoxLayout(); btns.addStretch()
        bcl = QPushButton("Cancelar"); bcl.setObjectName("btn_secondary"); bcl.clicked.connect(dlg.reject)
        bok = QPushButton(f"  Confirmar {action}")
        def _check():
            if self._pwd_e.text() == "admin": dlg.accept()
            else: self._pwd_err.setText("  Senha incorreta!"); self._pwd_e.clear(); self._pwd_e.setFocus()
        bok.clicked.connect(_check); self._pwd_e.returnPressed.connect(_check)
        btns.addWidget(bcl); btns.addWidget(bok); lay.addLayout(btns)
        return dlg.exec_() == QDialog.Accepted

    def _cancel_sale(self):
        s = self._sel_sale()
        if not s: return
        if s.get("status") == "cancelada":
            QMessageBox.information(self, "Aviso", "Esta venda ja esta cancelada."); return
        if not self.is_admin:
            if not self._request_auth("Cancelar Venda"): return
        msg = "Cancelar a venda abaixo?\n\nData: {}\nCliente: {}\nTotal: {}\n\nO estoque dos produtos sera restaurado automaticamente.".format(fmt_date(s.get('date','')), s.get('customer_name','--'), fmtR(s.get('total',0)))
        if QMessageBox.question(self, "Confirmar Cancelamento", msg,
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            ok, msg2 = self.dm.cancel_sale(s["id"])
            if ok:
                QMessageBox.information(self, "  Cancelada", msg2)
                self.refresh()
                self.data_changed.emit()
            else: QMessageBox.warning(self, "Erro", msg2)

    def _delete_sale(self):
        s = self._sel_sale()
        if not s: return
        if not self.is_admin:
            if not self._request_auth("Excluir Venda"): return
        msg = "ATENCAO: Acao IRREVERSIVEL!\n\nExcluir permanentemente?\n\nData: " + fmt_date(s.get('date','')) + "\nCliente: " + s.get('customer_name','--') + "\nTotal: " + fmtR(s.get('total',0)) + "\n\nDica: cancele antes de excluir para restaurar o estoque."
        if QMessageBox.question(self, "  Confirmar Exclusao", msg,
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.dm.delete_sale(s["id"])
            QMessageBox.information(self, "  Excluida", "Venda excluida permanentemente.")
            self.refresh()
            self.data_changed.emit()

    def _detail(self):
        r = self.tbl.currentRow()
        if r < 0 or r >= len(self._sales): return
        s = self._sales[r]; cancelled = s.get("status","") == "cancelada"
        dlg = QDialog(self); dlg.setWindowTitle(f"Venda #{s.get('id','')[:8]}")
        dlg.setMinimumSize(580, 460)
        l = QVBoxLayout(dlg); l.setContentsMargins(16,16,16,16); l.setSpacing(10)
        if cancelled:
            bn = QLabel("  VENDA CANCELADA  "); bn.setAlignment(Qt.AlignCenter)
            bn.setStyleSheet("background:#FF1744;color:#FFF;font-weight:bold;font-size:13px;padding:8px;border-radius:6px;")
            l.addWidget(bn)
            if s.get("cancelled_at"):
                l.addWidget(QLabel(f"<b>Cancelada em:</b> {fmt_date(s.get('cancelled_at',''))}",
                                   styleSheet="color:#FF1744;font-size:12px;"))
        l.addWidget(QLabel(f"<b>Data:</b> {fmt_date(s.get('date',''))}  "
                           f"<b>Cliente:</b> {s.get('customer_name','--')}  "
                           f"<b>Pagamento:</b> {s.get('payment_method','--')}"))
        t = QTableWidget(len(s.get("items",[])), 5)
        t.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Preco","Total"])
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers); t.verticalHeader().setVisible(False)
        for ri, it in enumerate(s.get("items",[])):
            rt = it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100)
            t.setItem(ri,0,QTableWidgetItem(it.get("code",""))); t.setItem(ri,1,QTableWidgetItem(it.get("name","")))
            t.setItem(ri,2,QTableWidgetItem(str(it["quantity"]))); t.setItem(ri,3,QTableWidgetItem(fmtR(it["unit_price"])))
            ti = QTableWidgetItem(fmtR(rt))
            ti.setForeground(QColor("#888888" if cancelled else "#00C853")); t.setItem(ri,4,ti)
        l.addWidget(t)
        sl = QHBoxLayout(); sl.addStretch()
        sl.addWidget(QLabel(f"Sub: {fmtR(s.get('subtotal',0))}  Desc: -{fmtR(s.get('discount_value',0))}  "
                            f"<b>TOTAL: {fmtR(s.get('total',0))}</b>"))
        l.addLayout(sl)
        if s.get("observations",""): l.addWidget(QLabel(f"<b>Obs:</b> {s['observations']}"))
        btns = QHBoxLayout()
        if not cancelled:
            bcl2 = QPushButton("  Cancelar esta Venda"); bcl2.setObjectName("btn_warning")
            bcl2.clicked.connect(lambda: (dlg.accept(), self._cancel_sale())); btns.addWidget(bcl2)
        bdl2 = QPushButton("  Excluir Venda"); bdl2.setObjectName("btn_danger")
        bdl2.clicked.connect(lambda: (dlg.accept(), self._delete_sale())); btns.addWidget(bdl2)
        bret = QPushButton("  Devolver Item"); bret.setObjectName("btn_warning")
        bret.clicked.connect(lambda: (dlg.accept(), self._return_item())); btns.addWidget(bret)
        btns.addStretch()
        bclose = QPushButton("Fechar"); bclose.setObjectName("btn_secondary")
        bclose.clicked.connect(dlg.accept); btns.addWidget(bclose)
        l.addLayout(btns)
        dlg.exec_()

    def _return_item(self):
        s = self._sel_sale()
        if not s: return
        if s.get("status") == "cancelada":
            QMessageBox.information(self, "Aviso", "Nao e possivel devolver itens de uma venda cancelada."); return
        dlg = ReturnItemDialog(self, s, self.dm)
        if dlg.exec_() == QDialog.Accepted:
            self.refresh()
            self.data_changed.emit()

# =============================================================================
#  FALTAS — Produtos em Falta
# =============================================================================

class FaltasManager:
    """Gerencia o arquivo faltas.json com produtos em falta."""
    def __init__(self):
        self.data = []
        self._load()

    def _load(self):
        if os.path.exists(FALTAS_FILE):
            try:
                with open(FALTAS_FILE, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            except Exception:
                self.data = []

    def _save(self):
        with open(FALTAS_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def get_all(self): return self.data

    def add(self, item):
        item["id"]   = str(uuid.uuid4())
        item["data"] = datetime.now().isoformat()
        self.data.append(item); self._save()

    def update(self, iid, upd):
        for i, it in enumerate(self.data):
            if it["id"] == iid:
                upd["id"]   = iid
                upd["data"] = it.get("data", datetime.now().isoformat())
                self.data[i] = upd; break
        self._save()

    def delete(self, iid):
        self.data = [x for x in self.data if x["id"] != iid]; self._save()


class FaltaDialog(ModernFormDialog):
    """Dialog para cadastrar/editar produto em falta."""
    URGENCIAS = ["Normal", "Alta", "Urgente"]

    def __init__(self, parent=None, item=None):
        super().__init__(parent, min_width=450); self.item = item
        self.add_title_bar("Produto em Falta" if not item else "Editar Falta")
        lay = self.main_lay
        g = QGridLayout(); g.setSpacing(10)
        self.nome = QLineEdit(); self.nome.setPlaceholderText("Nome do produto ou peca")
        self.qtd  = QSpinBox();  self.qtd.setRange(1,9999); self.qtd.setValue(1)
        self.urg  = QComboBox(); self.urg.addItems(self.URGENCIAS)
        self.obs  = QTextEdit(); self.obs.setPlaceholderText("Observacoes, referencia, fornecedor..."); self.obs.setMaximumHeight(72)
        for r,(lbl,w) in enumerate([("Produto / Peca *",self.nome),("Quantidade",self.qtd),("Urgencia",self.urg),("Observacoes",self.obs)]):
            g.addWidget(QLabel(lbl),r,0); g.addWidget(w,r,1)
        g.setColumnStretch(1,1); lay.addLayout(g)
        if item:
            self.nome.setText(item.get("nome",""))
            self.qtd.setValue(int(item.get("qtd",1)))
            idx = self.urg.findText(item.get("urgencia","Normal"))
            if idx>=0: self.urg.setCurrentIndex(idx)
            self.obs.setPlainText(item.get("obs",""))
        btns = QHBoxLayout(); btns.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar"); sv.setObjectName("btn_success"); sv.clicked.connect(self._save)
        btns.addWidget(cn); btns.addWidget(sv); lay.addLayout(btns)

    def _save(self):
        if not self.nome.text().strip(): QMessageBox.warning(self,"Atencao","Informe o nome."); return
        self.result_data = {"nome": self.nome.text().strip(), "qtd": self.qtd.value(),
                            "urgencia": self.urg.currentText(), "obs": self.obs.toPlainText().strip()}
        self.accept()


class FaltasTab(QWidget):
    """Aba de gestao de produtos em falta."""

    def __init__(self, mgr, parent=None):
        super().__init__(parent); self.mgr=mgr; self._ids=[]; self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Lista de Faltas — Produtos em Falta"))
        info = QFrame(); info.setObjectName("card")
        info.setStyleSheet("QFrame#card{border-left:4px solid #FFD600;}")
        il = QHBoxLayout(info); il.setContentsMargins(14,8,14,8)
        il.addWidget(QLabel("  Registre aqui os produtos em falta para facilitar o reabastecimento.",
                            objectName="subtitle", wordWrap=True))
        lay.addWidget(info)
        tb = QHBoxLayout()
        self.srch = SearchBar("Buscar produto..."); self.srch.search_changed.connect(self._filter)
        tb.addWidget(self.srch,2); tb.addStretch()
        bwh=QPushButton("  Enviar via WhatsApp"); bwh.setObjectName("btn_success"); bwh.clicked.connect(self._send_whatsapp); tb.addWidget(bwh)
        ba=QPushButton("  Novo"); ba.clicked.connect(self._add); tb.addWidget(ba)
        be=QPushButton("  Editar"); be.setObjectName("btn_secondary"); be.clicked.connect(self._edit); tb.addWidget(be)
        bd=QPushButton("  Excluir"); bd.setObjectName("btn_danger"); bd.clicked.connect(self._delete); tb.addWidget(bd)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0,5)
        self.tbl.setHorizontalHeaderLabels(["Produto / Peca","Quantidade","Urgencia","Observacoes","Registrado em"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.doubleClicked.connect(self._edit)
        for c,w in enumerate([-1,90,90,200,130]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl,1)
        self.cnt=QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)

    def refresh(self): self.mgr._load(); self._filter()

    def _filter(self):
        txt=self.srch.edit.text().lower()
        items=[x for x in self.mgr.get_all()
               if not txt or txt in x.get("nome","").lower() or txt in x.get("obs","").lower()]
        self.tbl.setRowCount(len(items)); self._ids=[]
        urg_col={"Normal":"#00C853","Alta":"#FFD600","Urgente":"#FF1744"}
        for r,it in enumerate(items):
            self._ids.append(it["id"])
            self.tbl.setItem(r,0,QTableWidgetItem(it.get("nome","")))
            qi=QTableWidgetItem(str(it.get("qtd",1))); qi.setTextAlignment(Qt.AlignCenter); self.tbl.setItem(r,1,qi)
            ui=QTableWidgetItem(it.get("urgencia","Normal")); ui.setTextAlignment(Qt.AlignCenter)
            ui.setForeground(QColor(urg_col.get(it.get("urgencia","Normal"),"#AAA"))); self.tbl.setItem(r,2,ui)
            self.tbl.setItem(r,3,QTableWidgetItem(it.get("obs","")))
            self.tbl.setItem(r,4,QTableWidgetItem(fmt_date(it.get("data",""),show_time=False)))
            self.tbl.setRowHeight(r,36)
        self.cnt.setText(f"{len(items)} produto(s) em falta")

    def _sid(self):
        r=self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _add(self):
        dlg=FaltaDialog(self)
        if dlg.exec_()==QDialog.Accepted: self.mgr.add(dlg.result_data); self.refresh()
    def _edit(self):
        iid=self._sid()
        if not iid: QMessageBox.information(self,"Atencao","Selecione um item."); return
        item=next((x for x in self.mgr.get_all() if x["id"]==iid),None)
        if not item: return
        dlg=FaltaDialog(self,item)
        if dlg.exec_()==QDialog.Accepted: self.mgr.update(iid,dlg.result_data); self.refresh()
    def _delete(self):
        iid=self._sid()
        if not iid: QMessageBox.information(self,"Atencao","Selecione um item."); return
        item=next((x for x in self.mgr.get_all() if x["id"]==iid),None)
        if item and QMessageBox.question(self,"Confirmar",f"Excluir '{item['nome']}'?",
                                         QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.mgr.delete(iid); self.refresh()

    def _send_whatsapp(self):
        items = self.mgr.get_all()
        if not items:
            QMessageBox.information(self, "Vazio", "Nenhum produto em falta para enviar."); return
        # Pegar numero do whatsapp das configuracoes
        parent_win = self.window()
        dm = getattr(parent_win, "dm", None)
        meu_wpp = ""
        if dm:
            meu_wpp = dm.get_settings().get("meu_whatsapp", "")
        if not meu_wpp:
            meu_wpp, ok = QInputDialog.getText(self, "WhatsApp",
                "Digite seu numero de WhatsApp com DDD (ex: 21999999999):")
            if not ok or not meu_wpp: return
            # Salvar para proxima vez
            if dm:
                s = dm.get_settings(); s["meu_whatsapp"] = meu_wpp; dm.save_settings(s)
        digits = "".join(filter(str.isdigit, meu_wpp))
        if len(digits) in [10, 11]: digits = "55" + digits
        msg = "*LISTA DE FALTAS - Moto Pecas*\n"
        msg += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
        for i, it in enumerate(items, 1):
            urg = it.get("urgencia", "Normal")
            urg_icon = {"Normal": "🟢", "Alta": "🟡", "Urgente": "🔴"}.get(urg, "⚪")
            msg += f"{i}. {urg_icon} *{it.get('nome', '')}* — Qtd: {it.get('qtd', 1)}"
            if it.get("obs"): msg += f" ({it['obs'][:30]})"
            msg += "\n"
        msg += f"\nTotal: {len(items)} item(ns)"
        encoded = urllib.parse.quote(msg)
        webbrowser.open(f"https://wa.me/{digits}?text={encoded}")


# ═════════════════════════════════════════════════════════════════════════════
#  RETURN ITEM DIALOG (devolucao parcial de item)
# ═════════════════════════════════════════════════════════════════════════════
class ReturnItemDialog(ModernFormDialog):
    def __init__(self, parent, sale, dm):
        super().__init__(parent, min_width=560); self.sale = sale; self.dm = dm
        self.add_title_bar("Devolver Item")
        lay = self.main_lay
        lay.addWidget(QLabel(f"<b>Venda:</b> {fmt_date(sale.get('date',''))} | <b>Cliente:</b> {sale.get('customer_name','')}"))
        # Tabela de itens
        items = sale.get("items", [])
        self.tbl = QTableWidget(len(items), 5)
        self.tbl.setHorizontalHeaderLabels(["", "Produto", "Qtd Comprada", "Ja Devolvido", "Devolver"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.verticalHeader().setVisible(False)
        self._spins = []
        self._checks = []
        for r, it in enumerate(items):
            already = it.get("returned_qty", 0)
            avail = it["quantity"] - already
            chk = QCheckBox()
            chk.setEnabled(avail > 0)
            self._checks.append(chk)
            self.tbl.setCellWidget(r, 0, chk)
            self.tbl.setItem(r, 1, QTableWidgetItem(it.get("name", "")))
            self.tbl.setItem(r, 2, QTableWidgetItem(str(it["quantity"])))
            self.tbl.setItem(r, 3, QTableWidgetItem(str(already)))
            sp = QSpinBox(); sp.setRange(0, max(0, avail)); sp.setValue(0 if avail > 0 else 0)
            self._spins.append(sp)
            self.tbl.setCellWidget(r, 4, sp)
            self.tbl.setRowHeight(r, 36)
        lay.addWidget(self.tbl)
        # Motivo
        mg = QGroupBox("Motivo da Devolucao"); ml = QHBoxLayout(mg)
        self.motivo_cb = QComboBox()
        self.motivo_cb.addItems(["Avaria", "Defeito", "Arrependimento", "Outro"])
        ml.addWidget(QLabel("Motivo:")); ml.addWidget(self.motivo_cb, 1)
        lay.addWidget(mg)
        info = QLabel("Avaria/Defeito: produto vai para a aba Troca.\nArrependimento/Outro: produto volta para o estoque.")
        info.setWordWrap(True); info.setObjectName("subtitle"); lay.addWidget(info)
        # Botoes
        btns = QHBoxLayout(); btns.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Confirmar Devolucao"); sv.setObjectName("btn_warning"); sv.clicked.connect(self._confirm)
        btns.addWidget(cn); btns.addWidget(sv); lay.addLayout(btns)

    def _confirm(self):
        returns = []
        items = self.sale.get("items", [])
        for r, it in enumerate(items):
            if self._checks[r].isChecked() and self._spins[r].value() > 0:
                returns.append({
                    "product_id": it["product_id"],
                    "qty": self._spins[r].value(),
                    "motivo": self.motivo_cb.currentText()
                })
        if not returns:
            AlertDialog.show_info(self, "Atencao", "Selecione pelo menos um item e a quantidade para devolver.")
            return
        ok, msg = self.dm.partial_return(self.sale["id"], returns)
        if ok:
            AlertDialog.show_info(self, "Devolucao Registrada", msg, icon="✅")
            self.accept()
        else:
            AlertDialog.show_info(self, "Erro", msg, icon="❌")


# ═════════════════════════════════════════════════════════════════════════════
#  TROCAS TAB (produtos com avaria/defeito)
# ═════════════════════════════════════════════════════════════════════════════
class TrocasTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm = dm; self._ids = []; self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Trocas — Avarias e Defeitos"))
        info = QFrame(); info.setObjectName("card")
        info.setStyleSheet("QFrame#card{border-left:4px solid #FF1744;}")
        il = QHBoxLayout(info); il.setContentsMargins(14,8,14,8)
        il.addWidget(QLabel("  Produtos devolvidos por avaria ou defeito. Gerencie trocas com fornecedores.",
                            objectName="subtitle", wordWrap=True))
        lay.addWidget(info)
        # Metricas
        mr = QHBoxLayout(); mr.setSpacing(10)
        self.m1 = MetricCard("Pendentes", "0", "", "#FFD600")
        self.m2 = MetricCard("Em Troca", "0", "", "#1E88E5")
        self.m3 = MetricCard("Resolvidos", "0", "", "#00C853")
        self.m4 = MetricCard("Valor Total", "R$ 0", "", "#FF6B35")
        for m in [self.m1, self.m2, self.m3, self.m4]: m.setMaximumHeight(95); mr.addWidget(m)
        lay.addLayout(mr)
        # Toolbar
        tb = QHBoxLayout(); tb.addStretch()
        b1 = QPushButton("  Solicitar Troca ao Fornecedor"); b1.setObjectName("btn_secondary")
        b1.clicked.connect(lambda: self._update_status("em_troca")); tb.addWidget(b1)
        b2 = QPushButton("  Devolver ao Estoque"); b2.setObjectName("btn_success")
        b2.clicked.connect(lambda: self._update_status("devolvido_estoque")); tb.addWidget(b2)
        b3 = QPushButton("  Descartar"); b3.setObjectName("btn_danger")
        b3.clicked.connect(lambda: self._update_status("descartado")); tb.addWidget(b3)
        lay.addLayout(tb)
        # Tabela
        self.tbl = QTableWidget(0, 8)
        self.tbl.setHorizontalHeaderLabels(["Produto", "Cod", "Qtd", "Motivo", "Cliente", "Data", "Valor Un.", "Status"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        lay.addWidget(self.tbl, 1)

    def refresh(self):
        trocas = self.dm.get_trocas(); self._ids = []
        pend = sum(1 for t in trocas if t.get("status") == "pendente")
        em_troca = sum(1 for t in trocas if t.get("status") == "em_troca")
        resolv = sum(1 for t in trocas if t.get("status") in ("devolvido_estoque", "descartado"))
        val_total = sum(t.get("unit_price", 0) * t.get("quantity", 1) for t in trocas if t.get("status") == "pendente")
        self.m1.set_value(str(pend)); self.m2.set_value(str(em_troca))
        self.m3.set_value(str(resolv)); self.m4.set_value(fmtR(val_total))
        self.tbl.setRowCount(len(trocas))
        status_lbl = {"pendente": "Pendente", "em_troca": "Em Troca", "devolvido_estoque": "No Estoque", "descartado": "Descartado"}
        status_clr = {"pendente": "#FFD600", "em_troca": "#1E88E5", "devolvido_estoque": "#00C853", "descartado": "#FF1744"}
        for r, t in enumerate(trocas):
            self._ids.append(t["id"])
            self.tbl.setItem(r, 0, QTableWidgetItem(t.get("product_name", "")))
            self.tbl.setItem(r, 1, QTableWidgetItem(t.get("product_code", "")))
            qi = QTableWidgetItem(str(t.get("quantity", 1))); qi.setTextAlignment(Qt.AlignCenter); self.tbl.setItem(r, 2, qi)
            mi = QTableWidgetItem(t.get("motivo", "")); mi.setForeground(QColor("#FF6B35")); self.tbl.setItem(r, 3, mi)
            self.tbl.setItem(r, 4, QTableWidgetItem(t.get("customer_name", "")))
            self.tbl.setItem(r, 5, QTableWidgetItem(fmt_date(t.get("date", ""))))
            self.tbl.setItem(r, 6, QTableWidgetItem(fmtR(t.get("unit_price", 0))))
            st = t.get("status", "pendente")
            si = QTableWidgetItem(status_lbl.get(st, st)); si.setTextAlignment(Qt.AlignCenter)
            si.setForeground(QColor(status_clr.get(st, "#AAA"))); self.tbl.setItem(r, 7, si)
            self.tbl.setRowHeight(r, 36)

    def _update_status(self, new_status):
        r = self.tbl.currentRow()
        if r < 0 or r >= len(self._ids):
            QMessageBox.information(self, "Atencao", "Selecione um item."); return
        tid = self._ids[r]
        self.dm.update_troca(tid, new_status)
        QMessageBox.information(self, "Atualizado", f"Status atualizado!")
        self.refresh()


# ═════════════════════════════════════════════════════════════════════════════
#  CAIXA TAB (controle de caixa PDV)
# ═════════════════════════════════════════════════════════════════════════════
class CaixaTab(QWidget):
    def __init__(self, dm, user=None, parent=None):
        super().__init__(parent); self.dm = dm; self.user = user or {}; self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(14)
        lay.addWidget(SectionTitle("  Controle de Caixa — PDV"))
        # Status do caixa
        self.status_frame = QFrame(); self.status_frame.setObjectName("card")
        sf_lay = QHBoxLayout(self.status_frame); sf_lay.setContentsMargins(16,12,16,12)
        self.status_icon = QLabel(""); self.status_icon.setFont(QFont("Segoe UI Emoji", 28))
        self.status_lbl = QLabel(""); self.status_lbl.setStyleSheet("font-size:16px;font-weight:bold;")
        self.status_info = QLabel(""); self.status_info.setObjectName("subtitle")
        sf_lay.addWidget(self.status_icon); sf_lay.addWidget(self.status_lbl); sf_lay.addStretch(); sf_lay.addWidget(self.status_info)
        lay.addWidget(self.status_frame)
        # Botoes
        tb = QHBoxLayout(); tb.setSpacing(10)
        self.btn_abrir = QPushButton("  Abrir Caixa"); self.btn_abrir.setObjectName("btn_success")
        self.btn_abrir.setMinimumHeight(44); self.btn_abrir.clicked.connect(self._abrir)
        self.btn_sangria = QPushButton("  Sangria"); self.btn_sangria.setObjectName("btn_warning")
        self.btn_sangria.setMinimumHeight(44); self.btn_sangria.clicked.connect(self._sangria)
        self.btn_suprimento = QPushButton("  Suprimento"); self.btn_suprimento.setObjectName("btn_secondary")
        self.btn_suprimento.setMinimumHeight(44); self.btn_suprimento.clicked.connect(self._suprimento)
        self.btn_fechar = QPushButton("  Fechar Caixa"); self.btn_fechar.setObjectName("btn_danger")
        self.btn_fechar.setMinimumHeight(44); self.btn_fechar.clicked.connect(self._fechar)
        tb.addWidget(self.btn_abrir); tb.addWidget(self.btn_sangria)
        tb.addWidget(self.btn_suprimento); tb.addStretch(); tb.addWidget(self.btn_fechar)
        lay.addLayout(tb)
        # Metricas
        mr = QHBoxLayout(); mr.setSpacing(10)
        self.mc1 = MetricCard("Valor Inicial", "R$ 0,00", "", "#1E88E5")
        self.mc2 = MetricCard("Vendas", "R$ 0,00", "", "#00C853")
        self.mc3 = MetricCard("Sangrias", "R$ 0,00", "", "#FF1744")
        self.mc4 = MetricCard("Saldo Esperado", "R$ 0,00", "", "#FF6B35")
        for m in [self.mc1, self.mc2, self.mc3, self.mc4]: m.setMaximumHeight(95); mr.addWidget(m)
        lay.addLayout(mr)
        # Movimentacoes do dia
        lay.addWidget(SectionTitle("  Movimentacoes"))
        self.tbl = QTableWidget(0, 4)
        self.tbl.setHorizontalHeaderLabels(["Hora", "Tipo", "Valor", "Motivo"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.verticalHeader().setVisible(False)
        lay.addWidget(self.tbl, 1)

    def refresh(self):
        cx = self.dm.get_caixa()
        aberto = cx.get("aberto", False)
        if aberto:
            ab = cx.get("abertura", {})
            vi = ab.get("valor_inicial", 0)
            self.status_icon.setText("🟢"); self.status_lbl.setText("CAIXA ABERTO")
            self.status_lbl.setStyleSheet("font-size:16px;font-weight:bold;color:#00C853;")
            self.status_info.setText(f"Aberto em: {fmt_date(ab.get('data',''))} por {ab.get('operador','')}")
            self.btn_abrir.setEnabled(False); self.btn_sangria.setEnabled(True)
            self.btn_suprimento.setEnabled(True); self.btn_fechar.setEnabled(True)
            # Calcular totais
            dt_ab = ab.get("data", "")
            vendas = [s for s in self.dm.get_sales() if s.get("date","") >= dt_ab and s.get("status") != "cancelada"]
            tv = sum(s.get("total", 0) for s in vendas)
            movs = ab.get("movimentacoes", [])
            ts = sum(m["valor"] for m in movs if m["tipo"] == "sangria")
            tsp = sum(m["valor"] for m in movs if m["tipo"] == "suprimento")
            self.mc1.set_value(fmtR(vi)); self.mc2.set_value(fmtR(tv))
            self.mc3.set_value(fmtR(ts)); self.mc4.set_value(fmtR(vi + tv + tsp - ts))
            # Movimentacoes
            self.tbl.setRowCount(len(movs))
            for r, m in enumerate(movs):
                self.tbl.setItem(r, 0, QTableWidgetItem(fmt_date(m.get("data",""))))
                ti = QTableWidgetItem(m["tipo"].capitalize())
                ti.setForeground(QColor("#FF1744" if m["tipo"]=="sangria" else "#00C853"))
                self.tbl.setItem(r, 1, ti)
                self.tbl.setItem(r, 2, QTableWidgetItem(fmtR(m["valor"])))
                self.tbl.setItem(r, 3, QTableWidgetItem(m.get("motivo","")))
                self.tbl.setRowHeight(r, 32)
        else:
            self.status_icon.setText("🔴"); self.status_lbl.setText("CAIXA FECHADO")
            self.status_lbl.setStyleSheet("font-size:16px;font-weight:bold;color:#FF1744;")
            self.status_info.setText("Abra o caixa para iniciar as operacoes do dia.")
            self.btn_abrir.setEnabled(True); self.btn_sangria.setEnabled(False)
            self.btn_suprimento.setEnabled(False); self.btn_fechar.setEnabled(False)
            self.mc1.set_value("---"); self.mc2.set_value("---")
            self.mc3.set_value("---"); self.mc4.set_value("---")
            self.tbl.setRowCount(0)

    def _abrir(self):
        val, ok = QInputDialog.getDouble(self, "Abrir Caixa", "Valor inicial do caixa (fundo de troco):", 200.0, 0, 99999, 2)
        if not ok: return
        ok2, msg = self.dm.abrir_caixa(val, self.user.get("name", ""))
        if ok2: QMessageBox.information(self, "Caixa Aberto", msg); self.refresh()
        else: QMessageBox.warning(self, "Erro", msg)

    def _sangria(self):
        val, ok = QInputDialog.getDouble(self, "Sangria", "Valor da sangria:", 0, 0.01, 99999, 2)
        if not ok or val <= 0: return
        motivo, ok2 = QInputDialog.getText(self, "Sangria", "Motivo da sangria:")
        if not ok2: return
        ok3, msg = self.dm.add_movimentacao_caixa("sangria", val, motivo, self.user.get("name", ""))
        if ok3: self.refresh()
        else: QMessageBox.warning(self, "Erro", msg)

    def _suprimento(self):
        val, ok = QInputDialog.getDouble(self, "Suprimento", "Valor do suprimento:", 0, 0.01, 99999, 2)
        if not ok or val <= 0: return
        motivo, ok2 = QInputDialog.getText(self, "Suprimento", "Motivo do suprimento:")
        if not ok2: return
        ok3, msg = self.dm.add_movimentacao_caixa("suprimento", val, motivo, self.user.get("name", ""))
        if ok3: self.refresh()
        else: QMessageBox.warning(self, "Erro", msg)

    def _fechar(self):
        val, ok = QInputDialog.getDouble(self, "Fechar Caixa", "Valor contado no caixa:", 0, 0, 999999, 2)
        if not ok: return
        ok2, msg, resumo = self.dm.fechar_caixa(val, self.user.get("name", ""))
        if ok2:
            dif = resumo.get("diferenca", 0)
            dif_txt = "CORRETO" if abs(dif) < 0.01 else (f"SOBRA: {fmtR(dif)}" if dif > 0 else f"FALTA: {fmtR(abs(dif))}")
            dif_color = "#00C853" if abs(dif) < 0.01 else ("#FFD600" if dif > 0 else "#FF1744")
            detail = f"FECHAMENTO DE CAIXA\n\n"
            detail += f"Valor Inicial: {fmtR(resumo.get('valor_inicial',0))}\n"
            detail += f"Total Vendas: {fmtR(resumo.get('total_vendas',0))} ({resumo.get('n_vendas',0)} vendas)\n"
            detail += f"Sangrias: -{fmtR(resumo.get('total_sangria',0))}\n"
            detail += f"Suprimentos: +{fmtR(resumo.get('total_suprimento',0))}\n"
            detail += f"\nEsperado: {fmtR(resumo.get('esperado',0))}\n"
            detail += f"Contado: {fmtR(val)}\n"
            detail += f"Diferenca: {dif_txt}\n\n"
            # Por pagamento
            pp = resumo.get("por_pagamento", {})
            if pp:
                detail += "POR FORMA DE PAGAMENTO:\n"
                for k, v in pp.items(): detail += f"  {k}: {fmtR(v)}\n"
            QMessageBox.information(self, "Caixa Fechado", detail)
            self.refresh()
        else:
            QMessageBox.warning(self, "Erro", msg)


# ═════════════════════════════════════════════════════════════════════════════
#  IMPORT XML DIALOG (inclusao de estoque via XML NF-e)
# ═════════════════════════════════════════════════════════════════════════════
class ImportXMLDialog(ModernFormDialog):
    def __init__(self, parent, dm):
        super().__init__(parent, min_width=700); self.dm = dm; self.items = []
        self.add_title_bar("Importar Estoque via XML")
        lay = self.main_lay
        info = QLabel("Selecione o arquivo XML da nota fiscal de entrada do fornecedor.\nOs produtos serao adicionados ou atualizados no estoque.")
        info.setWordWrap(True); info.setObjectName("subtitle"); lay.addWidget(info)
        # Botao abrir XML
        bx = QPushButton("  Abrir Arquivo XML"); bx.clicked.connect(self._open_xml); lay.addWidget(bx)
        self.status_lbl = QLabel(""); lay.addWidget(self.status_lbl)
        # Tabela
        self.tbl = QTableWidget(0, 7)
        self.tbl.setHorizontalHeaderLabels(["Importar", "Codigo", "Nome", "NCM", "Qtd", "Preco Custo", "Acao"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.verticalHeader().setVisible(False)
        lay.addWidget(self.tbl, 1)
        # Botoes
        btns = QHBoxLayout(); btns.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Importar Selecionados"); sv.setObjectName("btn_success"); sv.clicked.connect(self._import)
        btns.addWidget(cn); btns.addWidget(sv); lay.addLayout(btns)

    def _open_xml(self):
        path, _ = QFileDialog.getOpenFileName(self, "Selecionar XML NF-e", "", "XML (*.xml)")
        if not path: return
        items, err = self.dm.import_xml_nfe(path)
        if err:
            AlertDialog.show_info(self, "Erro ao ler XML", err, icon="❌"); return
        if not items:
            AlertDialog.show_info(self, "Vazio", "Nenhum item encontrado no XML."); return
        self.items = items
        self.status_lbl.setText(f"  {len(items)} produto(s) encontrados no XML")
        self.status_lbl.setStyleSheet("color:#00C853;font-size:13px;font-weight:bold;")
        self.tbl.setRowCount(len(items))
        self._checks = []
        for r, it in enumerate(items):
            chk = QCheckBox(); chk.setChecked(True); self._checks.append(chk)
            self.tbl.setCellWidget(r, 0, chk)
            self.tbl.setItem(r, 1, QTableWidgetItem(it.get("code","")))
            self.tbl.setItem(r, 2, QTableWidgetItem(it.get("name","")))
            self.tbl.setItem(r, 3, QTableWidgetItem(it.get("ncm","")))
            qi = QTableWidgetItem(f"{it.get('quantity',0):.0f}"); qi.setTextAlignment(Qt.AlignCenter); self.tbl.setItem(r, 4, qi)
            self.tbl.setItem(r, 5, QTableWidgetItem(fmtR(it.get("unit_price", 0))))
            # Verificar se produto existe
            existing = next((p for p in self.dm.get_products()
                           if p.get("code","").upper() == it.get("code","").upper()
                           or (it.get("ean") and p.get("barcode") == it.get("ean"))), None)
            action = "Atualizar estoque" if existing else "Criar novo"
            ai = QTableWidgetItem(action)
            ai.setForeground(QColor("#00C853" if existing else "#FFD600"))
            self.tbl.setItem(r, 6, ai)
            self.tbl.setRowHeight(r, 36)

    def _import(self):
        if not self.items:
            AlertDialog.show_info(self, "Atencao", "Abra um arquivo XML primeiro."); return
        count_new = 0; count_upd = 0
        for r, it in enumerate(self.items):
            if not self._checks[r].isChecked(): continue
            existing = next((p for p in self.dm.get_products()
                           if p.get("code","").upper() == it.get("code","").upper()
                           or (it.get("ean") and it["ean"] != "SEM GTIN" and p.get("barcode") == it.get("ean"))), None)
            if existing:
                existing["stock"] = existing.get("stock", 0) + int(it.get("quantity", 0))
                existing["cost_price"] = it.get("unit_price", existing.get("cost_price", 0))
                if it.get("ncm"): existing["ncm"] = it["ncm"]
                self.dm.update_product(existing["id"], existing)
                count_upd += 1
            else:
                new_prod = {
                    "code": it.get("code",""),
                    "barcode": it.get("ean", ""),
                    "name": it.get("name",""),
                    "category": "Outros",
                    "brand": "",
                    "unit": it.get("unit","UN")[:3].upper(),
                    "cost_price": it.get("unit_price", 0),
                    "sale_price": round(it.get("unit_price", 0) * 1.30, 2),
                    "stock": int(it.get("quantity", 0)),
                    "min_stock": 2,
                    "description": "",
                    "ncm": it.get("ncm",""),
                    "cfop": "5102",
                    "added_at": datetime.now().isoformat()
                }
                self.dm.add_product(new_prod)
                count_new += 1
        AlertDialog.show_info(self, "Importacao Concluida",
            f"Importacao finalizada!\n\n{count_new} produto(s) criado(s)\n{count_upd} produto(s) atualizado(s)", icon="✅")
        self.accept()

# ═════════════════════════════════════════════════════════════════════════════
#  USERS TAB  (admin only)
# ═════════════════════════════════════════════════════════════════════════════
class UsersTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._ids=[]; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gerenciamento de Usuarios"))
        info=QFrame(); info.setObjectName("card"); info.setStyleSheet("QFrame#card{border-left:4px solid #1E88E5;}")
        il=QHBoxLayout(info); il.setContentsMargins(14,10,14,10)
        it=QLabel("<b>Controle de Acesso</b><br>"
            "  <b>Administrador</b>: acesso completo (todas as abas + configuracoes)<br>"
            "  <b>Operador</b>: somente Produtos, Nova Venda, Historico de Vendas e Clientes")
        it.setWordWrap(True); it.setObjectName("subtitle"); il.addWidget(it); lay.addWidget(info)
        tb=QHBoxLayout(); tb.addStretch()
        ba=QPushButton("  Novo Usuario"); ba.clicked.connect(self._add); tb.addWidget(ba)
        be=QPushButton("  Editar"); be.setObjectName("btn_secondary"); be.clicked.connect(self._edit); tb.addWidget(be)
        bt=QPushButton("  Ativar/Desativar"); bt.setObjectName("btn_warning"); bt.clicked.connect(self._toggle); tb.addWidget(bt)
        bd=QPushButton("  Excluir"); bd.setObjectName("btn_danger"); bd.clicked.connect(self._delete); tb.addWidget(bd)
        lay.addLayout(tb)
        self.tbl=QTableWidget(0,5); self.tbl.setHorizontalHeaderLabels(["Nome","Login","Perfil","Status","ID"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        for c,w in enumerate([-1,130,120,100,220]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt=QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
    def refresh(self):
        users=self.dm.get_users(); self.tbl.setRowCount(len(users)); self._ids=[]
        for r,u in enumerate(users):
            self._ids.append(u["id"])
            self.tbl.setItem(r,0,QTableWidgetItem(u.get("name","")))
            self.tbl.setItem(r,1,QTableWidgetItem(u.get("username","")))
            rl="  Admin" if u.get("role")=="admin" else "  Operador"
            ri=QTableWidgetItem(rl); ri.setForeground(QColor("#FFD600" if u.get("role")=="admin" else "#1E88E5")); self.tbl.setItem(r,2,ri)
            act=u.get("active",True); si=QTableWidgetItem("  Ativo" if act else "  Inativo")
            si.setForeground(QColor("#00C853" if act else "#FF1744")); self.tbl.setItem(r,3,si)
            self.tbl.setItem(r,4,QTableWidgetItem(u.get("id","")[:20]+"...")); self.tbl.setRowHeight(r,40)
        self.cnt.setText(f"{len(users)} usuario(s)")
    def _sid(self):
        r=self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _protected(self,uid): return uid in ("admin-fixed","func-fixed")
    def _add(self):
        dlg=UserDialog(self)
        if dlg.exec_()==QDialog.Accepted:
            ok,msg=self.dm.add_user(dlg.result_data)
            QMessageBox.information(self,"  Sucesso",msg) if ok else QMessageBox.warning(self,"Erro",msg)
            self.refresh()
    def _edit(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        u=next((x for x in self.dm.get_users() if x["id"]==uid),None)
        if not u: return
        dlg=UserDialog(self,u)
        if dlg.exec_()==QDialog.Accepted: dlg.result_data["id"]=uid; self.dm.update_user(uid,dlg.result_data); self.refresh()
    def _toggle(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        if self._protected(uid): QMessageBox.warning(self,"Protegido","Nao e possivel desativar este usuario."); return
        self.dm.toggle_user_active(uid); self.refresh()
    def _delete(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        if self._protected(uid): QMessageBox.warning(self,"Protegido","Nao e possivel excluir este usuario."); return
        u=next((x for x in self.dm.get_users() if x["id"]==uid),None)
        if u and QMessageBox.question(self,"Confirmar",f"Excluir usuario '{u['username']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_user(uid); self.refresh()


# ═════════════════════════════════════════════════════════════════════════════
#  INTEGRAÇÕES (MERCADO PAGO MOCK)
# ═════════════════════════════════════════════════════════════════════════════
class PaymentProcessDialog(ModernFormDialog):
    def __init__(self, parent, value, mode="credit"):
        super().__init__(parent, min_width=400)
        self.add_title_bar("Processando Pagamento...")
        self.container.setFixedSize(400, 480)
        self._value = value
        self._mode = mode
        
        lay = self.main_lay
        
        lbl_mp = QLabel("MERCADO PAGO  |  INTEGRACAO")
        lbl_mp.setAlignment(Qt.AlignCenter)
        lbl_mp.setStyleSheet("color:#0EA5E9;font-weight:bold;font-size:13px;letter-spacing:2px;")
        lay.addWidget(lbl_mp)
        
        self.lbl_status = QLabel("Aguardando conexao...")
        self.lbl_status.setAlignment(Qt.AlignCenter)
        self.lbl_status.setStyleSheet("color:#0F172A;font-size:18px;font-weight:bold;")
        lay.addWidget(self.lbl_status)
        lay.addSpacing(10)
        
        self.lbl_val = QLabel(fmtR(self._value))
        self.lbl_val.setAlignment(Qt.AlignCenter)
        self.lbl_val.setStyleSheet("color:#00C853;font-size:36px;font-weight:bold;")
        lay.addWidget(self.lbl_val)
        
        lay.addStretch()
        
        if self._mode == "pix":
            import qrcode
            try:
                qr = qrcode.QRCode(version=1, box_size=8, border=2)
                qr.add_data(f"00020101021126360014br.gov.bcb.pix0114{value}5204000053039865802BR5910MOCK MP6009MOCK CITY62070503***63041234")
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                import io
                buf=io.BytesIO()
                img.save(buf, format="PNG")
                qpix = QPixmap()
                qpix.loadFromData(buf.getvalue())
                qr_lbl = QLabel()
                qr_lbl.setAlignment(Qt.AlignCenter)
                qr_lbl.setPixmap(qpix.scaled(220, 220, Qt.KeepAspectRatio))
                lay.addWidget(qr_lbl)
                self.lbl_status.setText("Escaneie o QR Code no APP")
            except Exception as e:
                qr_lbl = QLabel("[ MOCK PIX: ERRO NO QRCODE ]\n" + str(e))
                qr_lbl.setStyleSheet("color:#888;")
                qr_lbl.setAlignment(Qt.AlignCenter)
                lay.addWidget(qr_lbl)
            
            self.btn_cancel = QPushButton("Cancelar PIX")
        else:
            self.lbl_icon = QLabel("📡")
            self.lbl_icon.setFont(QFont("Segoe UI Emoji", 48))
            self.lbl_icon.setAlignment(Qt.AlignCenter)
            lay.addWidget(self.lbl_icon)
            self.btn_cancel = QPushButton("Cancelar Transacao na Maquina")

        lay.addStretch()
        
        self.btn_cancel.setObjectName("btn_danger")
        self.btn_cancel.clicked.connect(self.reject)
        lay.addWidget(self.btn_cancel)
        
        self.timer = QTimer(self)
        self.timer.timeout.connect(self._fake_step)
        if self._mode != "pix":
            self._step = 0
            self.timer.start(1800)
            self.btn_cancel.setDisabled(False)
        else:
            self._step = -2
            self.timer.start(3500)

    def _fake_step(self):
        if self._mode == "pix":
            self._step += 1
            if self._step == 1:
                self.lbl_status.setText("Pagamento Recebido!")
                self.lbl_status.setStyleSheet("color:#00C853;font-size:22px;font-weight:bold;")
            elif self._step == 2:
                self.accept(); self.timer.stop()
            return

        self._step += 1
        if self._step == 1:
            self.lbl_status.setText("Enviando valor p/ Maquininha...")
            self.lbl_icon.setText("💳")
        elif self._step == 2:
            self.lbl_status.setText("Aguardando insercao de senha / nfc...")
            self.lbl_status.setStyleSheet("color:#FFD600;font-size:16px;font-weight:bold;")
        elif self._step == 3:
            self.lbl_status.setText("MERCADO PAGO: Processando...")
        elif self._step == 4:
            self.lbl_status.setText("Transacao Aprovada!")
            self.lbl_status.setStyleSheet("color:#00C853;font-size:22px;font-weight:bold;")
            self.lbl_icon.setText("✅")
            self.btn_cancel.setDisabled(True)
        elif self._step == 5:
            self.accept(); self.timer.stop()

class IntegrationsTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm
        self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(20)
        lay.addWidget(SectionTitle("  Central de Integracoes"))
        
        ff = QFrame(); ff.setStyleSheet("QFrame{background:#1A1A1A;border:1px solid #333;border-radius:12px;}")
        fl = QVBoxLayout(ff); fl.setContentsMargins(24,24,24,24); fl.setSpacing(15)
        
        hdr = QHBoxLayout()
        m_logo = QLabel(" MERCADO PAGO OAUTH  ")
        m_logo.setStyleSheet("color:#009EE3;font-weight:bold;font-size:18px;letter-spacing:1px;background:rgba(0,158,227,0.1);padding:6px;border-radius:6px;")
        hdr.addWidget(m_logo); hdr.addStretch()
        
        self.st_lbl = QLabel()
        self.st_lbl.setStyleSheet("font-weight:bold;font-size:14px;")
        hdr.addWidget(self.st_lbl); fl.addLayout(hdr)
        
        form = QFormLayout(); form.setSpacing(12)
        self.e_token = QLineEdit(); self.e_token.setEchoMode(QLineEdit.Password)
        self.e_token.setPlaceholderText("APP_USR-XXXXX-XXXX-...")
        self.e_device = QLineEdit()
        self.e_device.setPlaceholderText("PDV_XXXX_CAIXA_1")
        
        form.addRow("Access Token (Prod):", self.e_token)
        form.addRow("ID da Maquininha (Point):", self.e_device)
        fl.addLayout(form)
        
        bl = QHBoxLayout()
        lb = QLabel("Obtenha essas chaves no Painel de Desenvolvedores Mercado Pago.")
        lb.setStyleSheet("color:#888;font-size:11px;")
        bl.addWidget(lb); bl.addStretch()
        bsv = QPushButton("  Salvar Credenciais"); bsv.clicked.connect(self._save)
        bl.addWidget(bsv); fl.addLayout(bl)
        lay.addWidget(ff)
        
        self.dash = QFrame(); self.dash.setStyleSheet("QFrame{background:#121212;border:1px solid #009EE3;border-radius:12px;}")
        dl = QVBoxLayout(self.dash); dl.setContentsMargins(24,24,24,24)
        dl.addWidget(QLabel("Saldo em Conta Corrente", styleSheet="color:#888;"))
        dl.addWidget(QLabel("R$ 4.250,80", styleSheet="color:white;font-size:36px;font-weight:bold;"))
        dl.addSpacing(10)
        dl.addWidget(QLabel("Ultimas Transacoes MP", styleSheet="color:#888;"))
        
        for tx in ["10:45 - PIX Recebido - R$ 125,00", "09:30 - Point (Cartao Credito) - R$ 89,90", "Ontem - Point (Cartao Debito) - R$ 35,00"]:
            rl = QLabel("🟢  " + tx)
            rl.setStyleSheet("color:#DDD;padding:4px;")
            dl.addWidget(rl)
            
        lay.addWidget(self.dash); lay.addStretch()

    def refresh(self):
        s = self.dm.get_settings(); tk = s.get("mp_access_token", ""); dv = s.get("mp_device_id", "")
        self.e_token.setText(tk); self.e_device.setText(dv)
        if tk and dv:
            self.st_lbl.setText(" STATUS:  CONECTADO")
            self.st_lbl.setStyleSheet("color:#00C853;font-weight:bold;")
            self.dash.show()
        else:
            self.st_lbl.setText(" STATUS:  PENDENTE CONFIG.")
            self.st_lbl.setStyleSheet("color:#FF1744;font-weight:bold;")
            self.dash.hide()

    def _save(self):
        tk = self.e_token.text().strip(); dv = self.e_device.text().strip()
        s = self.dm.get_settings(); s["mp_access_token"] = tk; s["mp_device_id"] = dv; self.dm.save_settings(s)
        QMessageBox.information(self, "Sucesso", "Credenciais do Mercado Pago conectadas ao sistema!")
        self.refresh()

# ═════════════════════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ═════════════════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self, dm, user):
        super().__init__(); self.dm=dm; self.user=user; self.is_admin=user.get("role")=="admin"
        self.faltas_mgr = FaltasManager()
        s=dm.get_settings(); self._theme=s.get("theme","dark"); self._bg=s.get("background_image",""); self._accent=s.get("accent_color","#FF6B35")
        self.setWindowTitle(f"  Sistema Moto Pecas & Mecanica  --  {user.get('name','')}")
        self.setMinimumSize(1200,750); self.resize(1400,860); self._build(); self._apply_theme()
    def _build(self):
        central=QWidget(); self.setCentralWidget(central)
        ml=QHBoxLayout(central); ml.setContentsMargins(0,0,0,0); ml.setSpacing(0)
        self.sidebar=QFrame(); self.sidebar.setObjectName("sidebar"); self.sidebar.setFixedWidth(228)
        sl=QVBoxLayout(self.sidebar); sl.setContentsMargins(12,16,12,12); sl.setSpacing(4)
        lf=QFrame(); ll=QVBoxLayout(lf); ll.setContentsMargins(8,8,8,16)
        li=QLabel("  "); li.setFont(QFont("Segoe UI Emoji",28)); li.setAlignment(Qt.AlignCenter); ll.addWidget(li)
        lt=QLabel("MOTO PECAS"); lt.setAlignment(Qt.AlignCenter); lt.setStyleSheet("font-size:14px;font-weight:bold;letter-spacing:2px;"); ll.addWidget(lt)
        rl=QLabel("  Admin" if self.is_admin else "  Operador"); rl.setAlignment(Qt.AlignCenter)
        rl.setStyleSheet(f"font-size:11px;color:{'#FFD600' if self.is_admin else '#1E88E5'};"); ll.addWidget(rl)
        sep=QFrame(); sep.setFrameShape(QFrame.HLine); ll.addWidget(sep); sl.addWidget(lf)
        
        QShortcut(QKeySequence("F11"), self, self.showFullScreen)
        QShortcut(QKeySequence("Esc"), self, self.showNormal)
        
        self.content=QStackedWidget()
        self.t_dash=DashboardTab(self.dm); self.t_prod=ProductsTab(self.dm, self.user)
        self.t_sale=SalesTab(self.dm); self.t_sale.sale_completed.connect(self._on_sale)
        self.t_hist=HistoryTab(self.dm, self.user); self.t_hist.data_changed.connect(self._on_sale)
        self.t_cust=CustomersTab(self.dm)
        self.t_rep=ReportsTab(self.dm); self.t_usr=UsersTab(self.dm)
        self.t_faltas=FaltasTab(self.faltas_mgr)
        self.t_int = IntegrationsTab(self.dm)
        self.t_trocas = TrocasTab(self.dm)
        # CaixaTab is no longer shown
        all_tabs=[
            ("  Dashboard",   self.t_dash,   True),
            ("  Nova Venda",   self.t_sale,   False),
            ("  Produtos",     self.t_prod,   False),
            ("  Historico",    self.t_hist,   False),
            ("  Clientes",     self.t_cust,   False),
            ("  Trocas",       self.t_trocas, False),
            ("  Faltas",       self.t_faltas, False),
            ("  Relatorios",   self.t_rep,    True),
            ("  Integracoes",  self.t_int,    True),
            ("  Usuarios",     self.t_usr,    True),
        ]
        self._nav=[]
        for label,widget,admin_only in all_tabs:
            idx=self.content.count(); self.content.addWidget(widget)
            if not admin_only or self.is_admin: self._nav.append((label,idx,widget))
        self._btns=[]
        for label,idx,_ in self._nav:
            btn=QPushButton(label); btn.setObjectName("nav_btn"); btn.setMinimumHeight(42)
            btn.clicked.connect(lambda _,i=idx: self._go(i)); sl.addWidget(btn); self._btns.append((btn,idx))
        sl.addStretch()
        sep2=QFrame(); sep2.setFrameShape(QFrame.HLine); sl.addWidget(sep2)
        ol=QLabel("OPCOES"); ol.setObjectName("metric_label"); ol.setContentsMargins(8,4,0,4); sl.addWidget(ol)
        if self.is_admin:
            bs=QPushButton("  Configuracoes"); bs.setObjectName("nav_btn"); bs.setMinimumHeight(42); bs.clicked.connect(self._settings); sl.addWidget(bs)
            bfiscal=QPushButton("  Config. Fiscal / NFC-e"); bfiscal.setObjectName("nav_btn"); bfiscal.setMinimumHeight(42); bfiscal.clicked.connect(self._config_fiscal); sl.addWidget(bfiscal)
            bb=QPushButton("  Backup"); bb.setObjectName("nav_btn"); bb.setMinimumHeight(42); bb.clicked.connect(self._backup); sl.addWidget(bb)
        btc=QPushButton("  Trocar Conta"); btc.setObjectName("nav_btn"); btc.setMinimumHeight(42); btc.clicked.connect(self._switch_account); sl.addWidget(btc)
        bo=QPushButton("  Sair"); bo.setObjectName("nav_btn"); bo.setMinimumHeight(42); bo.clicked.connect(self._logout); sl.addWidget(bo)
        ml.addWidget(self.sidebar); ml.addWidget(self.content,1)
        sb=QStatusBar(); self.setStatusBar(sb)
        role_s="Administrador" if self.is_admin else "Operador"
        sb.showMessage(f"  {self.user.get('name','')} ({role_s})  |    {DATA_FILE}  |    {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        if self._nav: self._go(self._nav[0][1])
        if self.dm.get_settings().get("fullscreen", False): self.showFullScreen()
    def _go(self, idx):
        self.content.setCurrentIndex(idx)
        for btn,i in self._btns:
            btn.setObjectName("nav_btn_active" if i==idx else "nav_btn")
            btn.style().unpolish(btn); btn.style().polish(btn)
        # Refresh current tab
        w = self.content.currentWidget()
        if hasattr(w, "refresh"): w.refresh()
    def _on_sale(self):
        self.t_dash.refresh(); self.t_hist.refresh(); self.t_prod.refresh(); self.t_sale.refresh_customers()
        if hasattr(self, 't_caixa'): self.t_caixa.refresh()
    def _config_fiscal(self):
        dlg = EmpresaDialog(self, self.dm)
        if dlg.exec_() == QDialog.Accepted:
            QMessageBox.information(self, "  Salvo",
                "Dados fiscais salvos com sucesso!\n\n"
                "Configure o CSC junto a SEFAZ do seu estado para transmissao real da NFC-e.")

    def _settings(self):
        dlg=SettingsDialog(self,self.dm)
        if dlg.exec_()==QDialog.Accepted:
            s=self.dm.get_settings(); self._theme=s.get("theme","dark"); self._bg=s.get("background_image",""); self._accent=s.get("accent_color","#FF6B35")
            self._apply_theme(); QMessageBox.information(self,"  Configuracoes Salvas","Configuracoes aplicadas com sucesso!\nO icone sera aplicado na proxima abertura do sistema.")
    def _apply_theme(self):
        QApplication.instance().setStyleSheet(build_stylesheet(self._theme,self._accent))
        # Apply icon
        icon_path = self.dm.get_settings().get("app_icon","")
        if icon_path and os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
            QApplication.instance().setWindowIcon(QIcon(icon_path))
        if self._bg and os.path.exists(self._bg):
            pix=QPixmap(self._bg); pal=self.palette()
            scaled=pix.scaled(self.size(),Qt.KeepAspectRatioByExpanding,Qt.SmoothTransformation)
            pal.setBrush(QPalette.Window,QBrush(scaled)); self.setPalette(pal); self.setAutoFillBackground(True)
        else: self.setPalette(QApplication.instance().palette()); self.setAutoFillBackground(False)
    def resizeEvent(self,e):
        super().resizeEvent(e)
        if self._bg and os.path.exists(self._bg): self._apply_theme()
    def _backup(self):
        p,_=QFileDialog.getSaveFileName(self,"Backup",f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json","JSON (*.json)")
        if p:
            import shutil
            try: shutil.copy2(DATA_FILE,p); QMessageBox.information(self,"  Backup OK",f"Salvo em:\n{p}")
            except Exception as e: QMessageBox.critical(self,"Erro",str(e))
    def _switch_account(self):
        """Close current session and show login without closing the app."""
        reply = QMessageBox.question(
            self, "Trocar Conta",
            f"Deseja trocar de conta?\n\nUsuario atual: {self.user.get('name', '')}",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.close()
            win = LoginWindow()
            win.show()
            scr = QApplication.instance().primaryScreen().geometry()
            win.move((scr.width() - win.width()) // 2, (scr.height() - win.height()) // 2)

    def toggle_focus_mode(self, enabled):
        if enabled:
            self.sidebar.hide()
            self.showFullScreen()
        else:
            self.sidebar.show()
            self.showNormal()

    def _logout(self):
        if QMessageBox.question(self,"Sair","Deseja sair do sistema?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.close(); w=LoginWindow(); w.show()
            scr=QApplication.instance().primaryScreen().geometry(); w.move((scr.width()-w.width())//2,(scr.height()-w.height())//2)

# =============================================================================
#  SISTEMA DE ATIVACAO POR PIN (primeira execucao por maquina)
# =============================================================================
MACHINE_CONFIG_FILE = os.path.join(BASE_DIR, "machine_config.json")
_PIN_CODE = "8602"   # PIN do proprietario

def _is_activated():
    """Retorna True se este computador ja foi ativado com o PIN."""
    try:
        if os.path.exists(MACHINE_CONFIG_FILE):
            with open(MACHINE_CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("activated", False)
    except Exception:
        pass
    return False

def _activate_machine():
    """Salva ativacao permanente neste computador."""
    try:
        import platform
        with open(MACHINE_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "activated":    True,
                "activated_at": datetime.now().isoformat(),
                "machine":      platform.node(),
            }, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


# ═════════════════════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ═════════════════════════════════════════════════════════════════════════════
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__(); self.dm=DataManager()
        self._needs_pin = not _is_activated()
        _h = 650 if self._needs_pin else 560
        self.setWindowTitle("Login"); self.setFixedSize(420, _h)
        self.setWindowFlags(Qt.FramelessWindowHint); self._dp=None; self._mw=None
        QApplication.instance().setStyleSheet(build_stylesheet("dark","#FF6B35"))
        _icon_p = self.dm.get_settings().get("app_icon","")
        if _icon_p and os.path.exists(_icon_p):
            QApplication.instance().setWindowIcon(QIcon(_icon_p))
        lay=QVBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        bg=QFrame(); bg.setStyleSheet("QFrame{background:#121212;border-radius:14px;}")
        bl=QVBoxLayout(bg); bl.setContentsMargins(40,24,40,24); bl.setSpacing(0)
        tb=QHBoxLayout(); tb.addWidget(QLabel("  ",styleSheet="color:#555;font-size:11px;")); tb.addStretch()
        cl=QPushButton("x"); cl.setFixedSize(28,28); cl.setStyleSheet("background:transparent;color:#888;font-size:14px;border:none;"); cl.clicked.connect(QApplication.quit)
        tb.addWidget(cl); bl.addLayout(tb); bl.addSpacing(4)
        li=QLabel("  "); li.setFont(QFont("Segoe UI Emoji",44)); li.setAlignment(Qt.AlignCenter); bl.addWidget(li)
        lt=QLabel("MOTO PECAS"); lt.setAlignment(Qt.AlignCenter); lt.setStyleSheet("font-size:22px;font-weight:bold;color:#FF6B35;letter-spacing:4px;"); bl.addWidget(lt)
        ls=QLabel("Sistema de Gestao e Vendas"); ls.setAlignment(Qt.AlignCenter); ls.setStyleSheet("font-size:12px;color:#888;margin-bottom:10px;"); bl.addWidget(ls)
        # Banner de ativacao
        if self._needs_pin:
            ab = QLabel("  ATIVACAO DO SISTEMA NECESSARIA")
            ab.setAlignment(Qt.AlignCenter)
            ab.setStyleSheet("font-size:10px;font-weight:bold;color:#FFD600;letter-spacing:1px;"
                             "background:rgba(255,214,0,0.1);border:1px solid rgba(255,214,0,0.35);"
                             "border-radius:6px;padding:8px;margin-top:4px;margin-bottom:2px;")
            bl.addWidget(ab)
        bl.addSpacing(8)
        ff=QFrame(); ff.setStyleSheet("QFrame{background:#1A1A1A;border:1px solid #2C2C2C;border-radius:10px;}")
        fl=QVBoxLayout(ff); fl.setContentsMargins(20,18,20,18); fl.setSpacing(12)
        ist="QLineEdit{background:#121212;border:1px solid #333;border-radius:8px;color:white;padding:8px 14px;font-size:14px;}QLineEdit:focus{border:2px solid #FF6B35;}"
        lbs="font-size:11px;color:#888;font-weight:bold;letter-spacing:1px;"
        # Campo PIN
        if self._needs_pin:
            fl.addWidget(QLabel("  PIN DE ATIVACAO",styleSheet=lbs))
            self.pin_e=QLineEdit(); self.pin_e.setEchoMode(QLineEdit.Password)
            self.pin_e.setPlaceholderText("Pin de 4 digitos")
            self.pin_e.setMaxLength(4); self.pin_e.setMinimumHeight(42); self.pin_e.setStyleSheet(ist)
            fl.addWidget(self.pin_e)
        fl.addWidget(QLabel("  USUARIO",styleSheet=lbs))
        self.ue=QLineEdit(); self.ue.setPlaceholderText("Login de acesso"); self.ue.setMinimumHeight(42); self.ue.setStyleSheet(ist)
        fl.addWidget(self.ue)
        fl.addWidget(QLabel("  SENHA",styleSheet=lbs))
        self.pe=QLineEdit(); self.pe.setEchoMode(QLineEdit.Password); self.pe.setPlaceholderText("Senha"); self.pe.setMinimumHeight(42); self.pe.setStyleSheet(ist)
        self.pe.returnPressed.connect(self._login); fl.addWidget(self.pe)
        bl.addWidget(ff); bl.addSpacing(10)
        self.el=QLabel(""); self.el.setAlignment(Qt.AlignCenter); self.el.setStyleSheet("color:#FF1744;font-size:12px;"); bl.addWidget(self.el)
        bl.addSpacing(4)
        btn=QPushButton("ENTRAR  >"); btn.setMinimumHeight(48)
        btn.setStyleSheet("QPushButton{background:#FF6B35;color:white;border:none;border-radius:10px;font-size:15px;font-weight:bold;letter-spacing:2px;}QPushButton:hover{background:#FFAA80;}QPushButton:pressed{background:#CC4F1F;}")
        btn.clicked.connect(self._login); bl.addWidget(btn); bl.addStretch()
        vl=QLabel("v2.0  |  Funcionario: funcionario / 123moto")
        vl.setAlignment(Qt.AlignCenter); vl.setStyleSheet("font-size:10px;color:#444;margin-top:6px;"); bl.addWidget(vl)
        lay.addWidget(bg)

    def _login(self):
        u=self.ue.text().strip(); p=self.pe.text()
        if not u or not p: self.el.setText("  Preencha o login e a senha."); return
        if self._needs_pin:
            pin = self.pin_e.text().strip()
            if not pin: self.el.setText("  Informe o PIN de ativacao."); return
            if pin != _PIN_CODE:
                self.el.setText("  PIN incorreto! Contate o proprietario.")
                self.pin_e.clear(); self.pin_e.setFocus(); return
        user=self.dm.authenticate(u,p)
        if user:
            if self._needs_pin: _activate_machine()
            self.el.setText(""); self._mw=MainWindow(self.dm,user); self._mw.show()
            scr=QApplication.instance().primaryScreen().geometry()
            self._mw.move((scr.width()-self._mw.width())//2,(scr.height()-self._mw.height())//2); self.close()
        else: self.el.setText("  Usuario ou senha incorretos!"); self.pe.clear(); self.pe.setFocus()

    def mousePressEvent(self,e):
        if e.button()==Qt.LeftButton: self._dp=e.globalPos()-self.pos(); e.accept()
    def mouseMoveEvent(self,e):
        if e.buttons()==Qt.LeftButton and self._dp: self.move(e.globalPos()-self._dp); e.accept()


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    app=QApplication(sys.argv); app.setApplicationName("MotoPecas Sistema"); app.setApplicationVersion("2.0.0")
    try: app.setAttribute(Qt.AA_EnableHighDpiScaling,True); app.setAttribute(Qt.AA_UseHighDpiPixmaps,True)
    except: pass
    win=LoginWindow(); win.show()
    scr=app.primaryScreen().geometry(); win.move((scr.width()-win.width())//2,(scr.height()-win.height())//2)
    sys.exit(app.exec_())

if __name__=="__main__":
    main()
